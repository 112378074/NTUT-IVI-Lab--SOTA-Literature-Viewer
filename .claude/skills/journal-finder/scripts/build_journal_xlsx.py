# -*- coding: utf-8 -*-
"""Generic journal-list workbook builder for the `journal-finder` skill.

Reads a JSON spec produced by the agent and emits a styled .xlsx with:
  - one combined sheet ("全部期刊")
  - one sheet per domain
  - a "說明" sheet with notes, verified-IF list, and source URLs.

Anti-fabrication rule: the agent must leave `if`/`quartile`/`issn` blank
("") for any value it did NOT verify from a cited source this run. This
script never invents values; it only renders what the JSON provides.

JSON schema
-----------
{
  "title":   "期刊清單_...",          # title shown on 說明 sheet
  "output":  "期刊清單_xxx.xlsx",      # output filename (relative to --outdir)
  "date":    "2026-06-01",
  "relevance_label": "對 X 關聯",      # header for the relevance column
  "domains": [ {"name": "...", "color": "E8F1FB"}, ... ],
  "journals": [
     {"domain","tier","name","abbrev","publisher",
      "if","quartile","issn","scope","relevance"}, ...
  ],
  "verified_if": ["Pattern Recognition 7.6 — <url>", ...],   # optional
  "sources":     [["label","url"], ...],                      # optional
  "notes":       [["key","value"], ...]                       # optional extra notes
}

Usage
-----
  python build_journal_xlsx.py --spec spec.json --outdir "C:/path/to/project"
"""
import argparse
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_COLORS = ["E8F1FB", "FCEFE3", "EAF6EC", "F3EAF8", "FBF6E3", "EAF0F6"]

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = [16, 11, 52, 14, 18, 15, 8, 13, 46, 14]
FIELDS = ["domain", "tier", "name", "abbrev", "publisher",
          "if", "quartile", "issn", "scope", "relevance"]


def build(spec, outdir):
    rel_label = spec.get("relevance_label", "關聯")
    headers = ["領域", "Tier", "期刊全名", "縮寫", "出版社",
               "IF (JCR)", "分區", "ISSN", "範疇 / Scope", rel_label]

    # domain -> fill color
    domains = spec.get("domains") or []
    color_of = {}
    for i, d in enumerate(domains):
        color_of[d["name"]] = d.get("color") or DEFAULT_COLORS[i % len(DEFAULT_COLORS)]

    journals = spec["journals"]

    def fill_for(domain):
        c = color_of.get(domain, DEFAULT_COLORS[0])
        return PatternFill("solid", fgColor=c)

    def style_sheet(ws, data):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        for r, j in enumerate(data, 2):
            fill = fill_for(j.get("domain", ""))
            for c, f in enumerate(FIELDS, 1):
                cell = ws.cell(row=r, column=c, value=j.get(f, ""))
                cell.border = BORDER
                if c in (6, 7, 8, 10):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 9)))
                cell.fill = fill
        for c, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
        ws.row_dimensions[1].height = 30

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "全部期刊"
    style_sheet(ws_all, journals)

    # per-domain sheets (sheet title <= 31 chars, no illegal chars)
    used = {"全部期刊"}
    for d in domains:
        name = d["name"]
        rows = [j for j in journals if j.get("domain") == name]
        if not rows:
            continue
        title = "".join(ch for ch in name if ch not in '[]:*?/\\')[:31] or "Domain"
        base, k = title, 2
        while title in used:
            title = (base[:28] + f"_{k}")
            k += 1
        used.add(title)
        style_sheet(wb.create_sheet(title), rows)

    # 說明 sheet
    ws_note = wb.create_sheet("說明")
    rows = [["說明 / Notes", ""], ["", ""],
            ["標題", spec.get("title", "")],
            ["整理日期", spec.get("date", "")],
            ["領域", " / ".join(d["name"] for d in domains)],
            ["IF 欄位", "僅填入本次經 web 查證的 JCR 影響因子；未查證者留白，請至 JCR/Scimago 確認。"],
            ["分區 欄位", "僅標示高把握之 Q 分區；留白＝未於本次確認（依 JCR 學科分類年度而異）。"],
            ["ISSN", "常用 (print) ISSN，供建檔參考；投稿前請於官方頁再確認。"]]
    for k, v in spec.get("notes", []):
        rows.append([k, v])
    vif = spec.get("verified_if", [])
    if vif:
        rows.append(["", ""])
        rows.append(["IF 已查證之期刊", ""])
        for item in vif:
            rows.append(["", item])
    src = spec.get("sources", [])
    if src:
        rows.append(["", ""])
        rows.append(["資料來源 (查證)", ""])
        for label, url in src:
            rows.append([label, url])

    for r, (a, b) in enumerate(rows, 1):
        ca = ws_note.cell(row=r, column=1, value=a)
        cb = ws_note.cell(row=r, column=2, value=b)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        if r == 1:
            ca.font = Font(bold=True, size=14, color="1F4E79")
        elif a and not b:
            ca.font = Font(bold=True, color="1F4E79")
        elif a:
            ca.font = Font(bold=True)
    ws_note.column_dimensions["A"].width = 42
    ws_note.column_dimensions["B"].width = 80

    out = os.path.join(outdir, spec.get("output", "journal_list.xlsx"))
    wb.save(out)
    return out, len(journals)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--spec", required=True, help="path to JSON spec")
    ap.add_argument("--outdir", default=".", help="output directory")
    args = ap.parse_args()
    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    out, n = build(spec, args.outdir)
    print(f"saved: {out}")
    print(f"journals: {n}")


if __name__ == "__main__":
    main()
