#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_fsod.py — Fill MS-COCO 1/5/10/30-shot novel AP from Meta-DETR + DeFRCN papers."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# Sources:
#   Meta-DETR (Zhang et al. TPAMI 2022) Table 3 — COCO few-shot, AP0.5:0.95 (novel AP)
#   DeFRCN (Qiao et al. ICCV 2021) Table 2 — COCO few-shot, novel AP without w/G
#
# Method paper-reported novel AP at each shot:
#   Meta-DETR:    1=7.5  5=15.4  10=19.0  30=22.2
#   DeFRCN:       1=9.3  5=16.1  10=18.5  30=22.6
#   TFA w/cos:    1=4.4* 5=7.7*  10=10.0  30=13.7  (* = reproduced in DeFRCN Table 2)
#   MPSR:         1=5.1* 5=8.7*  10=9.8   30=14.1
#   Meta R-CNN:                  10=8.7   30=12.4   (1/5-shot not reported)
#   FSCE:                        10=11.1  30=15.3   (1/5-shot not reported)
#   FSRW (Meta-YOLO):            10=5.6   30=9.1    (1/5-shot not reported)

# (sheet, method_prefix) -> (novel_AP, ap_text, source)
DATA = {
    # 1-shot
    ('MS-COCO 1-shot', 'Meta-DETR'):  (7.5, 'novel AP 7.5 (AP50 12.5, AP75 7.7) MS-COCO 1-shot', 'Meta-DETR TPAMI22 Table 3'),
    ('MS-COCO 1-shot', 'DeFRCN'):     (9.3, 'novel AP 9.3 MS-COCO 1-shot', 'DeFRCN ICCV21 Table 2 (without w/G)'),
    ('MS-COCO 1-shot', 'TFA w/cos'):  (4.4, 'novel AP 4.4 MS-COCO 1-shot (reproduced)', 'DeFRCN ICCV21 Table 2 (TFA reproduced)'),
    ('MS-COCO 1-shot', 'MPSR'):       (5.1, 'novel AP 5.1 MS-COCO 1-shot (reproduced)', 'DeFRCN ICCV21 Table 2 (MPSR reproduced)'),

    # 5-shot
    ('MS-COCO 5-shot', 'Meta-DETR'):  (15.4, 'novel AP 15.4 (AP50 25.0, AP75 15.8) MS-COCO 5-shot', 'Meta-DETR TPAMI22 Table 3'),
    ('MS-COCO 5-shot', 'DeFRCN'):     (16.1, 'novel AP 16.1 MS-COCO 5-shot', 'DeFRCN ICCV21 Table 2 (without w/G)'),
    ('MS-COCO 5-shot', 'TFA w/cos'):  (7.7,  'novel AP 7.7 MS-COCO 5-shot (reproduced)', 'DeFRCN ICCV21 Table 2 (TFA reproduced)'),
    ('MS-COCO 5-shot', 'MPSR'):       (8.7,  'novel AP 8.7 MS-COCO 5-shot (reproduced)', 'DeFRCN ICCV21 Table 2 (MPSR reproduced)'),

    # 10-shot
    ('MS-COCO 10-shot', 'Meta-DETR'): (19.0, 'novel AP 19.0 (AP50 30.5, AP75 19.7) MS-COCO 10-shot', 'Meta-DETR TPAMI22 Table 3'),
    ('MS-COCO 10-shot', 'DeFRCN'):    (18.5, 'novel AP 18.5 MS-COCO 10-shot', 'DeFRCN ICCV21 Table 2 (without w/G)'),
    ('MS-COCO 10-shot', 'TFA w/cos'): (10.0, 'novel AP 10.0 MS-COCO 10-shot', 'TFA ICML20 / Meta-DETR Table 3'),
    ('MS-COCO 10-shot', 'MPSR'):      (9.8,  'novel AP 9.8 MS-COCO 10-shot', 'MPSR ECCV20 / Meta-DETR Table 3'),
    ('MS-COCO 10-shot', 'Meta R-CNN'):(8.7,  'novel AP 8.7 MS-COCO 10-shot', 'Meta R-CNN ICCV19 / Meta-DETR Table 3'),
    ('MS-COCO 10-shot', 'FSCE'):      (11.1, 'novel AP 11.1 MS-COCO 10-shot', 'FSCE CVPR21 / Meta-DETR Table 3'),
    ('MS-COCO 10-shot', 'FSRW'):      (5.6,  'novel AP 5.6 MS-COCO 10-shot (Meta-YOLO)', 'FSRW ICCV19 / Meta-DETR Table 3'),

    # 30-shot
    ('MS-COCO 30-shot', 'Meta-DETR'): (22.2, 'novel AP 22.2 (AP50 35.0, AP75 22.8) MS-COCO 30-shot', 'Meta-DETR TPAMI22 Table 3'),
    ('MS-COCO 30-shot', 'DeFRCN'):    (22.6, 'novel AP 22.6 MS-COCO 30-shot', 'DeFRCN ICCV21 Table 2 (without w/G)'),
    ('MS-COCO 30-shot', 'TFA w/cos'): (13.7, 'novel AP 13.7 MS-COCO 30-shot', 'TFA ICML20 / Meta-DETR Table 3'),
    ('MS-COCO 30-shot', 'MPSR'):      (14.1, 'novel AP 14.1 MS-COCO 30-shot', 'MPSR ECCV20 / Meta-DETR Table 3'),
    ('MS-COCO 30-shot', 'Meta R-CNN'):(12.4, 'novel AP 12.4 MS-COCO 30-shot', 'Meta R-CNN ICCV19 / Meta-DETR Table 3'),
    ('MS-COCO 30-shot', 'FSCE'):      (15.3, 'novel AP 15.3 MS-COCO 30-shot', 'FSCE CVPR21 / Meta-DETR Table 3'),
    ('MS-COCO 30-shot', 'FSRW'):      (9.1,  'novel AP 9.1 MS-COCO 30-shot (Meta-YOLO)', 'FSRW ICCV19 / Meta-DETR Table 3'),

    # COCO 2017 FSOD = generic 10-shot
    ('COCO 2017 FSOD', 'Meta-DETR'):  (19.0, 'novel AP 19.0 MS-COCO 10-shot (general FSOD protocol)', 'Meta-DETR TPAMI22 Table 3'),
    ('COCO 2017 FSOD', 'DeFRCN'):     (18.5, 'novel AP 18.5 MS-COCO 10-shot', 'DeFRCN ICCV21 Table 2'),
    ('COCO 2017 FSOD', 'TFA w/cos'):  (10.0, 'novel AP 10.0 MS-COCO 10-shot', 'TFA ICML20'),
    ('COCO 2017 FSOD', 'MPSR'):       (9.8,  'novel AP 9.8 MS-COCO 10-shot', 'MPSR ECCV20'),
    ('COCO 2017 FSOD', 'Meta R-CNN'): (8.7,  'novel AP 8.7 MS-COCO 10-shot', 'Meta R-CNN ICCV19'),
    ('COCO 2017 FSOD', 'FSCE'):       (11.1, 'novel AP 11.1 MS-COCO 10-shot', 'FSCE CVPR21'),
    ('COCO 2017 FSOD', 'FSRW'):       (5.6,  'novel AP 5.6 MS-COCO 10-shot (Meta-YOLO)', 'FSRW ICCV19'),
}

# Methods to flag (couldn't verify or have broken refs)
FLAGS = {
    'DETReg': 'DETReg (CVPR22 Bar et al.) — pretraining-only paper; main FSOD results in supplementary or follow-up code. Paper-reported COCO novel AP not directly extracted; needs verification from DETReg paper or codebase',
    'FM-FSOD': 'FM-FSOD arxiv ID 2402.10433 in sheet links to a different paper (Str2Str protein folding). The correct FM-FSOD paper reference is missing — needs identification before metric verification',
    'FII-DETR': 'FII-DETR is a journal publication (Information Fusion 2024/2025, sciencedirect paywalled). No public arXiv preprint found for metric verification',
    'hANMCL': 'hANMCL arxiv ID 2404.00700 in sheet links to a different paper (Kleinian groups math). The correct hANMCL paper reference is missing',
}

wb = load_workbook(XLSX)
filled = 0
flagged = 0
SHEETS = ['MS-COCO 1-shot','MS-COCO 5-shot','MS-COCO 10-shot','MS-COCO 30-shot','COCO 2017 FSOD']

for sn in SHEETS:
    s = wb[sn]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not (len(row) > 1 and row[1]): continue
        method_full = str(row[1]).strip()
        # Extract base method name (remove "(N-shot)" suffix)
        base = method_full.split('(')[0].strip()
        cur_note = str(row[10] or '') if len(row) > 10 else ''
        if '✅' in cur_note: continue
        if '⚠' in cur_note and not row[6]: pass  # allow re-fill if mAP empty

        # Try verified data first
        key = (sn, base)
        if key in DATA:
            ap, ap_text, source = DATA[key]
            s.cell(row=i, column=7).value = ap
            s.cell(row=i, column=8).value = ap_text
            s.cell(row=i, column=11).value = f'✅ verified ({source}): {sn} novel AP={ap}'
            filled += 1
            print(f'FILL [{sn}] r{i} {method_full}: {ap}')
            continue

        # Flag broken/missing references
        if base in FLAGS and not row[6]:
            s.cell(row=i, column=11).value = '⚠️ ' + FLAGS[base]
            flagged += 1

wb.save(XLSX)
print(f'\nTotal: filled={filled}, flagged={flagged}')
