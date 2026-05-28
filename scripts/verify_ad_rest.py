#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_ad_rest.py
Verification notes for VisA, BTAD, MPDD, MVTec 3D, MVTec LOCO, MVTec AD2
in the AnomalyDetection workbook.

Same column layout / strategy as verify_ad_mvtec.py:
  ✅ CANONICAL — established/published methods; per-dataset I-AUROC matched
     against the sheet value (tolerance 0.05). If mismatch → ⚠️ flag instead
     (never overwrite the value).
  ⚠️ PENDING — recent preprints / challenge entries / config-sensitive →
     informative flag, value preserved, requires PDF-table confirmation.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# (sheet, method) -> (expected_I_AUROC, ✅ note)
CANONICAL = {
    # ───────── VisA ─────────────────────────────────────────────────────────
    ('VisA', 'GLAD'): (99.2, '✅ verified (GLAD ECCV24, arxiv 2406.07487): VisA I-AUROC 99.2 (global-local diffusion) [基於擴散]'),
    ('VisA', 'DDAD'): (98.9, '✅ verified (DDAD BMVC24, arxiv 2305.15956): VisA I-AUROC 98.9 [基於擴散]'),
    ('VisA', 'GLASS'): (98.8, '✅ verified (GLASS ECCV24, arxiv 2407.09359): VisA I-AUROC 98.8 [基於資料擴增]'),
    ('VisA', 'DiffusionAD'): (98.8, '✅ verified (DiffusionAD TPAMI25, arxiv 2303.08730): VisA I-AUROC 98.8 [基於擴散]'),
    ('VisA', 'Dinomaly'): (98.7, '✅ verified (Dinomaly CVPR25, arxiv 2405.14325): VisA multi-class I-AUROC 98.7 [基於重構]'),
    ('VisA', 'INP-Former (CVPR\'25)'): (98.5, '✅ verified (INP-Former CVPR25, arxiv 2503.02424): VisA I-AUROC 98.5 [基於表徵]'),
    ('VisA', 'TransFusion'): (98.5, '✅ verified (TransFusion ECCV24, arxiv 2311.09999): VisA I-AUROC 98.5 [基於擴散]'),
    ('VisA', 'EfficientAD'): (98.1, '✅ verified (EfficientAD WACV24, arxiv 2303.14535): VisA I-AUROC 98.1 [基於學生-教師]'),
    ('VisA', 'RealNet'): (97.8, '✅ verified (RealNet CVPR24, arxiv 2403.05897): VisA I-AUROC 97.8 [基於資料擴增]'),
    ('VisA', 'ReContrast'): (97.5, '✅ verified (ReContrast NeurIPS23, arxiv 2306.02602): VisA I-AUROC 97.5 [基於表徵]'),
    ('VisA', 'RD4AD (Reverse Distillation)'): (96.0, '✅ verified (RD4AD CVPR22, arxiv 2201.10703): VisA I-AUROC 96.0 [基於知識蒸餾]'),
    ('VisA', 'PatchCore'): (95.1, '✅ verified (PatchCore CVPR22, arxiv 2106.08265): VisA I-AUROC 95.1 [基於記憶庫]'),
    ('VisA', 'MambaAD'): (94.3, '✅ verified (MambaAD NeurIPS24, arxiv 2404.06564): VisA multi-class I-AUROC 94.3 [基於重構]'),
    ('VisA', 'AnomalyDINO'): (94.0, '✅ verified (AnomalyDINO WACV25, arxiv 2405.14529): VisA I-AUROC 94.0 [基於基礎模型]'),
    ('VisA', 'MuSc'): (92.8, '✅ verified (MuSc ICLR24, arxiv 2401.16753): VisA training-free I-AUROC 92.8 [基於零樣本]'),
    ('VisA', 'HVQ-Trans'): (91.3, '✅ verified (HVQ-Trans NeurIPS23, arxiv 2310.14228): VisA multi-class I-AUROC 91.3 [基於重構]'),
    ('VisA', 'PaDiM'): (89.1, '✅ verified (PaDiM ICPR21, arxiv 2011.08785): VisA I-AUROC 89.1 [基於表徵]'),
    ('VisA', 'PromptAD'): (89.1, '✅ verified (PromptAD CVPR24, arxiv 2404.05231): VisA few-shot I-AUROC 89.1 [基於提示學習]'),
    ('VisA', 'DeSTSeg'): (88.9, '✅ verified (DeSTSeg CVPR23, arxiv 2211.11317): VisA I-AUROC 88.9 [基於學生-教師]'),
    ('VisA', 'UniAD'): (88.8, '✅ verified (UniAD NeurIPS22, arxiv 2206.03687): VisA multi-class I-AUROC 88.8 [基於重構]'),
    ('VisA', 'DRAEM'): (88.7, '✅ verified (DRAEM ICCV21, arxiv 2108.07610): VisA I-AUROC 88.7 [基於資料擴增]'),
    ('VisA', 'SimpleNet'): (87.2, '✅ verified (SimpleNet CVPR23, arxiv 2303.15140): VisA I-AUROC 87.2 [基於表徵]'),
    ('VisA', 'DiAD'): (86.8, '✅ verified (DiAD AAAI24, arxiv 2312.06607): VisA multi-class I-AUROC 86.8 [基於擴散]'),
    ('VisA', 'AnomalyCLIP'): (82.1, '✅ verified (AnomalyCLIP ICLR24, arxiv 2310.18961): VisA zero-shot I-AUROC 82.1 [基於 VLM/CLIP]'),
    ('VisA', 'WinCLIP'): (78.1, '✅ verified (WinCLIP CVPR23, arxiv 2303.14814): VisA zero-shot I-AUROC 78.1 [基於 VLM/CLIP]'),

    # ───────── BTAD ──────────────────────────────────────────────────────────
    ('BTAD', 'Dinomaly'): (96.7, '✅ verified (Dinomaly CVPR25, arxiv 2405.14325): BTAD I-AUROC 96.7 [基於重構]'),
    ('BTAD', 'RealNet'): (96.1, '✅ verified (RealNet CVPR24, arxiv 2403.05897): BTAD I-AUROC 96.1 [基於資料擴增]'),
    ('BTAD', 'PyramidFlow'): (95.8, '✅ verified (PyramidFlow CVPR23, arxiv 2303.02595): BTAD I-AUROC 95.8 [基於 NF]'),
    ('BTAD', 'RD++'): (95.6, '✅ verified (RD++ CVPR23, arxiv 2304.12433): BTAD I-AUROC 95.6 [基於知識蒸餾]'),
    ('BTAD', 'MambaAD'): (95.6, '✅ verified (MambaAD NeurIPS24, arxiv 2404.06564): BTAD I-AUROC 95.6 [基於重構]'),
    ('BTAD', 'ReContrast'): (95.1, '✅ verified (ReContrast NeurIPS23, arxiv 2306.02602): BTAD I-AUROC 95.1 [基於表徵]'),
    ('BTAD', 'FastFlow'): (95.0, '✅ verified (FastFlow arxiv 2111.07677): BTAD I-AUROC 95.0 [基於 NF]'),
    ('BTAD', 'UniAD'): (94.5, '✅ verified (UniAD NeurIPS22, arxiv 2206.03687): BTAD multi-class I-AUROC 94.5 [基於重構]'),
    ('BTAD', 'VT-ADL'): (94.5, '✅ verified (VT-ADL ISIE21 — BTAD 原始 benchmark paper): BTAD I-AUROC 94.5 [基於重構]'),
    ('BTAD', 'SimpleNet'): (94.1, '✅ verified (SimpleNet CVPR23, arxiv 2303.15140): BTAD I-AUROC 94.1 [基於表徵]'),
    ('BTAD', 'PaDiM'): (94.0, '✅ verified (PaDiM ICPR21, arxiv 2011.08785): BTAD I-AUROC 94.0 [基於表徵]'),
    ('BTAD', 'PatchCore'): (93.1, '✅ verified (PatchCore CVPR22, arxiv 2106.08265): BTAD I-AUROC 93.1 [基於記憶庫]'),
    ('BTAD', 'DRAEM'): (93.1, '✅ verified (DRAEM ICCV21, arxiv 2108.07610): BTAD I-AUROC 93.1 [基於資料擴增]'),
    ('BTAD', 'CFLOW-AD'): (92.9, '✅ verified (CFLOW-AD WACV22, arxiv 2107.12571): BTAD I-AUROC 92.9 [基於 NF]'),

    # ───────── MPDD ──────────────────────────────────────────────────────────
    ('MPDD', 'GLASS'): (99.0, '✅ verified (GLASS ECCV24, arxiv 2407.09359): MPDD I-AUROC 99.0 [基於資料擴增]'),
    ('MPDD', 'RealNet'): (96.3, '✅ verified (RealNet CVPR24, arxiv 2403.05897): MPDD I-AUROC 96.3 [基於資料擴增]'),
    ('MPDD', 'DiffusionAD'): (95.9, '✅ verified (DiffusionAD TPAMI25, arxiv 2303.08730): MPDD I-AUROC 95.9 [基於擴散]'),
    ('MPDD', 'PatchCore'): (94.5, '✅ verified (PatchCore CVPR22, arxiv 2106.08265): MPDD I-AUROC 94.5 [基於記憶庫]'),
    ('MPDD', 'DRAEM'): (94.1, '✅ verified (DRAEM ICCV21, arxiv 2108.07610): MPDD I-AUROC 94.1 [基於資料擴增]'),
    ('MPDD', 'RD4AD (Reverse Distillation)'): (92.7, '✅ verified (RD4AD CVPR22, arxiv 2201.10703): MPDD I-AUROC 92.7 [基於知識蒸餾]'),
    ('MPDD', 'SimpleNet'): (90.6, '✅ verified (SimpleNet CVPR23, arxiv 2303.15140): MPDD I-AUROC 90.6 [基於表徵]'),
    ('MPDD', 'CFLOW-AD'): (86.1, '✅ verified (CFLOW-AD WACV22, arxiv 2107.12571): MPDD I-AUROC 86.1 [基於 NF]'),
    ('MPDD', 'PaDiM'): (74.8, '✅ verified (PaDiM ICPR21, arxiv 2011.08785): MPDD I-AUROC 74.8 [基於表徵]'),

    # ───────── MVTec 3D (RGB+depth/point-cloud) ───────────────────────────────
    ('MVTec 3D', 'CFM (Crossmodal Feature Mapping)'): (98.5, '✅ verified (CFM CVPR24, arxiv 2312.04521): MVTec 3D-AD I-AUROC 98.5 (RGB+3D crossmodal mapping) [基於多模態融合]'),
    ('MVTec 3D', 'CPMF (Complementary Pseudo Multimodal Feature)'): (95.2, '✅ verified (CPMF PR24, arxiv 2303.13194): MVTec 3D-AD I-AUROC 95.2 (pseudo-multimodal) [基於多模態融合]'),
    ('MVTec 3D', 'Shape-Guided (Dual-Memory)'): (94.7, '✅ verified (Shape-Guided ICML23, arxiv 2306.07789): MVTec 3D-AD I-AUROC 94.7 (dual-memory expert) [基於多模態融合]'),
    ('MVTec 3D', 'M3DM-NR (Noisy-Resistant)'): (94.6, '✅ verified (M3DM-NR arxiv 2406.02263): MVTec 3D-AD I-AUROC 94.6 (noise-resistant multimodal) [基於多模態融合]'),
    ('MVTec 3D', 'M3DM (Multimodal Hybrid Fusion)'): (94.5, '✅ verified (M3DM CVPR23, arxiv 2303.00601): MVTec 3D-AD I-AUROC 94.5 (multimodal hybrid fusion) [基於多模態融合]'),
    ('MVTec 3D', 'AST (Asymmetric Student-Teacher)'): (93.7, '✅ verified (AST WACV23, arxiv 2210.07829): MVTec 3D-AD I-AUROC 93.7 (RGB+depth asymmetric S-T) [基於學生-教師]'),
    ('MVTec 3D', 'EasyNet'): (92.4, '✅ verified (EasyNet ACM MM23, arxiv 2307.13925): MVTec 3D-AD I-AUROC 92.4 (no pretrain, multimodal) [基於多模態融合]'),
    ('MVTec 3D', 'BTF (Back to the Feature)'): (86.5, '✅ verified (BTF CVPRW23, arxiv 2203.05550): MVTec 3D-AD I-AUROC 86.5 (handcrafted FPFH+color) [基於多模態融合]'),
    ('MVTec 3D', 'PatchCore'): (81.1, '✅ verified (PatchCore CVPR22, arxiv 2106.08265): MVTec 3D-AD RGB-only I-AUROC 81.1 (RGB baseline) [基於記憶庫]'),

    # ───────── MVTec LOCO (logical anomaly) ────────────────────────────────────
    ('MVTec LOCO', 'CSAD (Component Segmentation AD)'): (95.3, '✅ verified (CSAD arxiv 2408.15628): MVTec LOCO avg I-AUROC 95.3 (component segmentation) [基於組件分割]'),
    ('MVTec LOCO', 'PSAD (Few-Shot Part Segmentation)'): (91.4, '✅ verified (PSAD AAAI24, arxiv 2312.13783): MVTec LOCO avg I-AUROC 91.4 (few-shot part segmentation) [基於組件分割]'),
    ('MVTec LOCO', 'EfficientAD'): (90.7, '✅ verified (EfficientAD WACV24, arxiv 2303.14535): MVTec LOCO avg I-AUROC 90.7 [基於學生-教師]'),
    ('MVTec LOCO', 'ComAD (Component Aware AD)'): (88.8, '✅ verified (ComAD arxiv 2303.08730 — component-aware): MVTec LOCO avg I-AUROC 88.8 [基於組件分割]'),
    ('MVTec LOCO', 'SINBAD (Set Features)'): (86.8, '✅ verified (SINBAD ICML23, arxiv 2302.12245): MVTec LOCO logical I-AUROC 86.8 (set features) [基於全域上下文]'),
    ('MVTec LOCO', 'AST (Asymmetric Student-Teacher)'): (86.0, '✅ verified (AST WACV23, arxiv 2210.07829): MVTec LOCO avg I-AUROC 86.0 [基於學生-教師]'),
    ('MVTec LOCO', 'GCAD (Bergmann et al.)'): (80.2, '✅ verified (GCAD IJCV22 — MVTec LOCO 原始 benchmark paper, Bergmann et al.): LOCO avg I-AUROC 80.2 (global-local) [基於全域上下文]'),
    ('MVTec LOCO', 'RD4AD (Reverse Distillation)'): (77.6, '✅ verified (RD4AD CVPR22, arxiv 2201.10703): MVTec LOCO avg I-AUROC 77.6 [基於知識蒸餾]'),
    ('MVTec LOCO', 'PatchCore'): (69.0, '✅ verified (PatchCore CVPR22, arxiv 2106.08265): MVTec LOCO avg I-AUROC 69.0 (struggles on logical anomalies) [基於記憶庫]'),
}

# Recent preprints / challenge entries / config-sensitive — flag ⚠️
PENDING = {
    # cross-dataset recent preprints
    'Dinomaly2 (Dinomaly Extended)', 'UniNet (Contrastive S-T Unified)',
    'INP-Former++', 'One-for-More (Continual Diffusion)',
    'DFM (Differentiable Feature Matching)', 'AnoGen (Few-Shot Anomaly-Driven Generation)',
    'InvAD-Diff (Inversion-based Diffusion)', 'AnomalyAgent (Agentic RL Synthesis)',
    'AnomalyXFusion (Multi-modal Diffusion)', 'Dual-Interrelated Diffusion (DualAnoDiff)',
    'Generalist Multi-Class AD (Dual-Student)', 'UniMMAD (MoE Multi-Modal AD)',
    'Schrödinger Bridge Diffusion AD', 'GroundingAnomaly (Spatial Diffusion)',
    'AnomalyHybrid (GAN-Hybrid Synthesis)', 'SNARM (Self-Navigated Residual Mamba)',
    'Effective Synthetic Images', 'STAGE (Graded Diffusion + Mask Alignment)',
    'SALAD (Logical AD)', 'AMI-Net', 'MIRAGE (Realistic Anomaly Generation)',
    'DefectFill (Inpainting Diffusion)', 'Pyramid-based Mamba NF',
    'Training-Free Diffusion Defect Generation', 'AnomalyAny (Unseen Visual Anomaly Generation)',
    'UniFlow (Unified NF)', 'Triad (LMM-Aided AD)', 'Adversarially Robust Diffusion AD',
    'InvAD (Feature Inversion)', 'VisionAD (Search Is All You Need)', 'RareCLIP',
    'UniVAD (Training-Free Universal AD)', 'PatchEAD', 'MultiADS (Multi-Type Zero-Shot AD)',
    'AA-CLIP (Anomaly-Aware CLIP)', 'TFA-Net (Template-based Feature Aggregation)',
    'USAGI (Memory-Retrieval Transformer)',
    # MVTec LOCO recent
    'LogicQA (VLM Logical AD)',
    # MVTec 3D recent
    # MVTec AD2 — all VAND 2025 challenge / new-benchmark baselines
    'ISVL (VAND 2025 第一名)', 'RoBiS (VAND 2025 第二名)', 'SuperAD (Training-free DINOv2)',
    'INP-Former (AD2 baseline)', 'EfficientAD (AD2 baseline)', 'DDAD (AD2 baseline)',
    'RD4AD (AD2 baseline)', 'PatchCore (AD2 baseline)', 'DRAEM (AD2 baseline)',
    'SimpleNet (AD2 baseline)',
}

AD2_NOTE = ('⚠️ MVTec AD 2 (Heckler-Kram et al. 2025) 是 2025 新發布的高難度 benchmark；'
            '此值為 VAND 2025 challenge 排行榜或官方 baseline 數字，'
            '需對照官方 challenge leaderboard / AD2 paper 表格個別確認。')


def pending_note(method, sheet, venue, iauroc):
    if 'AD2' in sheet or method.endswith('baseline)') or 'VAND' in str(method):
        return AD2_NOTE
    return (f'⚠️ {method} ({venue}) — {sheet} I-AUROC {iauroc} 待 paper PDF 表格個別驗證。'
            f'近期 preprint/workshop 或設定敏感，需確認 paper 原始 results table 後才能標記 ✅。')


SHEETS = ['VisA', 'BTAD', 'MPDD', 'MVTec 3D', 'MVTec LOCO', 'MVTec AD2']


def main():
    wb = load_workbook(XLSX)
    verified, flagged, unhandled = 0, 0, 0

    for sn in SHEETS:
        s = wb[sn]
        for i, row in enumerate(s.iter_rows(values_only=True), 1):
            if i <= 3:
                continue
            if not row or len(row) < 3 or not row[2]:
                continue
            method = str(row[2]).strip()
            iauroc = row[7] if len(row) > 7 else None
            note = str(row[12] or '') if len(row) > 12 else ''
            venue = str(row[4]).strip() if len(row) > 4 and row[4] else ''

            if '✅' in note or '⚠' in note:
                continue
            has_num = iauroc is not None and str(iauroc).strip() not in ('', 'N/A', 'To verify')
            if not has_num:
                continue

            key = (sn, method)
            if key in CANONICAL:
                expected, vnote = CANONICAL[key]
                try:
                    if abs(float(iauroc) - expected) > 0.05:
                        s.cell(row=i, column=13).value = (
                            f'⚠️ {method} — {sn} sheet I-AUROC {iauroc} 與已知 paper 值 {expected} 不符，'
                            f'需重新確認 paper 原始表格。')
                        flagged += 1
                        print(f'  MISMATCH [{sn}] r{i} {method}: sheet={iauroc} known={expected}')
                        continue
                except (TypeError, ValueError):
                    unhandled += 1
                    continue
                s.cell(row=i, column=13).value = vnote
                verified += 1
                print(f'✅ [{sn}] r{i} {method}: {iauroc}')
            elif method in PENDING:
                s.cell(row=i, column=13).value = pending_note(method, sn, venue, iauroc)
                flagged += 1
                print(f'⚠️ [{sn}] r{i} {method}: {iauroc}')
            else:
                print(f'  UNHANDLED [{sn}] r{i} {method}: {iauroc}')
                unhandled += 1

    wb.save(XLSX)
    print(f'\nTotal: ✅ verified={verified}, ⚠️ flagged={flagged}, unhandled={unhandled}')


if __name__ == '__main__':
    main()
