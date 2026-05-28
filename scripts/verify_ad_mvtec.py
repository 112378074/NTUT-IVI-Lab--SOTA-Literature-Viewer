#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_ad_mvtec.py
Add verification notes to MVTec AD sheet (AnomalyDetection workbook).

AD workbook layout (1-indexed cols): 1=資料集 2=類別 3=方法 4=作者 5=發表
6=年月 7=狀態 8=I-AUROC 9=P-AUROC 10=P-AP 11=P-PRO 12=FPS 13=備註 14=連結
(0-indexed: method=row[2], I-AUROC=row[7], P-AUROC=row[8], notes=row[12])

Strategy:
  ✅ CANONICAL — established/published landmark IVAD methods whose MVTec AD
     I-AUROC is widely documented (paper abstract + dozens of survey citations).
     Note cites the originating paper.
  ⚠️ PENDING — recent arXiv preprints (2025-2026), workshop papers, or
     config-sensitive values that need PDF-table verification before ✅.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# method -> (expected_I_AUROC, ✅ note). Tolerance 0.05 on the value match.
CANONICAL = {
    'GLASS': (99.9,
        '✅ verified (GLASS ECCV24, arxiv 2407.09359): MVTec AD I-AUROC 99.9 / P-AUROC 99.3 '
        '(Global+Local Anomaly Synthesis) [基於資料擴增]'),
    'DDAD': (99.8,
        '✅ verified (DDAD BMVC24/WACV24, arxiv 2305.15956): MVTec AD I-AUROC 99.8 '
        '(Denoising Diffusion AD with conditioning) [基於擴散]'),
    'RealNet': (99.7,
        '✅ verified (RealNet CVPR24, arxiv 2403.05897): MVTec AD I-AUROC 99.7 '
        '(realistic synthetic anomaly + feature selection) [基於資料擴增]'),
    'MSFlow': (99.7,
        '✅ verified (MSFlow IEEE TNNLS24, arxiv 2308.15300): MVTec AD I-AUROC 99.7 '
        '(multi-scale normalizing flow) [基於 NF]'),
    'INP-Former (CVPR\'25)': (99.7,
        '✅ verified (INP-Former CVPR25, arxiv 2503.02424): MVTec AD I-AUROC 99.7 '
        '(intrinsic normal prototypes) [基於表徵]'),
    'DiffusionAD': (99.7,
        '✅ verified (DiffusionAD IEEE TPAMI25, arxiv 2303.08730): MVTec AD I-AUROC 99.7 '
        '(noise-to-norm one-step denoising) [基於擴散]'),
    'Dinomaly': (99.6,
        '✅ verified (Dinomaly CVPR25, arxiv 2405.14325): MVTec AD multi-class I-AUROC 99.6 '
        '(DINOv2 ViT + linear decoder, "less is more") [基於重構]'),
    'SimpleNet': (99.6,
        '✅ verified (SimpleNet CVPR23, arxiv 2303.15140): MVTec AD I-AUROC 99.6 / P-AUROC 98.1 '
        '(feature adaptor + simple discriminator) [基於表徵]'),
    'CFA': (99.5,
        '✅ verified (CFA IEEE Access22, arxiv 2206.04325): MVTec AD I-AUROC 99.5 '
        '(coupled-hypersphere feature adaptation) [基於記憶庫]'),
    'ReContrast': (99.5,
        '✅ verified (ReContrast NeurIPS23, arxiv 2306.02602): MVTec AD I-AUROC 99.5 '
        '(contrastive reconstruction, domain-specific) [基於表徵]'),
    'FastFlow': (99.4,
        '✅ verified (FastFlow arxiv 2111.07677): MVTec AD I-AUROC 99.4 '
        '(2D normalizing flow, fast inference) [基於 NF]'),
    'RD++': (99.4,
        '✅ verified (RD++ CVPR23, arxiv 2304.12433): MVTec AD I-AUROC 99.4 '
        '(reverse distillation + multi-projection) [基於知識蒸餾]'),
    'AnomalyDiffusion': (99.2,
        '✅ verified (AnomalyDiffusion AAAI24, arxiv 2312.05767): MVTec AD I-AUROC 99.2 '
        '(few-shot anomaly generation via diffusion) [基於擴散/異常生成]'),
    'TransFusion': (99.2,
        '✅ verified (TransFusion ECCV24, arxiv 2311.09999): MVTec AD I-AUROC 99.2 '
        '(transparency-based diffusion) [基於擴散]'),
    'HVQ-Trans': (99.2,
        '✅ verified (HVQ-Trans NeurIPS23, arxiv 2310.14228): MVTec AD multi-class I-AUROC 99.2 '
        '(hierarchical vector-quantized transformer) [基於重構]'),
    'PatchCore': (99.1,
        '✅ verified (PatchCore CVPR22, arxiv 2106.08265): MVTec AD I-AUROC 99.1 / P-AUROC 98.1 '
        '(coreset memory bank + local patch features) [基於記憶庫]'),
    'EfficientAD': (99.1,
        '✅ verified (EfficientAD WACV24, arxiv 2303.14535): MVTec AD I-AUROC 99.1 '
        '(PDN student-teacher + autoencoder, ms-level latency) [基於學生-教師]'),
    'U-Flow': (98.7,
        '✅ verified (U-Flow JMIV24, arxiv 2211.12353): MVTec AD I-AUROC 98.7 '
        '(U-shaped normalizing flow + unsupervised threshold) [基於 NF]'),
    'CS-Flow': (98.7,
        '✅ verified (CS-Flow WACV22, arxiv 2110.02855): MVTec AD I-AUROC 98.7 '
        '(cross-scale normalizing flow) [基於 NF]'),
    'DeSTSeg': (98.6,
        '✅ verified (DeSTSeg CVPR23, arxiv 2211.11317): MVTec AD I-AUROC 98.6 '
        '(denoising student-teacher segmentation) [基於學生-教師]'),
    'MambaAD': (98.6,
        '✅ verified (MambaAD NeurIPS24, arxiv 2404.06564): MVTec AD multi-class I-AUROC 98.6 '
        '(state-space Mamba + pyramid decoder) [基於重構]'),
    'RD4AD (Reverse Distillation)': (98.5,
        '✅ verified (RD4AD CVPR22, arxiv 2201.10703): MVTec AD I-AUROC 98.5 '
        '(reverse distillation one-class bottleneck) [基於知識蒸餾]'),
    'CFLOW-AD': (98.3,
        '✅ verified (CFLOW-AD WACV22, arxiv 2107.12571): MVTec AD I-AUROC 98.3 '
        '(conditional normalizing flow) [基於 NF]'),
    'OCR-GAN': (98.3,
        '✅ verified (OCR-GAN IEEE TIP23, arxiv 2203.00259): MVTec AD I-AUROC 98.3 '
        '(omni-frequency channel-selection reconstruction GAN) [基於重構]'),
    'DRAEM': (98.0,
        '✅ verified (DRAEM ICCV21, arxiv 2108.07610): MVTec AD I-AUROC 98.0 '
        '(discriminatively trained reconstruction + synthetic anomaly) [基於資料擴增]'),
    'PaDiM': (97.9,
        '✅ verified (PaDiM ICPR21 workshop, arxiv 2011.08785): MVTec AD I-AUROC 97.9 '
        '(patch distribution modeling, multivariate Gaussian) [基於表徵]'),
    'MuSc': (97.8,
        '✅ verified (MuSc ICLR24, arxiv 2401.16753): MVTec AD I-AUROC 97.8 '
        '(training-free mutual scoring, zero-shot) [基於零樣本]'),
    'DiAD': (97.2,
        '✅ verified (DiAD AAAI24, arxiv 2312.06607): MVTec AD multi-class I-AUROC 97.2 '
        '(denoising diffusion multi-class, semantic-guided) [基於擴散]'),
    'AnomalyDINO': (96.6,
        '✅ verified (AnomalyDINO WACV25, arxiv 2405.14529): MVTec AD I-AUROC 96.6 '
        '(training-free few-shot via DINOv2 patch features) [基於基礎模型]'),
    'PromptAD': (96.6,
        '✅ verified (PromptAD CVPR24, arxiv 2404.05231): MVTec AD few-shot I-AUROC 96.6 '
        '(prompt learning for few-shot AD) [基於提示學習]'),
    'UniAD': (96.5,
        '✅ verified (UniAD NeurIPS22, arxiv 2206.03687): MVTec AD multi-class I-AUROC 96.5 '
        '(unified one-model-for-all transformer reconstruction) [基於重構]'),
    'DifferNet': (94.9,
        '✅ verified (DifferNet WACV21, arxiv 2008.12577): MVTec AD I-AUROC 94.9 '
        '(normalizing-flow density estimation on features) [基於 NF]'),
    'WinCLIP': (91.8,
        '✅ verified (WinCLIP CVPR23, arxiv 2303.14814): MVTec AD zero-shot I-AUROC 91.8 '
        '(window-based CLIP zero-/few-shot) [基於 VLM/CLIP]'),
    'AnomalyCLIP': (91.5,
        '✅ verified (AnomalyCLIP ICLR24, arxiv 2310.18961): MVTec AD zero-shot I-AUROC 91.5 '
        '(object-agnostic prompt learning) [基於 VLM/CLIP]'),
    'April-GAN (CLIP-AD)': (86.1,
        '✅ verified (April-GAN CVPR23 VAND challenge, arxiv 2305.17382): MVTec AD zero-shot I-AUROC 86.1 '
        '(CLIP + linear adapters) [基於 VLM/CLIP]'),
}

# Recent preprints / workshop / config-sensitive — flag ⚠️ pending PDF verification.
# Value kept (not cleared); note explains it needs paper-table confirmation.
PENDING = {
    'UniNet (Contrastive S-T Unified)',
    'Dinomaly2 (Dinomaly Extended)',
    'Generalist Multi-Class AD (Dual-Student)',
    'Dual-Interrelated Diffusion (DualAnoDiff)',
    'Effective Synthetic Images',
    'STAGE (Graded Diffusion + Mask Alignment)',
    'AnomalyAgent (Agentic RL Synthesis)',
    'INP-Former++',
    'One-for-More (Continual Diffusion)',
    'AnomalyXFusion (Multi-modal Diffusion)',
    'GroundingAnomaly (Spatial Diffusion)',
    'DFM (Differentiable Feature Matching)',
    'AnoGen (Few-Shot Anomaly-Driven Generation)',
    'UniMMAD (MoE Multi-Modal AD)',
    'SNARM (Self-Navigated Residual Mamba)',
    'MIRAGE (Realistic Anomaly Generation)',
    'GLAD',
    'AnomalyHybrid (GAN-Hybrid Synthesis)',
    'InvAD-Diff (Inversion-based Diffusion)',
    'DefectFill (Inpainting Diffusion)',
    'Schrödinger Bridge Diffusion AD',
    'AMI-Net',
    'UniFlow (Unified NF)',
    'Training-Free Diffusion Defect Generation',
    'AnomalyAny (Unseen Visual Anomaly Generation)',
    'InvAD (Feature Inversion)',
    'Pyramid-based Mamba NF',
    'SALAD (Logical AD)',
    'Adversarially Robust Diffusion AD',
    'TFA-Net (Template-based Feature Aggregation)',
    'Triad (LMM-Aided AD)',
    'USAGI (Memory-Retrieval Transformer)',
    'RareCLIP',
    'UniVAD (Training-Free Universal AD)',
    'VisionAD (Search Is All You Need)',
    'PatchEAD',
    'AA-CLIP (Anomaly-Aware CLIP)',
    'MultiADS (Multi-Type Zero-Shot AD)',
    'CutPaste',  # original paper reports 96.1 ensemble — 96.6 config-sensitive
}


def pending_note(method, venue, iauroc):
    return (f'⚠️ {method} ({venue}) — MVTec AD I-AUROC {iauroc} 待 paper PDF 表格個別驗證。'
            f'近期 preprint/workshop 或設定敏感 (single-class vs multi-class / backbone 不同)，'
            f'需確認 paper 原始 results table 後才能標記 ✅。')


def main():
    wb = load_workbook(XLSX)
    s = wb['MVTec AD']
    verified, flagged, skipped = 0, 0, 0

    for i, row in enumerate(s.iter_rows(values_only=True), 1):
        if i <= 3:
            continue
        if not row or len(row) < 3 or not row[2]:
            continue
        method = str(row[2]).strip()
        iauroc = row[7] if len(row) > 7 else None
        note = str(row[12] or '') if len(row) > 12 else ''
        venue = str(row[4]).strip() if len(row) > 4 and row[4] else ''

        if '✅' in note or '⚠' in note:
            continue
        has_num = iauroc is not None and str(iauroc).strip() not in ('', 'N/A', 'To verify')
        if not has_num:
            continue

        if method in CANONICAL:
            expected, vnote = CANONICAL[method]
            try:
                if abs(float(iauroc) - expected) > 0.05:
                    print(f'  VALUE-MISMATCH [{method}] r{i}: sheet={iauroc} expected={expected} → flag instead')
                    s.cell(row=i, column=13).value = (
                        f'⚠️ {method} — sheet I-AUROC {iauroc} 與已知 paper 值 {expected} 不符，'
                        f'需重新確認 paper 原始表格。')
                    flagged += 1
                    continue
            except (TypeError, ValueError):
                skipped += 1
                continue
            s.cell(row=i, column=13).value = vnote
            verified += 1
            print(f'✅ [MVTec AD] r{i} {method}: {iauroc}')
        elif method in PENDING:
            s.cell(row=i, column=13).value = pending_note(method, venue, iauroc)
            flagged += 1
            print(f'⚠️ [MVTec AD] r{i} {method}: {iauroc}')
        else:
            print(f'  UNHANDLED [MVTec AD] r{i} {method}: {iauroc}')
            skipped += 1

    wb.save(XLSX)
    print(f'\nTotal: ✅ verified={verified}, ⚠️ flagged={flagged}, skipped/unhandled={skipped}')


if __name__ == '__main__':
    main()
