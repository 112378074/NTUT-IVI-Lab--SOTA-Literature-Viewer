#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batchA.py — COCO/LVIS/ODinW prominent methods (per-paper verified)."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# Columns: 7=mAP, 8=AP-text, 11=備註
FILLS = [
    # ----- COCO 2017 (zero-shot or trained) -----
    ('COCO 2017', 22, 56.0, 'AP 56.0 zero-shot',
     '✅ 已驗證 (DINO-X arXiv24 abstract)：DINO-X Pro 56.0% AP on COCO zero-shot transfer [Open-vocab Foundation Model]'),
    ('COCO 2017', 23, 52.5, 'AP 52.5 zero-shot',
     '✅ 已驗證 (Grounding DINO ECCV24 paper)：52.5 AP on COCO zero-shot transfer [Open-set OD]'),
    ('COCO 2017', 27, 57.7, 'AP 57.7 (APE-B cite)',
     '✅ 已驗證 (Grounding DINO 1.5 Pro Table 1 row)：57.7 AP zero-shot COCO transfer (line APE(B) cite); see also 55.7 LVIS-minival [Open-set OD]'),
    ('COCO 2017', 32, 50.8, 'AP 50.8 test-dev',
     '✅ 已驗證 (DAMO-YOLO arXiv22 Table 1)：DAMO-YOLO-L 50.8% AP COCO test-dev, 42ms@T4 [Real-Time OD]'),
    ('COCO 2017', 37, 54.7, 'AP 54.7 test-dev',
     '✅ 已驗證 (PP-YOLOE+ arXiv22 Table 1)：PP-YOLOE+-X 54.7% AP COCO test-dev [Real-Time OD]'),
    # ----- COCO test-dev -----
    ('COCO test-dev', 16, 64.5, 'box AP 64.5 test-dev',
     '✅ 已驗證 (EVA-02-Det arXiv23 abstract)：EVA-02 64.5 box AP / 55.8 mask AP on COCO test-dev [Foundation Model OD]'),
    ('COCO test-dev', 17, 54.7, 'AP 54.7 test-dev',
     '✅ 已驗證 (PP-YOLOE+ arXiv22 Table 1)：PP-YOLOE+-X 54.7% AP COCO test-dev [Real-Time OD]'),
    # ----- LVIS v1.0 val -----
    ('LVIS v1.0 val', 2, 41.7, 'AP 41.7 LVIS',
     '✅ 已驗證 (Detic CVPR22 abstract)：41.7 mAP on standard LVIS benchmark [Open-vocab OD via ImageNet-21K]'),
    ('LVIS v1.0 val', 3, 55.7, 'AP 55.7 LVIS-minival',
     '✅ 已驗證 (Grounding DINO 1.5 Pro paper)：55.7 AP on LVIS-minival zero-shot [Open-set OD]'),
    ('LVIS v1.0 val', 7, 59.8, 'AP 59.8 LVIS-minival',
     '✅ 已驗證 (DINO-X arXiv24 abstract)：DINO-X Pro 59.8 AP on LVIS-minival, 52.4 AP on LVIS-val [Open-vocab OD]'),
    ('LVIS v1.0 val', 8, 35.4, 'AP 35.4 LVIS',
     '✅ 已驗證 (YOLO-World CVPR24 abstract)：35.4 AP on LVIS with 52 FPS on V100 [Open-vocab Real-Time OD]'),
    ('LVIS v1.0 val', 15, 65.2, 'box AP 65.2 LVIS val',
     '✅ 已驗證 (EVA-02-Det arXiv23 abstract)：65.2 box AP / 57.3 mask AP on LVIS val [Foundation Model OD]'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
