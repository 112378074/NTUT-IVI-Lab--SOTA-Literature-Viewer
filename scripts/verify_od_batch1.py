#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batch1.py — fill COCO AP numbers from YOLOv13 Table I, D-FINE Table 1,
and Grounding DINO Table 2 PDF verifications."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# OD columns: 1=subarea, 2=method, ... 7=mAP, 8=AP(text), 9=FPS, 10=params, 11=notes, 12=link, 13=github
# (sheet, row, mAP_value, AP_text, note)
FILLS = [
    # COCO 2017 val
    ('COCO 2017 val', 17, 54.8, 'AP 54.8 val',
     '✅ 已驗證 (YOLOv13 Table I, COCO val2017)：YOLOv13-X 54.8% AP, 14.7ms latency [Real-Time OD]'),
    ('COCO 2017 val', 19, 54.0, 'AP 54.0 val',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：D-FINE-L 54.0% AP, 8.07ms latency [Real-Time OD end-to-end]'),
    ('COCO 2017 val', 22, 54.3, 'AP 54.3 val',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RT-DETRv2-R101 54.3% AP, 13.66ms [Real-Time OD]'),
    ('COCO 2017 val', 23, 54.3, 'AP 54.3 val',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RT-DETR-R101 54.3% AP, 13.61ms [Real-Time OD]'),
    ('COCO 2017 val', 25, 52.8, 'AP 52.8 val',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RTMDet-X 52.8% AP, 21.59ms [Real-Time OD]'),
    ('COCO 2017 val', 27, 52.8, 'AP 52.8 val',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：YOLOv6-L 52.8% AP, 9.04ms [Real-Time OD]'),
    # COCO 2017 (combined)
    ('COCO 2017', 16, 54.8, 'AP 54.8 val2017',
     '✅ 已驗證 (YOLOv13 Table I, COCO val2017)：YOLOv13-X 54.8% AP [Real-Time OD]'),
    ('COCO 2017', 19, 54.0, 'AP 54.0 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：D-FINE-L 54.0% AP [Real-Time OD]'),
    ('COCO 2017', 21, 54.4, 'AP 54.4 val2017',
     '✅ 已驗證 (YOLOv13 Table I, COCO val2017)：YOLOv10-X 54.4% AP [Real-Time OD]'),
    ('COCO 2017', 25, 54.3, 'AP 54.3 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RT-DETRv2-R101 54.3% AP [Real-Time OD]'),
    ('COCO 2017', 26, 54.3, 'AP 54.3 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RT-DETR-R101 54.3% AP [Real-Time OD]'),
    ('COCO 2017', 29, 52.9, 'AP 52.9 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：YOLOv7-X 52.9% AP [Real-Time OD]'),
    ('COCO 2017', 31, 52.8, 'AP 52.8 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：RTMDet-X 52.8% AP [Real-Time OD]'),
    ('COCO 2017', 35, 52.8, 'AP 52.8 val2017',
     '✅ 已驗證 (D-FINE Table 1, COCO val2017)：YOLOv6-L 52.8% AP [Real-Time OD]'),
    # COCO test-dev
    ('COCO test-dev', 14, 52.5, 'AP 52.5 (zero-shot Swin-L)',
     '✅ 已驗證 (Grounding DINO Table 2, COCO zero-shot)：Grounding DINO-L 52.5% AP zero-shot [Open-Vocab OD]'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))

if __name__ == '__main__':
    main()
