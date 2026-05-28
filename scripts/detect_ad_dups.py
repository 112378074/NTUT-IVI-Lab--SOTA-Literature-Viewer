#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""detect_ad_dups.py — report-only scan for duplicate paper entries in AD workbook.

Duplicate pattern: same paper appears twice — once as curated short method name
(+ arXiv link), once as CVF auto-harvested full title (+ CVF link).
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from itertools import combinations
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'


def arxiv_id(link):
    m = re.search(r'arxiv\.org/abs/(\d+\.\d+)', str(link or ''))
    return m.group(1) if m else None


def cvf_slug(link):
    m = re.search(r'/html/([^/]+)_(?:CVPR|ICCV|WACV|ECCV)', str(link or ''))
    return m.group(1).lower() if m else None


def main():
    wb = load_workbook(XLSX, read_only=True)
    SKIP = {'說明 Notes', '總覽 All Papers'}
    total_pairs = 0
    for sn in wb.sheetnames:
        if sn in SKIP:
            continue
        s = wb[sn]
        entries = []
        for i, row in enumerate(s.iter_rows(values_only=True), 1):
            if i <= 3:
                continue
            if not row or len(row) < 3 or not row[2]:
                continue
            entries.append({
                'r': i,
                'method': str(row[2]).strip(),
                'authors': str(row[3] or '').strip(),
                'I': row[7] if len(row) > 7 else None,
                'P': row[8] if len(row) > 8 else None,
                'link': row[13] if len(row) > 13 else '',
                'note': str(row[12] or ''),
                'axid': arxiv_id(row[13] if len(row) > 13 else ''),
                'cvf': cvf_slug(row[13] if len(row) > 13 else ''),
            })

        dups = []
        for a, b in combinations(entries, 2):
            score = 0
            reasons = []
            try:
                if a['I'] is not None and b['I'] is not None and abs(float(a['I']) - float(b['I'])) < 0.001:
                    score += 1
                    reasons.append('I=')
            except (TypeError, ValueError):
                pass
            try:
                if a['P'] is not None and b['P'] is not None and abs(float(a['P']) - float(b['P'])) < 0.001:
                    score += 1
                    reasons.append('P=')
            except (TypeError, ValueError):
                pass
            if a['axid'] and b['axid'] and a['axid'] == b['axid']:
                score += 3
                reasons.append('SAME-ARXIV')
            ma, mb = a['method'], b['method']
            # method short-name mentioned in the other entry's note
            if len(ma) > 4 and ma.lower() in b['note'].lower():
                score += 2
                reasons.append('A-method-in-Bnote')
            if len(mb) > 4 and mb.lower() in a['note'].lower():
                score += 2
                reasons.append('B-method-in-Anote')
            # author surname matches the other's CVF link slug
            for x, y in [(a, b), (b, a)]:
                if y['cvf'] and x['authors']:
                    toks = x['authors'].split()
                    surn = (toks[1] if len(toks) > 1 else toks[0]).lower().strip(',') if toks else ''
                    if surn and len(surn) > 2 and surn in y['cvf']:
                        score += 1
                        reasons.append('author-cvf')
            if score >= 3:
                dups.append((a, b, reasons))

        if dups:
            total_pairs += len(dups)
            print('\n=== %s — %d candidate duplicate pair(s) ===' % (sn, len(dups)))
            for a, b, reasons in dups:
                print('  r%-3d "%s" (I=%s P=%s) link=%s' % (a['r'], a['method'][:40], a['I'], a['P'], str(a['link'])[:45]))
                print('  r%-3d "%s" (I=%s P=%s) link=%s' % (b['r'], b['method'][:40], b['I'], b['P'], str(b['link'])[:45]))
                print('       reasons: %s' % ', '.join(reasons))
                print()
    print('\nTOTAL candidate duplicate pairs: %d' % total_pairs)


if __name__ == '__main__':
    main()
