#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batchB.py — more COCO/LVIS/ODinW + classic OD methods."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

FILLS = [
    # ----- ViTDet (ViT-H, MAE) on COCO 2017 + val -----
    ('COCO 2017', 33, 61.3, 'AP 61.3 (plain ViT-H, MAE)',
     '✅ 已驗證 (ViTDet ECCV22 abstract)：61.3 box AP on COCO with plain ViT-H, MAE pre-training [Plain ViT OD]'),
    ('COCO 2017 val', 26, 61.3, 'AP 61.3 (ViT-H)',
     '✅ 已驗證 (ViTDet ECCV22 abstract)：61.3 box AP on COCO with plain ViT-H, MAE pre-training [Plain ViT OD]'),
    # ----- RT-DETR-R101 @ COCO test-dev -----
    ('COCO test-dev', 15, 54.3, 'AP 54.3 (R101)',
     '✅ 已驗證 (RT-DETR CVPR24 abstract)：RT-DETR-R101 achieves 54.3% AP on COCO, 74 FPS T4 [Real-Time DETR]'),
    # ----- DETReg @ COCO + val + PASCAL VOC -----
    ('COCO 2017', 38, 45.5, 'AP 45.5 (semi-sup, Table 1)',
     '✅ 已驗證 (DETReg CVPR22 Table 1)：DETReg 45.5 AP COCO val under semi-supervised fine-tuning [Self-supervised pre-train OD]'),
    ('COCO 2017 val', 28, 45.5, 'AP 45.5 (semi-sup, Table 1)',
     '✅ 已驗證 (DETReg CVPR22 Table 1)：DETReg 45.5 AP COCO val under semi-supervised fine-tuning [Self-supervised pre-train OD]'),
    # ----- DAB-Deformable-DETR (workbook entry says DAB-DETR but the paper headline is DAB-Deformable-DETR-R50 46.8) -----
    ('COCO 2017', 41, 46.8, 'AP 46.8 (DAB-Deformable-DETR-R50)',
     '✅ 已驗證 (DAB-DETR ICLR22 Table 2)：DAB-Deformable-DETR-R50 46.8% AP COCO val (50 epochs) [DETR variant]'),
    ('COCO 2017 val', 30, 46.8, 'AP 46.8 (DAB-Deformable-DETR-R50)',
     '✅ 已驗證 (DAB-DETR ICLR22 Table 2)：DAB-Deformable-DETR-R50 46.8% AP COCO val [DETR variant]'),
    # ----- DN-DETR (DN-Deformable-DETR-R50, 12 ep) -----
    ('COCO 2017', 40, 43.4, 'AP 43.4 (DN-Deformable-DETR-R50, 12ep)',
     '✅ 已驗證 (DN-DETR CVPR22 Table 1)：DN-Deformable-DETR-R50 43.4% AP COCO val (12 epochs) [DETR denoising training]'),
    ('COCO 2017 val', 29, 43.4, 'AP 43.4 (DN-Deformable-DETR-R50, 12ep)',
     '✅ 已驗證 (DN-DETR CVPR22 Table 1)：DN-Deformable-DETR-R50 43.4% AP COCO val [DETR denoising training]'),
    # ----- LVIS APr (rare-class AP) for open-vocab methods -----
    ('LVIS v1.0 val', 5, 28.2, 'AP 28.2 APr (LVIS rare-class)',
     '✅ 已驗證 (GLIP CVPR22 Table 3 / Anomaly-OV cite)：GLIP-L Swin-L 28.2 LVIS APr (zero-shot rare-class) [Grounded VLM]'),
    ('LVIS v1.0 val', 9, 32.3, 'AP 32.3 LVIS',
     '✅ 已驗證 (RegionCLIP CVPR22 Table)：RegionCLIP 32.3 LVIS AP (open-vocab) [Region-aware CLIP for OD]'),
    # ----- DAMO-YOLO-l @ PASCAL VOC etc — only fills COCO; skip if no PASCAL data -----
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
