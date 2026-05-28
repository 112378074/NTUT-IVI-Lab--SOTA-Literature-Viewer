#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dedup_ad.py — merge + delete confirmed duplicate paper entries in AD workbook.

For each duplicate (same paper, two rows): keep the curated short-name entry,
merge any metric the keep-row is missing (N/A) from the delete-row, append a
provenance note, then delete the full-title auto-harvested duplicate row.

Cols (1-indexed): 8=I-AUROC 9=P-AUROC 10=P-AP 11=P-PRO 12=FPS 13=備註 14=連結
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

METRIC_COLS = {'I-AUROC': 8, 'P-AUROC': 9, 'P-AP': 10, 'P-PRO': 11, 'FPS': 12}
NOTE_COL = 13

# Confirmed duplicates: (sheet, keep_row, delete_row, provenance_suffix)
DUPLICATES = [
    ('MVTec AD', 11, 91,
     '（P-AP 已併自 CVF 重複條目「Exploring Intrinsic Normal Prototypes…」；該重複條目已刪除）'),
    ('VisA', 13, 74,
     '（P-AP 已併自 CVF 重複條目「Exploring Intrinsic Normal Prototypes…」，該條目多類別設定報 I=98.9；重複條目已刪除）'),
]


def is_empty(v):
    return v is None or str(v).strip() in ('', 'N/A', 'None')


def main():
    wb = load_workbook(XLSX)
    # Process; delete rows last (and we delete bottom-most first per sheet).
    # Group deletions by sheet to delete safely.
    to_delete = {}  # sheet -> list of row indices

    for sheet, keep_r, del_r, prov in DUPLICATES:
        s = wb[sheet]
        merged = []
        for mname, col in METRIC_COLS.items():
            keep_val = s.cell(row=keep_r, column=col).value
            del_val = s.cell(row=del_r, column=col).value
            if is_empty(keep_val) and not is_empty(del_val):
                s.cell(row=keep_r, column=col).value = del_val
                merged.append(f'{mname}={del_val}')
        # append provenance to keep-row note
        note = str(s.cell(row=keep_r, column=NOTE_COL).value or '')
        if prov not in note:
            s.cell(row=keep_r, column=NOTE_COL).value = (note + ' ' + prov).strip()
        print(f'[{sheet}] kept r{keep_r}, merged: {merged or "(no missing metrics)"}')
        to_delete.setdefault(sheet, []).append(del_r)

    # delete duplicate rows (highest index first within each sheet)
    for sheet, rows in to_delete.items():
        s = wb[sheet]
        for r in sorted(rows, reverse=True):
            method = s.cell(row=r, column=3).value
            s.delete_rows(r, 1)
            print(f'[{sheet}] deleted r{r}: "{method}"')

    wb.save(XLSX)
    print('\nSaved.')


if __name__ == '__main__':
    main()
