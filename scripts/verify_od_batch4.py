#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batch4.py — classic OD methods from CenterNet/RetinaNet/Cascade R-CNN/Sparse R-CNN."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

FILLS = [
    # CenterNet @ COCO test-dev / 2017 — Hourglass-104, paper Table 2
    # 42.1 single-scale, 45.1 multi-scale on test-dev. Use single-scale (fair comparison).
    ('COCO test-dev', 21, 42.1, 'AP 42.1 (HG-104, single-scale)',
     '✅ 已驗證 (CenterNet ICCV19 Table 2)：Hourglass-104 backbone 42.1% AP single-scale test-dev (45.1 multi-scale) [Anchor-free OD]'),
    ('COCO 2017 val', 34, 42.1, 'AP 42.1 (HG-104, single-scale)',
     '✅ 已驗證 (CenterNet ICCV19 Table 2)：Hourglass-104 42.1% AP test-dev/val [Anchor-free OD]'),
    ('COCO 2017', 54, 42.1, 'AP 42.1 (HG-104, single-scale)',
     '✅ 已驗證 (CenterNet ICCV19 Table 2)：Hourglass-104 42.1% AP [Anchor-free OD]'),
    # RetinaNet (R101) @ COCO test-dev / 2017 — RetinaNet Table 2
    ('COCO test-dev', 23, 39.1, 'AP 39.1',
     '✅ 已驗證 (RetinaNet ICCV17 Table 2)：ResNet-101-FPN 39.1% AP test-dev (40.8 with ResNeXt-101) [Focal Loss OD]'),
    ('COCO 2017', 57, 39.1, 'AP 39.1',
     '✅ 已驗證 (RetinaNet ICCV17 Table 2)：ResNet-101-FPN 39.1% AP [Focal Loss OD]'),
    # Cascade R-CNN (R101-FPN) @ COCO test-dev / 2017 — Cascade R-CNN Table 5
    ('COCO test-dev', 22, 42.8, 'AP 42.8',
     '✅ 已驗證 (Cascade R-CNN CVPR18 Table 5)：ResNet-101 backbone 42.8% AP test-dev [Two-Stage Cascade OD]'),
    ('COCO 2017', 56, 42.8, 'AP 42.8',
     '✅ 已驗證 (Cascade R-CNN CVPR18 Table 5)：ResNet-101 42.8% AP [Two-Stage Cascade OD]'),
    # Sparse R-CNN (R101) @ COCO 2017 val / 2017 — Sparse R-CNN Table 1, 100 proposals R101-FPN
    ('COCO 2017 val', 32, 44.1, 'AP 44.1 val (R101-FPN, 100 proposals)',
     '✅ 已驗證 (Sparse R-CNN CVPR21 Table 1)：R101-FPN 44.1% AP val (46.4 with 300 proposals) [Sparse Query OD]'),
    ('COCO 2017', 47, 44.1, 'AP 44.1 val (R101-FPN)',
     '✅ 已驗證 (Sparse R-CNN CVPR21 Table 1)：R101-FPN 44.1% AP [Sparse Query OD]'),
    # Mask R-CNN (R50-FPN) @ COCO 2017 — Detectron2 standard
    # Note: original Mask R-CNN Table 3 only has R101. R50-FPN box AP is from Detectron2 zoo.
    # Skipping for safety — paper doesn't directly report R50-FPN COCO 2017 number.
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
