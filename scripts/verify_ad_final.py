#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_ad_final.py — finishing AD batch: 2 fills + 8 removes."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

FILLS = [
    # SNARM @ MVTec 3D (SNARM Table 1, T=10^5, multi-class)
    # Columns in Table 1: P-AP | PRO | P-AUROC | I-AUROC
    # MVTec-3D SNARM: 63.6 / 97.4 / 99.2 / 93.9
    ('MVTec 3D', 19, 93.9, 99.2, 63.6, 97.4,
     '✅ 已驗證 (SNARM arXiv25 Table 1, multi-class, T=10^5)：Self-Navigated Residual Mamba；I=93.9 P=99.2 AP=63.6 PRO=97.4 [基於 Mamba SSM]'),
    # Generalist Multi-Class AD @ MVTec LOCO (Table 1, multi-class)
    ('MVTec LOCO', 17, 86.1, None, None, None,
     '✅ 已驗證 (Generalist Multi-Class AD arXiv25 Table 1, multi-class)：Dual-Student ensemble；MVTec-LOCO I=86.1 (image-AUROC) [Student-Teacher 雙學生]'),
]

DELETES = [
    # PSAD @ MVTec AD — paper only tests MVTec LOCO AD
    ('MVTec AD', 91, 'PSAD AAAI24 only evaluates MVTec LOCO AD (logical anomaly), not MVTec AD'),
    # LogicQA @ MVTec AD — paper only tests MVTec LOCO AD + Semiconductor SEM
    ('MVTec AD', 92, 'LogicQA ACL25 only evaluates MVTec LOCO AD + Semiconductor SEM'),
    # RD++ @ 3 datasets — arxiv 2304.02216 is the AeBAD/MMR paper, not actual RD++
    ('VisA', 71, 'RD++ link arxiv:2304.02216 is the AeBAD/MMR paper (different); RD++ not verifiable from this link'),
    ('MVTec LOCO', 20, 'RD++ link is wrong paper (AeBAD/MMR)'),
    ('MVTec 3D', 18, 'RD++ link is wrong paper (AeBAD/MMR)'),
    # STAGE @ BTAD — synthesis paper, uses mIoU/Acc not AUROC ranking
    ('BTAD', 21, 'STAGE is anomaly synthesis paper; evaluates segmentation mIoU/Acc, not AD AUROC ranking'),
    # AnomalyXFusion @ LOCO — synthesis-focused paper; main metrics are IS/FID
    ('MVTec LOCO', 19, 'AnomalyXFusion is generation paper; main eval is IS/FID/IC-LPIPS for synthesis, not LOCO AD ranking'),
    # Adversarially Robust Diffusion AD @ BTAD — robustness paper without standard BTAD AUROC
    ('BTAD', 24, 'Adversarially Robust Diffusion AD paper does not provide standard BTAD AUROC in main results'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> i={i}, p={p}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason[:60]}...)')

    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
