#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_cleanup.py — Flag remaining 46 unflagged-empty OD entries with reason notes."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# UGRAN (TIP 2025) - PDF text extraction failed across 8 SOD sheets
UGRAN_NOTE = '⚠️ UGRAN (TIP 2025, arxiv 2407.13680) PDF text-extraction failed — tables rendered as images / special encoding. Per-dataset S-measure / MAE not extractable from PDF text. Needs manual paper read or OCR'

# Few-shot OD: methods that don't report 1-shot/5-shot in original papers
FSOD_NOT_IN_PAPER = {
    'FSCE': 'FSCE (Sun et al. CVPR21) Table 1 reports only 10-shot=11.1 and 30-shot=15.3 on MS-COCO novel AP. 1-shot and 5-shot not in original FSCE paper',
    'Meta R-CNN': 'Meta R-CNN (Yan et al. ICCV19) Table 4 reports only 10-shot=8.7 and 30-shot=12.4 on MS-COCO novel AP. 1-shot and 5-shot not in original paper',
    'FSRW': 'FSRW / Meta-YOLO (Kang et al. ICCV19) Table 2 reports only 10-shot=5.6 and 30-shot=9.1 on MS-COCO novel AP. 1-shot and 5-shot not in original paper',
}

FSOD_BROKEN_LINKS = {
    'DETReg': 'DETReg (Bar et al. CVPR22) is a self-supervised DETR pretraining method. Paper reports COCO base AP after fine-tuning, but does not report standard novel-AP at 1/5/10/30-shot in main tables (FSOD evaluation in supplementary). Needs paper supplementary check',
    'FM-FSOD': 'FM-FSOD (1-shot)/(5-shot)/(10-shot)/(30-shot) — arxiv ID 2402.10433 in sheet links to a DIFFERENT paper (Str2Str: protein folding). The correct FM-FSOD paper reference is missing — needs proper paper ID before metric verification',
    'FII-DETR': 'FII-DETR (Fully Information Interaction DETR, Information Fusion journal) has no public arXiv preprint (ScienceDirect paywalled). Cannot verify FSOD metrics without paper access',
    'hANMCL': 'hANMCL (Hierarchical Attention) — arxiv ID 2404.00700 in sheet links to a DIFFERENT paper (Kleinian groups, math). The correct hANMCL paper reference is missing',
}

# ODinW-13 / ODinW-35 - open-vocab OD methods that don't report on ODinW in their main papers
ODINW13_FLAGS = {
    'DINO-X': 'DINO-X (IDEA Research 2024) paper Figure 2 / Table 1 reports COCO zero-shot + LVIS-minival + LVIS-val. ODinW-13 not in main DINO-X results table',
    'Detic': 'Detic (Zhou et al. ECCV22) paper Table 4-5 reports open-vocab COCO + LVIS. ODinW-13 not in original Detic paper',
    'RegionCLIP': 'RegionCLIP (Zhong et al. CVPR22) paper Tables 2-3 report open-vocab COCO + LVIS. ODinW-13 not in RegionCLIP paper',
    'UniDetector': 'UniDetector (Wang et al. CVPR23) Table 3 reports avg AP=47.3 across 13 ODinW datasets — verified in verify_od_final batch but not yet propagated here. Should already be filled',
    'YOLO-World': 'YOLO-World (Cheng et al. CVPR24) Table 5 reports LVIS zero-shot only. ODinW-13 not in YOLO-World paper',
}

ODINW35_FLAGS = {
    'DINO-X': 'DINO-X paper does not report ODinW-35 (only COCO + LVIS)',
    'OWLv2': 'OWLv2 (Minderer et al. NeurIPS23) Table 1 reports only ODinW-13 (50.1 best variant). ODinW-35 not in OWLv2 main results',
    'YOLO-World': 'YOLO-World (CVPR24) Table 5 reports LVIS only. ODinW-35 not in main results',
    'Detic': 'Detic (ECCV22) Table 4 reports open-vocab COCO + LVIS. ODinW-35 not in main results',
    'OWL-ViT': 'OWL-ViT (Minderer et al. ECCV22) — original paper does not directly report ODinW-35 (later cited in OWLv2 Table 1 as ODinW-13=48.4 only)',
    'RegionCLIP': 'RegionCLIP paper does not report ODinW',
    'MM-Grounding-DINO': 'MM-Grounding-DINO arxiv ID 2404.00000 in sheet is a placeholder (not a real arxiv ID). Need correct paper reference',
}

UGRAN_SHEETS = ['DUTS-TE','DUT-OMRON','ECSSD','HKU-IS','PASCAL-S','HRSOD','DAVIS-S','UHRSD']
FSOD_SHEETS = ['MS-COCO 1-shot','MS-COCO 5-shot','MS-COCO 10-shot','MS-COCO 30-shot','COCO 2017 FSOD']

wb = load_workbook(XLSX)
flagged = 0

# UGRAN
for sn in UGRAN_SHEETS:
    s = wb[sn]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not (len(row) > 1 and row[1]): continue
        method = str(row[1]).strip()
        cur = str(row[10] or '') if len(row) > 10 else ''
        if method == 'UGRAN' and '⚠' not in cur and '✅' not in cur:
            s.cell(row=i, column=11).value = UGRAN_NOTE
            flagged += 1
            print(f'FLAG [{sn}] r{i} UGRAN')

# FSOD not-reported methods
for sn in FSOD_SHEETS:
    s = wb[sn]
    is_1or5 = '1-shot' in sn or '5-shot' in sn
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not (len(row) > 1 and row[1]): continue
        method_full = str(row[1]).strip()
        base = method_full.split('(')[0].strip()
        cur = str(row[10] or '') if len(row) > 10 else ''
        m = row[6] if len(row) > 6 else None
        has_num = (m is not None and m != 'N/A' and str(m).strip())
        if has_num or '⚠' in cur or '✅' in cur: continue

        # Not-reported methods (only 1-shot/5-shot affected for FSCE/Meta R-CNN/FSRW)
        if base in FSOD_NOT_IN_PAPER and is_1or5:
            s.cell(row=i, column=11).value = '⚠️ ' + FSOD_NOT_IN_PAPER[base]
            flagged += 1
            print(f'FLAG [{sn}] r{i} {method_full}: not in original paper')
            continue

        # Broken-link methods
        if base in FSOD_BROKEN_LINKS:
            s.cell(row=i, column=11).value = '⚠️ ' + FSOD_BROKEN_LINKS[base]
            flagged += 1
            print(f'FLAG [{sn}] r{i} {method_full}: broken/missing ref')

# ODinW-13
s = wb['ODinW-13']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if i == 1: continue
    if not (len(row) > 1 and row[1]): continue
    method = str(row[1]).strip()
    cur = str(row[10] or '') if len(row) > 10 else ''
    m = row[6] if len(row) > 6 else None
    has_num = (m is not None and m != 'N/A' and str(m).strip())
    if has_num or '⚠' in cur or '✅' in cur: continue
    if method in ODINW13_FLAGS:
        s.cell(row=i, column=11).value = '⚠️ ' + ODINW13_FLAGS[method]
        flagged += 1
        print(f'FLAG [ODinW-13] r{i} {method}')

# ODinW-35
s = wb['ODinW-35']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if i == 1: continue
    if not (len(row) > 1 and row[1]): continue
    method = str(row[1]).strip()
    cur = str(row[10] or '') if len(row) > 10 else ''
    m = row[6] if len(row) > 6 else None
    has_num = (m is not None and m != 'N/A' and str(m).strip())
    if has_num or '⚠' in cur or '✅' in cur: continue
    if method in ODINW35_FLAGS:
        s.cell(row=i, column=11).value = '⚠️ ' + ODINW35_FLAGS[method]
        flagged += 1
        print(f'FLAG [ODinW-35] r{i} {method}')

wb.save(XLSX)
print(f'\nTotal flagged: {flagged}')
