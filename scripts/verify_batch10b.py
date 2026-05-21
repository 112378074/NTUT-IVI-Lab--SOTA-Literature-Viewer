#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_batch10b.py — 2 fills + 1 remove."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

FILLS = [
    # GLAD @ MPDD (Table 2 average)
    ('MPDD', 17, 97.5, 98.7, None, None,
     '✅ 已驗證 (GLAD ECCV24 Table 2 average)：Global-Local Anomaly Detection；LDM + DINO fine-tuned；MPDD I=97.5 P=98.7 [基於擴散]'),
    # April-GAN @ VisA (Table 4 zero-shot)
    ('VisA', 72, 78.0, 94.2, None, 86.8,
     '✅ 已驗證 (April-GAN VAND2023 Table 4 zero-shot)：CLIP-AD with linear projection adapters；VAND 2023 Challenge zero-shot 第一名 [基於 VLM]'),
]

DELETES = [
    # TFA-Net — arxiv ID 2603.22874 is a placeholder (year 2026 prefix) for a not-yet-existent paper
    ('MVTec LOCO', 21, 'TFA-Net arxiv 2603.22874 is a placeholder ID; paper not actually available'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> i={i}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason[:60]}...)')

    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))

if __name__ == '__main__':
    main()
