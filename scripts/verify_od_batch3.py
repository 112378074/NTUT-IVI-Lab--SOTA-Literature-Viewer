#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batch3.py — classic OD methods from DETR/Deformable DETR/EfficientDet PDFs."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

FILLS = [
    # EfficientDet-D7 @ COCO 2017 + test-dev (EfficientDet paper abstract, single-model single-scale)
    ('COCO 2017', 51, 55.1, 'AP 55.1 test-dev',
     '✅ 已驗證 (EfficientDet CVPR20 paper abstract)：EfficientDet-D7 55.1% AP on COCO test-dev, single-model single-scale [Real-Time OD]'),
    ('COCO test-dev', 20, 55.1, 'AP 55.1',
     '✅ 已驗證 (EfficientDet CVPR20 paper abstract)：EfficientDet-D7 55.1% AP, 77M params, 410B FLOPs [Real-Time OD]'),
    # Deformable DETR (R50) on COCO 2017 test-dev (Deformable DETR Table 3)
    ('COCO 2017 val', 31, 46.9, 'AP 46.9 (ResNet-50)',
     '✅ 已驗證 (Deformable DETR ICLR21 Table 3)：R50 backbone, 46.9% AP on COCO 2017 test-dev [DETR-based OD]'),
    ('COCO 2017', 46, 46.9, 'AP 46.9 (ResNet-50)',
     '✅ 已驗證 (Deformable DETR ICLR21 Table 3)：R50 backbone, 46.9% AP, end-to-end transformer detector [DETR-based]'),
    # DETR (R101) on COCO 2017 val (DETR Table 1)
    ('COCO 2017 val', 33, 43.5, 'AP 43.5 val',
     '✅ 已驗證 (DETR ECCV20 Table 1)：DETR-R101 43.5% AP on COCO 2017 validation [DETR-based]'),
    ('COCO 2017', 48, 43.5, 'AP 43.5 val',
     '✅ 已驗證 (DETR ECCV20 Table 1)：DETR-R101 43.5% AP on COCO 2017 val [DETR-based OD]'),
    # Faster R-CNN (R50-FPN) on COCO 2017 (DETR Table 1 Detectron2 cite)
    ('COCO 2017', 59, 40.2, 'AP 40.2 val (Detectron2)',
     '✅ 已驗證 (DETR Table 1 cites Detectron2 baseline)：Faster R-CNN R50-FPN 40.2% AP on COCO 2017 val [Two-Stage OD baseline]'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
