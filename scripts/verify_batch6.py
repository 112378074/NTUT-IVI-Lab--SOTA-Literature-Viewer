#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verify_batch6.py — fill PDF-verified metrics into AD workbook rows previously
flagged "⚠️ 指標待驗證". Mark them with "✅ 已驗證" + 1-line source.
Also removes one misplaced row (EfficientAD @ MVTec 3D — the EfficientAD paper
only tests MVTec AD / VisA / MVTec LOCO, not MVTec 3D-AD).

Run from project root:
    python scripts/verify_batch6.py
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# (sheet, row_1based, i_auroc, p_auroc, p_ap, p_pro, note_appendix)
FILLS = [
    # Dinomaly @ MPDD — Dinomaly paper Table A11 (multi-class)
    ('MPDD', 16, 97.2, 99.1, 59.5, 96.6,
     '✅ 已驗證 (Dinomaly Table A11 multi-class)：DINOv2 ViT encoder + ViT decoder 特徵重構；Less is More [基於重構]'),
    # PromptAD @ MPDD — PromptAD paper Table 9 (4-shot)
    ('MPDD', 19, 87.2, 97.3, None, None,
     '✅ 已驗證 (PromptAD Table 9, 4-shot)：Few-shot prompt learning；ViT-B/16+ CLIP backbone [基於 VLM]'),
    # AnomalyCLIP @ MPDD — AnomalyCLIP Table 1 (zero-shot)
    ('MPDD', 23, 77.0, 96.5, None, 88.7,
     '✅ 已驗證 (AnomalyCLIP Table 1, zero-shot)：Object-agnostic prompt learning；零樣本通用異常 prompts [基於 VLM]'),
    # MambaAD @ MVTec 3D — MambaAD Table A5/A6 (multi-class)
    ('MVTec 3D', 22, 86.2, 98.6, 37.5, 93.6,
     '✅ 已驗證 (MambaAD Table A5/A6 multi-class)：Mamba SSM 多類別異常偵測；Hilbert Hybrid Scanning [基於 SSM]'),
    # Dinomaly2 @ BTAD — Dinomaly2 paper Table 2 (Dinomaly2-L, multi-class)
    ('BTAD', 20, 97.5, 98.2, 73.8, 79.8,
     '✅ 已驗證 (Dinomaly2 Table 2, Dinomaly2-L multi-class)：DINOv2 ViT-L Encoder + ViT decoder；BTAD 多類別 SOTA [基於重構]'),
    # MuSc @ BTAD — MuSc Table 14 (zero-shot)
    ('BTAD', 27, 94.8, 97.3, None, None,
     '✅ 已驗證 (MuSc Table 14, zero-shot)：未標註測試集內互評分；BTAD 0-shot 超越多 full-shot 方法 [零樣本互評]'),
    # AnomalyCLIP @ BTAD — AnomalyCLIP Table 1 (zero-shot)
    ('BTAD', 30, 88.3, 94.2, None, 74.8,
     '✅ 已驗證 (AnomalyCLIP Table 1, zero-shot)：Object-agnostic prompt learning；零樣本 BTAD [基於 VLM]'),
    # PromptAD @ MVTec LOCO — PromptAD Table 9 (4-shot)
    ('MVTec LOCO', 24, 73.5, 74.5, None, None,
     '✅ 已驗證 (PromptAD Table 9, 4-shot)：Few-shot prompt learning；MVTec LOCO 4-shot [基於 VLM]'),
]

# (sheet, row_1based, reason) — rows to delete (misplaced)
DELETES = [
    ('MVTec 3D', 20,
     'EfficientAD paper (Table 1/2) only tests MVTec AD / VisA / MVTec LOCO; '
     'does not evaluate on MVTec 3D-AD. Row was auto-drafted from passing mention.'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []

    # Fills first (row indices don't shift)
    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur_method = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur_method} → i={i}, p={p}, ap={ap}, pro={pro}')

    # Deletes (do in reverse row order per sheet to keep indices stable)
    from collections import defaultdict
    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur_method = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur_method} ({reason})')

    wb.save(XLSX)
    print('Saved', XLSX)
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
