#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_papers.py
Auto-fetch new AD / OD papers from arXiv, classify, append to xlsx,
regenerate JSON + index.html, and git push.

Triggered by Windows Task Scheduler every Wed and Fri at 02:00.

Usage:
    python update_papers.py          # full run with git push
    python update_papers.py --dry    # fetch + classify + log only, no writes
    python update_papers.py --no-push # update files but skip git push
"""

import os, sys, json, re, time, argparse, subprocess, ssl, smtplib
import urllib.request, urllib.parse
import xml.etree.ElementTree as ET
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate, make_msgid

# Resolve SSL: prefer certifi's CA bundle (works in Anaconda/Windows where
# the default urllib doesn't always find a usable trust store).
try:
    import certifi
    SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except Exception:
    SSL_CTX = ssl.create_default_context()
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from copy import copy

# ====================================================================
# Configuration
# ====================================================================
PROJECT_DIR = Path(__file__).resolve().parent.parent
AD_XLSX     = PROJECT_DIR / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'
OD_XLSX     = PROJECT_DIR / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'
CLS_XLSX    = PROJECT_DIR / 'Image_Classification_Papers_Ranking_2021_2026.xlsx'
AS_XLSX     = PROJECT_DIR / 'Anomaly_Synthesis_Papers_Benchmark_2021_2026.xlsx'
INDEX_HTML  = PROJECT_DIR / 'index.html'
AD_JSON     = PROJECT_DIR / 'papers_data.json'
OD_JSON     = PROJECT_DIR / 'od_data.json'
CLS_JSON    = PROJECT_DIR / 'cls_data.json'
AS_JSON     = PROJECT_DIR / 'as_data.json'
LOG_FILE    = PROJECT_DIR / 'scripts' / 'update_log.txt'
ENV_FILE    = PROJECT_DIR / 'scripts' / '.env'

ARXIV_API = 'http://export.arxiv.org/api/query'
ATOM_NS   = '{http://www.w3.org/2005/Atom}'

# Search queries — derived from CLAUDE.md (AD) and CLAUDE_updated.md (OD).
# Polite to arXiv: 5s sleep between queries, 60s timeout, 3 retries.

AD_QUERIES = [
    # Single broad AD query: anomaly + relevant dataset/method tokens
    'cat:cs.CV+AND+%28all:%22anomaly+detection%22+OR+all:%22anomaly+segmentation%22+OR+all:%22defect+detection%22+OR+all:%22defect+localization%22%29',
]

OD_QUERIES = [
    # Broad OD query: detection + COCO/LVIS/DUTS/ODinW + key model tokens (incl. 3D as sub-area)
    'cat:cs.CV+AND+%28all:%22object+detection%22+OR+all:%22salient+object%22+OR+all:%22few-shot+detection%22+OR+all:%22YOLO%22+OR+all:%22DETR%22+OR+all:%223D+object+detection%22+OR+all:%22LiDAR%22%29',
]

CLS_QUERIES = [
    # ImageNet / CIFAR / fine-grained / long-tail / few-shot CLS
    'cat:cs.CV+AND+%28all:%22image+classification%22+AND+%28all:%22ImageNet%22+OR+all:%22CIFAR%22+OR+all:%22fine-grained%22+OR+all:%22long-tailed%22+OR+all:%22few-shot%22%29%29',
]

AS_QUERIES = [
    # Anomaly synthesis / synthetic defect / pseudo anomaly
    'cat:cs.CV+AND+%28all:%22anomaly+synthesis%22+OR+all:%22synthetic+anomaly%22+OR+all:%22pseudo+anomaly%22+OR+all:%22defect+generation%22+OR+all:%22DRAEM%22+OR+all:%22GLASS%22+OR+all:%22DualAnoDiff%22%29',
]

# 3D OD is now a sub-area (per user spec).  Keep medical / segmentation-only /
# tracking-only / adversarial / niche-domain exclusions.
OD_EXCLUSION_KEYWORDS = [
    'medical image', 'lesion detection', 'tumor detection', 'polyp detection',
    'tooth', 'caries', 'dental', 'retinal', 'airway-tree', 'cardiac',
    'sar object', 'remote sensing',
    'underwater', 'weed detection', 'pedestrian tracking',
    'adversarial patch', 'adversarial attack',
]
AD_EXCLUSION_KEYWORDS = [
    '3d lidar', 'lidar anomaly', 'point cloud anomaly',  # only 2D RGB / RGB-D AD
    'medical image anomaly', 'retinal abnormalit', 'lesion ',
    'video anomaly detection', 'temporal anomaly',  # not industrial visual AD
    'time series anomaly', 'log anomaly', 'network anomaly',
    'speech anomaly', 'audio anomaly',
    'surveillance', 'expressway', 'traffic anomaly',
    'fraud detection', 'cyber',
    'llm adaptation', 'language model adaptation',
]
def is_od_excluded(text):
    t = text.lower()
    return any(kw in t for kw in OD_EXCLUSION_KEYWORDS)
def is_ad_excluded(text):
    t = text.lower()
    return any(kw in t for kw in AD_EXCLUSION_KEYWORDS)

# AD category classification (priority order)
AD_CATEGORIES_BY_KEYWORDS = [
    ('基於擴散 (Diffusion)',           ['diffusion model', 'denoising diffusion', 'ddpm', 'ddim', 'score-based']),
    ('基於 NF (Normalizing Flow)',     ['normalizing flow', 'flow-based', 'fastflow', 'msflow', 'cflow', 'real-nvp']),
    ('基於資料擴增 (Data Augmentation)', ['anomaly synthesis', 'synthetic anomal', 'cutpaste', 'draem', 'glass',
                                        'pseudo anomal', 'data augmentation', 'synthesizing anomal']),
    ('基於重構 (Reconstruction)',       ['reconstruction', 'autoencoder', 'reconstruct', 'inpainting',
                                        'dinomaly', 'masked image modeling']),
    ('基於表徵 (Representation)',       ['patchcore', 'memory bank', 'student-teacher', 'student teacher',
                                        'knowledge distillation', 'efficientad', 'simplenet', 'embedding',
                                        'representation learning', 'feature distribution']),
]
AD_DEFAULT_CATEGORY = '基於表徵 (Representation)'

# AD dataset detection — order: more specific first
AD_DATASETS_BY_PATTERNS = [
    ('MVTec AD 2',  [r'mvtec\s*ad\s*2', r'mvtec\s*ad2']),
    ('MVTec LOCO',  [r'mvtec\s*loco', r'\bloco-?ad\b']),
    ('MVTec 3D',    [r'mvtec\s*3d', r'mvtec-3d-ad']),
    ('MVTec AD',    [r'mvtec\s*ad\b', r'mvtec-ad\b', r'\bmvtec\b']),
    ('VisA',        [r'\bvisa\b']),
    ('MPDD',        [r'\bmpdd\b']),
    ('BTAD',        [r'\bbtad\b', r'beantech']),
]
AD_DEFAULT_DATASET = 'MVTec AD'

# OD category classification — by alias / topic keywords. CLAUDE_updated.md §3.
OD_CATEGORIES_BY_KEYWORDS = [
    ('Few-Shot OD',    ['few-shot object detection', 'fsod', 'novel ap', 'base-to-novel',
                        'open-vocabulary detection', 'open vocabulary detection',
                        'open-vocabulary object detection', 'open-set detection',
                        'class-incremental detection', 'k-shot detection',
                        'cd-vito', 'unifs', 'grounding dino', 'odinw']),
    ('RGB Salient OD', ['salient object detection', 'saliency detection',
                        'salient region', 'birefnet', 'inspyrenet', 'duts-te',
                        'dut-omron', 'high-resolution salient']),
    ('Real-Time OD',   ['real-time object detection', 'real-time detection',
                        'yolo', 'rt-detr', 'rtdetr', 'd-fine', 'dfine ',
                        'deim ', 'rf-detr', 'rfdetr',
                        'rtmdet', 'efficientdet', 'nanodet', 'mobiledet',
                        'real time detector', 'streaming detection',
                        'edge detection', 'lightweight detection']),
]
OD_DEFAULT_CATEGORY = 'General OD'

# OD dataset detection (subset — only the ones already in workbook)
OD_DATASETS_BY_PATTERNS = [
    # Few-shot specific (most-specific first)
    ('MS-COCO 30-shot',      [r'30-shot', r'30\s+shot']),
    ('MS-COCO 10-shot',      [r'10-shot', r'10\s+shot']),
    ('MS-COCO 5-shot',       [r'5-shot',  r'5\s+shot']),
    ('MS-COCO 1-shot',       [r'1-shot',  r'1\s+shot', r'one-shot']),
    ('PASCAL VOC 2007 15+5', [r'voc\s*15\+?5', r'15\+5\s*split']),
    ('LVIS v1.0 test-dev',   [r'lvis\s*v?1\.?0?\s*test-dev', r'lvis.*test-dev']),
    ('LVIS v1.0 val',        [r'\blvis\b']),
    ('ODinW-35',             [r'odinw-?35']),
    ('ODinW-13',             [r'odinw-?13']),
    ('COCO 2017 FSOD',       [r'coco.*fsod', r'fsod.*coco']),
    # General OD
    ('COCO test-dev',        [r'coco\s*test-?dev', r'\btest-dev\b']),
    ('COCO minival',         [r'coco\s*minival', r'\bminival\b']),
    ('COCO-O',               [r'coco-o\b']),
    ('COCO 2017 val',        [r'coco\s*2017\s*val', r'coco-val', r'val2017']),
    ('COCO 2017',            [r'coco\s*2017']),
    ('PASCAL VOC 2007',      [r'pascal\s*voc\s*2007', r'\bvoc\s*2007']),
    ('CrowdHuman',           [r'crowdhuman']),
    ('GraZPEDWRI-DX',        [r'grazpedwri', r'pedwri']),
    ('CPPE-5',               [r'cppe-5', r'\bcppe\b']),
    ('Waymo 2D',             [r'waymo\s*2d']),
    # Argoverse-HD splits (most-specific first)
    ('Argoverse-HD FS Test', [r'argoverse.*full-stack.*test']),
    ('Argoverse-HD DO Test', [r'argoverse.*detection-only.*test']),
    ('Argoverse-HD FS Val',  [r'argoverse.*full-stack']),
    ('Argoverse-HD DO Val',  [r'argoverse']),
    # Salient OD
    ('DUTS-TE',              [r'duts-te', r'\bduts\b']),
    ('DUT-OMRON',            [r'dut-omron', r'omron']),
    ('ECSSD',                [r'ecssd']),
    ('HKU-IS',               [r'hku-is']),
    ('PASCAL-S',             [r'pascal-s']),
    ('HRSOD',                [r'\bhrsod\b']),
    ('UHRSD',                [r'\buhrsd\b']),
    ('DAVIS-S',              [r'davis-s']),
    ('SBU-Refine',           [r'sbu-refine']),
    ('ISTD',                 [r'\bistd\b']),
    ('CAMO-FS',              [r'camo-fs']),
]
OD_DEFAULT_DATASET = 'COCO 2017 val'

# CLS classification + dataset detection
CLS_CATEGORIES_BY_KEYWORDS = [
    ('Few-Shot Image Classification-based',     ['few-shot classification', 'miniimagenet', 'tieredimagenet', 'cifar-fs', 'fc100',
                                                  'meta-learning classification', 'k-shot image classification', 'low-shot']),
    ('Long-Tailed / Imbalanced Image Classification-based', ['long-tailed', 'long-tail classification', 'imbalanced classification',
                                                              'imagenet-lt', 'places-lt', 'class-imbalance']),
    ('Fine-Grained Image Classification-based', ['fine-grained classification', 'fine-grained visual', 'cub-200', 'stanford cars',
                                                  'fgvc-aircraft', 'oxford flowers', 'nabirds', 'stanford dogs']),
]
CLS_DEFAULT_CATEGORY = 'General Image Classification-based'

CLS_DATASETS_BY_PATTERNS = [
    ('miniImageNet 5w-1s',  [r'miniimagenet.*1[-\s]?shot']),
    ('miniImageNet 5w-5s',  [r'miniimagenet.*5[-\s]?shot']),
    ('tieredImageNet 5w-1s',[r'tieredimagenet.*1[-\s]?shot']),
    ('tieredImageNet 5w-5s',[r'tieredimagenet.*5[-\s]?shot']),
    ('CIFAR-FS 5w-1s',      [r'cifar-fs.*1[-\s]?shot']),
    ('CIFAR-FS 5w-5s',      [r'cifar-fs.*5[-\s]?shot']),
    ('FC100 5w-1s',         [r'fc100.*1[-\s]?shot']),
    ('FC100 5w-5s',         [r'fc100.*5[-\s]?shot']),
    ('CUB few-shot 5w-1s',  [r'cub.*few[-\s]?shot.*1[-\s]?shot']),
    ('CUB few-shot 5w-5s',  [r'cub.*few[-\s]?shot.*5[-\s]?shot']),
    ('CIFAR-100-LT IF=100', [r'cifar-?100-lt', r'cifar100-?lt']),
    ('CIFAR-10-LT IF=100',  [r'cifar-?10-lt', r'cifar10-?lt']),
    ('iNaturalist 2018',    [r'inaturalist.*2018']),
    ('ImageNet-LT',         [r'imagenet-lt']),
    ('Places-LT',           [r'places-lt']),
    ('ImageNet-V2',         [r'imagenet-?v2']),
    ('ImageNet-ReaL',       [r'imagenet-?real']),
    ('ImageNet-A',          [r'imagenet-?a\b']),
    ('ImageNet-R',          [r'imagenet-?r\b']),
    ('ImageNet-Sketch',     [r'imagenet-?sketch']),
    ('ImageNet-1K',         [r'imagenet-?1k', r'imagenet[^-]']),
    ('CIFAR-10',            [r'cifar-?10\b']),
    ('CIFAR-100',           [r'cifar-?100\b']),
    ('STL-10',              [r'stl-?10']),
    ('Places365',           [r'places365']),
    ('Food-101',            [r'food-?101']),
    ('CUB-200-2011',        [r'cub-?200', r'cub-200-2011']),
    ('Stanford Cars',       [r'stanford cars']),
    ('FGVC-Aircraft',       [r'fgvc-?aircraft']),
    ('NABirds',             [r'nabirds']),
    ('Oxford Flowers-102',  [r'oxford flowers', r'flowers-?102']),
    ('Stanford Dogs',       [r'stanford dogs']),
    ('Oxford-IIIT Pets',    [r'oxford.*pets']),
    ('iNaturalist',         [r'inaturalist']),
]
CLS_DEFAULT_DATASET = 'ImageNet-1K'

CLS_EXCLUSION_KEYWORDS = [
    '3d classification', 'point cloud classification',
    'medical image classification', 'pathology classification',
    'audio classification', 'video classification',
    'time series classification', 'graph classification',
    'text classification',
]
def is_cls_excluded(t):
    t = t.lower()
    return any(kw in t for kw in CLS_EXCLUSION_KEYWORDS)

# AS classification + dataset detection
AS_CATEGORIES_BY_KEYWORDS = [
    ('Vision-language / Foundation-model-based Anomaly Synthesis', ['vlm', 'vision-language', 'clip', 'dinov2', 'foundation model',
                                                                      'gpt', 'agent', 'agentic']),
    ('Generative-model-based Anomaly Synthesis',                   ['diffusion', 'gan-based', 'stable diffusion', 'ddpm', 'generative',
                                                                      'sdas', 'realnet', 'dualanodiff']),
    ('Distribution-hypothesis / Feature-space Anomaly Synthesis',  ['feature-space', 'feature space', 'distribution hypothesis',
                                                                      'glass', 'simplenet']),
    ('Hand-crafted / Rule-based Anomaly Synthesis',                ['cutpaste', 'draem', 'perlin noise', 'rule-based', 'hand-crafted',
                                                                      'procedural defect']),
]
AS_DEFAULT_CATEGORY = 'Downstream Synthesis-based AD / Segmentation'

AS_DATASETS_BY_PATTERNS = [
    ('MVTec AD 2 TESTpub',  [r'mvtec\s*ad\s*2.*testpub', r'mvtec\s*ad2.*testpub']),
    ('MVTec AD 2 TESTpriv', [r'mvtec\s*ad\s*2.*testpriv']),
    ('MVTec LOCO AD',       [r'mvtec\s*loco']),
    ('MVTec AD pixel-level',[r'mvtec\s*ad.*pixel']),
    ('MVTec AD image-level',[r'mvtec\s*ad']),
    ('VisA pixel-level',    [r'\bvisa\b.*pixel']),
    ('VisA image-level',    [r'\bvisa\b']),
    ('Real-IAD pixel-level',[r'real-iad.*pixel']),
    ('Real-IAD image-level',[r'real-iad']),
    ('BTAD',                [r'\bbtad\b', r'beantech']),
    ('MPDD',                [r'\bmpdd\b']),
    ('DAGM',                [r'\bdagm\b']),
    ('KSDD',                [r'\bksdd\b']),
    ('KSDD2',               [r'\bksdd2\b']),
    ('Magnetic Tile Defects',[r'magnetic\s*tile']),
    ('BTech',               [r'\bbtech\b']),
    ('Synthesis Quality',   [r'fid', r'kid score', r'synthesis quality', r'fréchet inception']),
]
AS_DEFAULT_DATASET = 'MVTec AD image-level'

AS_EXCLUSION_KEYWORDS = [
    'video anomaly', 'time series anomaly', 'medical anomaly',
    'log anomaly', 'network anomaly', 'speech anomaly',
]
def is_as_excluded(t):
    t = t.lower()
    return any(kw in t for kw in AS_EXCLUSION_KEYWORDS)

# How far back to look for new papers
LOOKBACK_DAYS = 7

# ====================================================================
# Utilities
# ====================================================================
def log(msg):
    line = f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {msg}'
    # On Windows the default stdout encoding is cp950/cp1252 and crashes on
    # non-ASCII paper titles. Re-encode the print call defensively so the run
    # never aborts on a stray accent/CJK char.
    try:
        print(line, flush=True)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, 'encoding', None) or 'ascii'
        print(line.encode(enc, 'replace').decode(enc, 'replace'), flush=True)
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

def fetch_arxiv(query, max_results=200, retries=3):
    url = f'{ARXIV_API}?search_query={query}&sortBy=submittedDate&sortOrder=descending&max_results={max_results}'
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'NTUT-AIL-PaperBot/1.0'})
            with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as r:
                return ET.fromstring(r.read())
        except Exception as e:
            last_err = e
            wait = 5 * (attempt + 1)  # 5s, 10s, 15s
            log(f'  arxiv fetch retry {attempt+1}/{retries} after {wait}s ({type(e).__name__})')
            time.sleep(wait)
    log(f'  arxiv fetch failed for query [{query[:60]}…]: {last_err}')
    return None

def parse_entries(root):
    if root is None:
        return []
    out = []
    for e in root.findall(f'{ATOM_NS}entry'):
        eid = (e.findtext(f'{ATOM_NS}id') or '').strip()
        title = re.sub(r'\s+', ' ', e.findtext(f'{ATOM_NS}title') or '').strip()
        summary = re.sub(r'\s+', ' ', e.findtext(f'{ATOM_NS}summary') or '').strip()
        published = (e.findtext(f'{ATOM_NS}published') or '').strip()
        authors = [a.findtext(f'{ATOM_NS}name') for a in e.findall(f'{ATOM_NS}author')]
        authors = [a for a in authors if a]
        m = re.search(r'arxiv\.org/abs/(\d+\.\d+)', eid)
        arxiv_id = m.group(1) if m else None
        out.append({
            'arxiv_id': arxiv_id,
            'title': title,
            'summary': summary,
            'authors': ', '.join(authors[:8]),
            'published': published,
            'date': published[:7] if published else '',
            'link': f'https://arxiv.org/abs/{arxiv_id}' if arxiv_id else eid,
        })
    return out

def classify_category(text, mapping):
    text_l = text.lower()
    for label, kws in mapping:
        for kw in kws:
            if kw and kw in text_l:
                return label
    return None

def detect_dataset(text, mapping):
    text_l = text.lower()
    for ds, patterns in mapping:
        for p in patterns:
            if re.search(p, text_l):
                return ds
    return None

def is_within_window(published_iso, days):
    try:
        dt = datetime.strptime(published_iso[:10], '%Y-%m-%d').replace(tzinfo=timezone.utc)
    except Exception:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    return dt >= cutoff

# ====================================================================
# Existing-papers index (to dedupe)
# ====================================================================
# ====================================================================
# CVF Open Access scan (CVPR / ICCV / ECCV / WACV proceedings)
# ====================================================================
CVF_BASE = 'https://openaccess.thecvf.com'
# (conference, year) pairs to scan every run.  Dedup makes re-scans harmless;
# new proceedings (e.g. CVPR 2026) are picked up automatically once published.
CVF_TARGETS = [
    ('CVPR', '2025'), ('ICCV', '2025'), ('ECCV', '2024'), ('WACV', '2025'),
    ('CVPR', '2026'), ('WACV', '2026'),
]
CVF_AD_KW  = ['anomaly detection', 'anomaly segmentation', 'defect detection',
              'defect localization', 'mvtec', 'visa', 'mpdd', 'btad',
              'industrial anomaly', 'surface defect']
CVF_OD_KW  = ['object detection', 'salient object', 'few-shot detection',
              'real-time detection', '3d object detection', 'detr', 'yolo',
              'open-vocabulary detection']
CVF_CLS_KW = ['image classification', 'fine-grained classification',
              'few-shot classification', 'semi-supervised classification',
              'long-tailed classification', 'small data classification',
              'image recognition']
CVF_EXCLUDE = ['autonomous driv', 'lidar', 'medical', 'x-ray', 'oral ', 'dental',
               'retinal', 'tooth', 'video anomaly', 'audio', 'speech', 'eeg',
               'remote sensing', 'satellite', 'underwater', 'whole slide',
               'histopath', 'climate']
CVF_VENUE_DATE = {'CVPR': '2025-06', 'ICCV': '2025-10', 'ECCV': '2024-10', 'WACV': '2025-01'}


def fetch_cvf_conference(conf, year):
    """Return list of {title, paper_url} for one CVF proceedings page.
    Tries two URL forms; each with up to 2 retries on a transient error."""
    out = []
    for suffix in (f'{conf}{year}?day=all', f'{conf}{year}'):
        html = None
        for attempt in range(2):
            try:
                req = urllib.request.Request(f'{CVF_BASE}/{suffix}',
                                             headers={'User-Agent': 'NTUT-AIL-PaperBot/1.0'})
                with urllib.request.urlopen(req, timeout=45, context=SSL_CTX) as r:
                    html = r.read().decode('utf-8', errors='replace')
                break
            except Exception as e:
                if attempt == 0:
                    time.sleep(8)
                    continue
                log(f'  CVF fetch failed {conf}{year}: {e}')
        if not html:
            continue
        for m in re.finditer(
                r'<dt class="ptitle">(?:<br>)?<a href="(/content/[^"]+)">([^<]+)</a>', html):
            out.append({'title': re.sub(r'\s+', ' ', m.group(2)).strip(),
                        'paper_url': CVF_BASE + m.group(1)})
        if out:
            break
    return out


def _cvf_domain(text):
    t = text.lower()
    if any(k in t for k in CVF_EXCLUDE): return None
    if any(k in t for k in CVF_AD_KW):   return 'AD'
    if any(k in t for k in CVF_OD_KW):   return 'OD'
    if any(k in t for k in CVF_CLS_KW):  return 'CLS'
    return None


def process_cvf():
    """Scan CVF Open Access, append genuinely-new papers to the All Papers
    sheets of each workbook.  Returns a list of added paper dicts."""
    # Existing method names for dedup
    existing = set()
    for path, mcol in [(AD_XLSX, 3), (OD_XLSX, 2), (CLS_XLSX, 2)]:
        if not path.exists(): continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            for row in wb[sn].iter_rows(values_only=True):
                if len(row) >= mcol and isinstance(row[mcol-1], str):
                    base = re.sub(r'\(.*?\)', '', row[mcol-1]).strip().lower()
                    existing.add(re.sub(r'[^a-z0-9]', '', base))
        wb.close()

    def method_of(title):
        if ':' in title:
            cand = title.split(':')[0].strip()
            if 1 <= len(cand.split()) <= 5:
                return cand
        return title[:60]

    new = []
    for conf, year in CVF_TARGETS:
        papers = fetch_cvf_conference(conf, year)
        if not papers:
            continue
        log(f'  CVF {conf}{year}: {len(papers)} papers')
        for p in papers:
            dom = _cvf_domain(p['title'])
            if not dom:
                continue
            method = method_of(p['title'])
            if re.sub(r'[^a-z0-9]', '', method.lower()) in existing:
                continue
            existing.add(re.sub(r'[^a-z0-9]', '', method.lower()))
            new.append({'domain': dom, 'conf': conf, 'year': year,
                        'title': p['title'], 'method': method,
                        'venue': f'{conf} {year}',
                        'date': CVF_VENUE_DATE.get(conf, f'{year}-00'),
                        'link': p['paper_url']})
        time.sleep(3)
    log(f'CVF new papers: {len(new)}')
    return new


def append_cvf_rows(new_cvf):
    """Append CVF papers to the All Papers sheets (metrics blank — pending verification)."""
    from copy import copy as _copy
    def _style(sh, src, dest, n):
        for c in range(1, n+1):
            s = sh.cell(row=src, column=c); d = sh.cell(row=dest, column=c)
            if s.has_style:
                d.font=_copy(s.font); d.fill=_copy(s.fill); d.border=_copy(s.border)
                d.alignment=_copy(s.alignment); d.number_format=s.number_format
    def _last(sh, kc):
        last = sh.max_row
        while last > 1 and sh.cell(row=last, column=kc).value in (None, ''):
            last -= 1
        return last
    note = lambda t: f'CVF Open Access 自動收錄；指標待人工驗證。{t}'

    ad = [r for r in new_cvf if r['domain'] == 'AD']
    od = [r for r in new_cvf if r['domain'] == 'OD']
    cl = [r for r in new_cvf if r['domain'] == 'CLS']

    if ad and AD_XLSX.exists():
        wb = load_workbook(AD_XLSX); sh = wb['總覽 All Papers']; last = _last(sh, 3)
        for r in ad:
            cat = classify_category(r['title'], AD_CATEGORIES_BY_KEYWORDS) or AD_DEFAULT_CATEGORY
            nr = last + 1; last = nr
            vals = ['—', cat, r['method'], '—', r['venue'], r['date'], '已發表',
                    None, None, None, None, None, note(r['title']), r['link'], 'N/A']
            for c, v in enumerate(vals, 1): sh.cell(row=nr, column=c).value = v
            _style(sh, nr-1 if nr > 2 else 2, nr, 15)
        wb.save(AD_XLSX)
        log(f'  CVF -> AD 總覽 All Papers: +{len(ad)}')

    if od and OD_XLSX.exists():
        wb = load_workbook(OD_XLSX); sh = wb['OD all papers']; last = _last(sh, 2)
        for r in od:
            cat = classify_category(r['title'], OD_CATEGORIES_BY_KEYWORDS) or OD_DEFAULT_CATEGORY
            nr = last + 1; last = nr
            vals = [cat, r['method'], '—', r['venue'], r['date'], note(r['title']), r['link'], 'N/A']
            for c, v in enumerate(vals, 1): sh.cell(row=nr, column=c).value = v
            _style(sh, nr-1 if nr > 2 else 2, nr, 8)
        wb.save(OD_XLSX)
        log(f'  CVF -> OD all papers: +{len(od)}')

    if cl and CLS_XLSX.exists():
        wb = load_workbook(CLS_XLSX); sh = wb['Classification all papers']; last = _last(sh, 2)
        for r in cl:
            tl = r['title'].lower()
            cat = ('Few-Shot Image Classification-based' if 'few-shot' in tl or 'few shot' in tl
                   else 'Fine-Grained Image Classification-based' if 'fine-grained' in tl
                   else 'Semi-Supervised Image Classification-based' if 'semi-supervised' in tl
                   else 'Small Data Image Classification-based' if 'small data' in tl or 'small-data' in tl
                   else 'General Image Classification-based')
            nr = last + 1; last = nr
            vals = [cat, r['method'], '—', r['venue'], r['date'], '已發表',
                    None, None, None, None, None, None, note(r['title']), r['link'], 'N/A']
            for c, v in enumerate(vals, 1): sh.cell(row=nr, column=c).value = v
            _style(sh, nr-1 if nr > 2 else 2, nr, 15)
        wb.save(CLS_XLSX)
        log(f'  CVF -> CLS all papers: +{len(cl)}')


def collect_existing_arxiv_ids(xlsx_path):
    ids = set()
    titles = set()
    if not xlsx_path.exists():
        return ids, titles
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            sh = wb[sn]
            for row in sh.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, str):
                        for m in re.finditer(r'arxiv\.org/abs/(\d+\.\d+)', v):
                            ids.add(m.group(1))
                # Method names sometimes appear in column 3 (idx 2). Capture for fallback dedupe.
                if len(row) >= 3 and isinstance(row[2], str):
                    titles.add(row[2].strip().lower())
    finally:
        wb.close()
    return ids, titles

# ====================================================================
# Excel append helpers
# ====================================================================
def find_last_row(sh, key_col=1):
    last = sh.max_row
    while last > 1 and sh.cell(row=last, column=key_col).value in (None, ''):
        last -= 1
    return last

def style_from(sh, src_row, dest_row, ncols):
    for c in range(1, ncols + 1):
        s = sh.cell(row=src_row, column=c)
        d = sh.cell(row=dest_row, column=c)
        if s.has_style:
            d.font          = copy(s.font)
            d.fill          = copy(s.fill)
            d.border        = copy(s.border)
            d.alignment     = copy(s.alignment)
            d.number_format = s.number_format
            d.protection    = copy(s.protection)

def append_ad_row(wb, dataset_sheet, paper):
    """Append a row to a per-dataset AD sheet matching its 15-col format."""
    if dataset_sheet not in wb.sheetnames:
        return False
    sh = wb[dataset_sheet]
    last = find_last_row(sh, key_col=3)  # method col
    new_row = last + 1
    # Columns: 資料集 / 類別 / 方法 / 作者 / 發表 / 年月 / 狀態 / I-AUROC / P-AUROC / P-AP / P-PRO / FPS / 備註 / 連結 / GitHub
    sh.cell(row=new_row, column=1).value = dataset_sheet if dataset_sheet != 'MVTec AD2' else 'MVTec AD 2'
    sh.cell(row=new_row, column=2).value = paper['category']
    sh.cell(row=new_row, column=3).value = paper['method']
    sh.cell(row=new_row, column=4).value = paper['authors']
    sh.cell(row=new_row, column=5).value = 'arXiv preprint'
    sh.cell(row=new_row, column=6).value = paper['date']
    sh.cell(row=new_row, column=7).value = '預印本'
    # Metrics blank — to be filled in manually after review
    for c in range(8, 13):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=13).value = (paper.get('note') or '自動抓取 (Auto-fetched)') + ' [' + paper['category'] + ']'
    sh.cell(row=new_row, column=14).value = paper['link']
    sh.cell(row=new_row, column=15).value = None
    style_from(sh, last if last >= 3 else 3, new_row, 15)
    return True

def append_od_dataset_row(wb, dataset_sheet, paper):
    """Append to an OD per-dataset sheet (13 cols)."""
    if dataset_sheet not in wb.sheetnames:
        return False
    sh = wb[dataset_sheet]
    last = find_last_row(sh, key_col=2)  # method col
    new_row = last + 1
    # Cols: 類別 / 方法 / 作者 / 發表 / 年月 / 狀態 / mAP / AP / FPS / params / 備註 / 連結 / GitHub
    sh.cell(row=new_row, column=1).value = paper['category']
    sh.cell(row=new_row, column=2).value = paper['method']
    sh.cell(row=new_row, column=3).value = paper['authors']
    sh.cell(row=new_row, column=4).value = 'arXiv'
    sh.cell(row=new_row, column=5).value = paper['date']
    sh.cell(row=new_row, column=6).value = 'Preprint'
    for c in range(7, 11):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=11).value = paper.get('note') or 'Auto-fetched from arXiv'
    sh.cell(row=new_row, column=12).value = paper['link']
    sh.cell(row=new_row, column=13).value = None
    style_from(sh, last if last >= 2 else 2, new_row, 13)
    return True

def append_od_all_papers_row(wb, paper):
    sn = 'OD all papers'
    if sn not in wb.sheetnames:
        return False
    sh = wb[sn]
    last = find_last_row(sh, key_col=2)
    new_row = last + 1
    # Cols: 類別 / 方法 / 作者 / 發表 / 年月 / 備註 / 連結 / GitHub
    sh.cell(row=new_row, column=1).value = paper['category']
    sh.cell(row=new_row, column=2).value = paper['method']
    sh.cell(row=new_row, column=3).value = paper['authors']
    sh.cell(row=new_row, column=4).value = 'arXiv'
    sh.cell(row=new_row, column=5).value = paper['date']
    sh.cell(row=new_row, column=6).value = paper.get('note') or 'Auto-fetched from arXiv'
    sh.cell(row=new_row, column=7).value = paper['link']
    sh.cell(row=new_row, column=8).value = None
    style_from(sh, last if last >= 1 else 1, new_row, 8)
    return True

def append_cls_dataset_row(wb, dataset_sheet, paper):
    """CLS per-dataset sheet (15 cols)."""
    if dataset_sheet not in wb.sheetnames: return False
    sh = wb[dataset_sheet]
    last = find_last_row(sh, key_col=2)
    new_row = last + 1
    # Cols: 類別 / 方法 / 作者 / 發表 / 年月 / 狀態 / Top-1 / Top-5 / Acc / F1 / Params / FLOPs / 備註 / 連結 / GitHub
    sh.cell(row=new_row, column=1).value  = paper['category']
    sh.cell(row=new_row, column=2).value  = paper['method']
    sh.cell(row=new_row, column=3).value  = paper['authors']
    sh.cell(row=new_row, column=4).value  = 'arXiv'
    sh.cell(row=new_row, column=5).value  = paper['date']
    sh.cell(row=new_row, column=6).value  = 'arXiv'
    for c in range(7, 13):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=13).value = paper.get('note') or 'Auto-fetched; metrics To verify'
    sh.cell(row=new_row, column=14).value = paper['link']
    sh.cell(row=new_row, column=15).value = None
    style_from(sh, last if last >= 2 else 2, new_row, 15)
    return True

def append_cls_all_papers_row(wb, paper):
    sn = 'Classification all papers'
    if sn not in wb.sheetnames: return False
    sh = wb[sn]
    last = find_last_row(sh, key_col=2)
    new_row = last + 1
    sh.cell(row=new_row, column=1).value  = paper['category']
    sh.cell(row=new_row, column=2).value  = paper['method']
    sh.cell(row=new_row, column=3).value  = paper['authors']
    sh.cell(row=new_row, column=4).value  = 'arXiv'
    sh.cell(row=new_row, column=5).value  = paper['date']
    sh.cell(row=new_row, column=6).value  = 'arXiv'
    for c in range(7, 13):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=13).value = paper.get('note') or 'Auto-fetched; metrics To verify'
    sh.cell(row=new_row, column=14).value = paper['link']
    sh.cell(row=new_row, column=15).value = None
    style_from(sh, last if last >= 1 else 1, new_row, 15)
    return True

def append_as_dataset_row(wb, dataset_sheet, paper):
    """AS per-dataset sheet (20 cols)."""
    if dataset_sheet not in wb.sheetnames: return False
    sh = wb[dataset_sheet]
    last = find_last_row(sh, key_col=2)
    new_row = last + 1
    # Cols: 類別 / 方法 / 作者 / 發表 / 年月 / 狀態 / 合成類型 / 任務設定 / 資料集 / Split/Protocol /
    #       I-AUROC / P-AUROC / AUPRO/PRO / AP/AUPRC / F1/Dice/IoU / 合成品質指標 / Backbone/Detector / 備註 / 連結 / GitHub
    sh.cell(row=new_row, column=1).value  = paper['category']
    sh.cell(row=new_row, column=2).value  = paper['method']
    sh.cell(row=new_row, column=3).value  = paper['authors']
    sh.cell(row=new_row, column=4).value  = 'arXiv'
    sh.cell(row=new_row, column=5).value  = paper['date']
    sh.cell(row=new_row, column=6).value  = 'arXiv'
    sh.cell(row=new_row, column=7).value  = paper.get('synth_type') or 'TBD'
    sh.cell(row=new_row, column=8).value  = paper.get('task_setting') or 'TBD'
    sh.cell(row=new_row, column=9).value  = dataset_sheet
    for c in range(10, 18):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=18).value = paper.get('note') or 'Auto-fetched from arXiv'
    sh.cell(row=new_row, column=19).value = paper['link']
    sh.cell(row=new_row, column=20).value = None
    style_from(sh, last if last >= 2 else 2, new_row, 20)
    return True

def append_as_all_papers_row(wb, paper):
    sn = 'AS all papers'
    if sn not in wb.sheetnames: return False
    sh = wb[sn]
    last = find_last_row(sh, key_col=2)
    new_row = last + 1
    sh.cell(row=new_row, column=1).value  = paper['category']
    sh.cell(row=new_row, column=2).value  = paper['method']
    sh.cell(row=new_row, column=3).value  = paper['authors']
    sh.cell(row=new_row, column=4).value  = 'arXiv'
    sh.cell(row=new_row, column=5).value  = paper['date']
    sh.cell(row=new_row, column=6).value  = 'arXiv'
    sh.cell(row=new_row, column=7).value  = paper.get('synth_type') or 'TBD'
    sh.cell(row=new_row, column=8).value  = paper.get('task_setting') or 'TBD'
    sh.cell(row=new_row, column=9).value  = paper['dataset']
    for c in range(10, 18):
        sh.cell(row=new_row, column=c).value = None
    sh.cell(row=new_row, column=18).value = paper.get('note') or 'Auto-fetched from arXiv'
    sh.cell(row=new_row, column=19).value = paper['link']
    sh.cell(row=new_row, column=20).value = None
    style_from(sh, last if last >= 1 else 1, new_row, 20)
    return True

# ====================================================================
# JSON regeneration (mirrors the original extraction logic)
# ====================================================================
import math
def _clean(v):
    if v is None: return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
    if isinstance(v, str) and v.strip() == '': return None
    return v

def regenerate_ad_json():
    import pandas as pd
    xl = pd.ExcelFile(AD_XLSX)
    sheets = ['MVTec AD', 'MVTec LOCO', 'MVTec AD2', 'VisA', 'MPDD', 'BTAD', 'MVTec 3D']
    rows = []
    for sn in sheets:
        if sn not in xl.sheet_names:
            continue
        df = pd.read_excel(xl, sheet_name=sn, header=2)
        df.columns = ['dataset','category','method','authors','venue','date','status',
                      'i_auroc','p_auroc','p_ap','p_pro','fps','notes','link','github']
        df = df.dropna(subset=['method'])
        for _, r in df.iterrows():
            rec = {k: _clean(v) for k, v in r.items()}
            if rec.get('method'):
                if rec.get('dataset') == 'MVTec AD2':
                    rec['dataset'] = 'MVTec AD 2'
                rows.append(rec)
    # Merge hand-curated sub-area seed rows (One-Class / Supervised / Graph AD)
    supp = PROJECT_DIR / 'ad_supplemental.json'
    if supp.exists():
        try:
            for r in json.loads(supp.read_text(encoding='utf-8')).get('rows', []):
                if r.get('method') and r.get('dataset'):
                    rows.append(r)
        except Exception as e:
            log(f'  ad_supplemental merge failed: {e}')
    AD_JSON.write_text(json.dumps(rows, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    return rows

def _extract_hyperlink_display(s):
    """=HYPERLINK("url", "display")  ->  "display".  Falls back to s unchanged."""
    if not isinstance(s, str): return s
    if not s.startswith('=HYPERLINK('): return s
    m = re.search(r'=HYPERLINK\("[^"]*"\s*,\s*"([^"]+)"\)', s)
    return m.group(1) if m else s

# ---- OD per-dataset primary metric (label, extraction strategy) ----
# Each dataset's primary ranking metric depends on the dataset.  For datasets
# that only report metric values inside the 'AP' string column (e.g. SOD's
# S-measure, Argoverse's sAP, Few-Shot novel AP), we parse the number with a
# regex.  For COCO-family datasets the primary number is already in the 'mAP'
# column.
def _to_float(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and (v != v)) else float(v)
    s = str(v).strip()
    m = re.search(r'-?\d+\.?\d*', s)
    return float(m.group(0)) if m else None

def _parse_metric_re(s, pattern):
    if not isinstance(s, str): return None
    m = re.search(pattern, s, re.IGNORECASE)
    return _to_float(m.group(1)) if m else None

# (label, kind, value-extractor, lower_is_better)
# kind: 'higher' = larger is better;  'lower' = smaller is better.
def _od_extract(row):
    ds = row.get('dataset', '')
    mAP_v = _to_float(row.get('mAP'))
    AP_str = str(row.get('AP') or '')

    # COCO family: primary in mAP col
    if ds in ('COCO test-dev','COCO 2017 val','COCO 2017','COCO minival'):
        return ('COCO AP', 'higher', mAP_v)
    if ds == 'COCO-O':
        return ('Effective robustness AP', 'higher', mAP_v)
    if ds == 'PASCAL VOC 2007':
        return ('mAP@0.5', 'higher', mAP_v)
    if ds in ('GraZPEDWRI-DX', 'CPPE-5'):
        return ('mAP@0.5', 'higher', mAP_v)
    if ds == 'CrowdHuman':
        return ('AP',  'higher', mAP_v)
    if ds == 'Waymo 2D':
        return ('Waymo 2D AP', 'higher', mAP_v)

    # Argoverse-HD: streaming AP from AP-string
    if ds.startswith('Argoverse-HD'):
        return ('Streaming AP (sAP)', 'higher',
                _parse_metric_re(AP_str, r'sAP[^\d]*([\d.]+)') or mAP_v)

    # SOD datasets: S-measure from AP-string
    if ds in ('DUTS-TE','DUT-OMRON','ECSSD','HKU-IS','PASCAL-S',
              'HRSOD','UHRSD','DAVIS-S'):
        return ('S-measure', 'higher',
                _parse_metric_re(AP_str, r'\bS[α]?\s*[:= ]?\s*([\d.]+)'))
    if ds in ('SBU-Refine','ISTD'):
        return ('MAE (lower better)', 'lower',
                _parse_metric_re(AP_str, r'\bMAE[^\d]*([\d.]+)'))
    if ds == 'CAMO-FS':
        return ('AP (CD-FSOD)', 'higher',
                _parse_metric_re(AP_str, r'\bAP[^\d]*([\d.]+)') or mAP_v)

    # Few-shot detection: novel AP from AP-string
    if ds in ('PASCAL VOC 2007 15+5',):
        return ('mAP / novel AP', 'higher',
                _parse_metric_re(AP_str, r'\b(?:nAP|novel AP|AP)[^\d]*([\d.]+)') or mAP_v)
    if ds in ('MS-COCO 1-shot','MS-COCO 5-shot','MS-COCO 10-shot','MS-COCO 30-shot','COCO 2017 FSOD'):
        return ('novel AP', 'higher',
                _parse_metric_re(AP_str, r'\b(?:nAP|novel AP|AP)[^\d]*([\d.]+)') or mAP_v)
    if ds in ('LVIS v1.0 val','LVIS v1.0 test-dev'):
        return ('AP / AP_rare', 'higher',
                _parse_metric_re(AP_str, r'\b(?:AP_?rare|AP)[^\d]*([\d.]+)') or mAP_v)
    if ds in ('ODinW-13','ODinW-35'):
        return ('AP (avg)', 'higher',
                _parse_metric_re(AP_str, r'\bAP[^\d]*([\d.]+)') or mAP_v)

    # Default: COCO-style mAP
    return ('mAP', 'higher', mAP_v)


def regenerate_od_json():
    import pandas as pd
    xl = pd.ExcelFile(OD_XLSX)
    # Index → category map.  pandas returns NaN for cells whose cached value
    # is empty (the "資料集 Sheet" cells contain =HYPERLINK formulas without
    # a cached display value), so we read this sheet via openpyxl directly.
    ds_cat = {}
    if 'Index' in xl.sheet_names:
        from openpyxl import load_workbook
        wb = load_workbook(OD_XLSX, read_only=True)
        sh = wb['Index']
        for r in range(2, sh.max_row + 1):
            cat   = sh.cell(row=r, column=1).value
            raw   = sh.cell(row=r, column=2).value
            sheet = _extract_hyperlink_display(raw) if isinstance(raw, str) else raw
            cat   = _clean(cat)
            sheet = _clean(sheet)
            if cat and sheet:
                ds_cat[sheet] = cat
        wb.close()
    # All papers
    all_papers = []
    if 'OD all papers' in xl.sheet_names:
        ap = pd.read_excel(xl, sheet_name='OD all papers')
        for _, row in ap.iterrows():
            rec = {
                'category': _clean(row.get('類別')),
                'method':   _clean(row.get('方法')),
                'authors':  _clean(row.get('作者')),
                'venue':    _clean(row.get('發表')),
                'date':     str(_clean(row.get('年月')) or ''),
                'notes':    _clean(row.get('備註(特色/based)')),
                'link':     _clean(row.get('連結')),
                'github':   _clean(row.get('GitHub')),
            }
            if rec['method']:
                all_papers.append(rec)
    # Dataset rows
    dataset_sheets = [s for s in xl.sheet_names if s not in ('OD all papers', 'Index')]
    rows = []
    for sn in dataset_sheets:
        df = pd.read_excel(xl, sheet_name=sn)
        if '方法' not in df.columns: continue
        for _, row in df.iterrows():
            rec = {
                'dataset': sn,
                'dataset_category': ds_cat.get(sn, ''),
                'category': _clean(row.get('類別')),
                'method':   _clean(row.get('方法')),
                'authors':  _clean(row.get('作者')),
                'venue':    _clean(row.get('發表')),
                'date':     str(_clean(row.get('年月')) or ''),
                'status':   _clean(row.get('狀態')),
                'mAP':      _clean(row.get('mAP')),
                'AP':       _clean(row.get('AP')),
                'FPS':      _clean(row.get('FPS')),
                'params':   _clean(row.get('params')),
                'notes':    _clean(row.get('備註(特色/based)')),
                'link':     _clean(row.get('連結')),
                'github':   _clean(row.get('GitHub')),
            }
            if rec['method']:
                # Attach dataset-specific primary metric for ranking + display
                label, kind, val = _od_extract(rec)
                rec['primary_label'] = label
                rec['primary_kind']  = kind
                rec['primary_value'] = val
                rows.append(rec)
    # Build per-dataset primary metric metadata (for See All headers / sorting)
    ds_primary = {}
    for r in rows:
        ds = r['dataset']
        if ds not in ds_primary:
            ds_primary[ds] = {'label': r.get('primary_label'), 'kind': r.get('primary_kind')}
    # Merge hand-curated sub-area seed rows (3D Object Detection)
    supp = PROJECT_DIR / 'od_supplemental.json'
    if supp.exists():
        try:
            for r in json.loads(supp.read_text(encoding='utf-8')).get('rows', []):
                if r.get('method') and r.get('dataset'):
                    rows.append(r)
                    if r['dataset'] not in dataset_sheets:
                        dataset_sheets.append(r['dataset'])
        except Exception as e:
            log(f'  od_supplemental merge failed: {e}')
    payload = {
        'datasets': dataset_sheets,
        'dataset_category_map': ds_cat,
        'dataset_primary_metric': ds_primary,
        'all_papers': all_papers,
        'rows': rows,
    }
    OD_JSON.write_text(json.dumps(payload, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    return payload

def reinject_html(ad_data, od_data, cls_data=None, as_data=None):
    html = INDEX_HTML.read_text(encoding='utf-8')
    ad_js = json.dumps(ad_data, ensure_ascii=False, separators=(',', ':'))
    od_js = json.dumps(od_data, ensure_ascii=False, separators=(',', ':'))
    new_html = re.sub(r'const AD_RAW = (\[[\s\S]*?\]);',
                      lambda _: 'const AD_RAW = ' + ad_js + ';', html, count=1)
    new_html = re.sub(r'const OD_DATA = (\{[\s\S]*?\});\s*\n',
                      lambda _: 'const OD_DATA = ' + od_js + ';\n', new_html, count=1)
    if cls_data is not None:
        cls_js = json.dumps(cls_data, ensure_ascii=False, separators=(',', ':'))
        new_html = re.sub(r'const CLS_DATA = (\{[\s\S]*?\});\s*\n',
                          lambda _: 'const CLS_DATA = ' + cls_js + ';\n', new_html, count=1)
    if as_data is not None:
        as_js  = json.dumps(as_data,  ensure_ascii=False, separators=(',', ':'))
        new_html = re.sub(r'const AS_DATA = (\{[\s\S]*?\});\s*\n',
                          lambda _: 'const AS_DATA = ' + as_js + ';\n', new_html, count=1)
    INDEX_HTML.write_text(new_html, encoding='utf-8')

def regenerate_cls_json():
    import pandas as pd
    if not CLS_XLSX.exists(): return None
    xl = pd.ExcelFile(CLS_XLSX)
    all_papers = []
    if 'Classification all papers' in xl.sheet_names:
        ap = pd.read_excel(xl, sheet_name='Classification all papers')
        for _, row in ap.iterrows():
            rec = {
                'category': _clean(row.get('類別')),
                'method':   _clean(row.get('方法')),
                'authors':  _clean(row.get('作者')),
                'venue':    _clean(row.get('發表')),
                'date':     str(_clean(row.get('年月')) or ''),
                'status':   _clean(row.get('狀態')),
                'top1':     _clean(row.get('Top-1')),
                'top5':     _clean(row.get('Top-5')),
                'acc':      _clean(row.get('Acc')),
                'f1':       _clean(row.get('F1')),
                'params':   _clean(row.get('Params')),
                'flops':    _clean(row.get('FLOPs')),
                'notes':    _clean(row.get('備註(特色/based)')),
                'link':     _clean(row.get('連結')),
                'github':   _clean(row.get('GitHub')),
            }
            if rec['method']: all_papers.append(rec)
    dataset_sheets = [s for s in xl.sheet_names if s != 'Classification all papers']
    rows = []
    for sn in dataset_sheets:
        df = pd.read_excel(xl, sheet_name=sn)
        if '方法' not in df.columns: continue
        for _, row in df.iterrows():
            rec = {
                'dataset':  sn,
                'category': _clean(row.get('類別')),
                'method':   _clean(row.get('方法')),
                'authors':  _clean(row.get('作者')),
                'venue':    _clean(row.get('發表')),
                'date':     str(_clean(row.get('年月')) or ''),
                'status':   _clean(row.get('狀態')),
                'top1':     _clean(row.get('Top-1')),
                'top5':     _clean(row.get('Top-5')),
                'acc':      _clean(row.get('Acc')),
                'f1':       _clean(row.get('F1')),
                'params':   _clean(row.get('Params')),
                'flops':    _clean(row.get('FLOPs')),
                'notes':    _clean(row.get('備註(特色/based)')),
                'link':     _clean(row.get('連結')),
                'github':   _clean(row.get('GitHub')),
            }
            if rec['method']: rows.append(rec)
    # Merge supplemental hand-curated entries (e.g. Semi-Supervised, Small-Data)
    supp_path = PROJECT_DIR / 'cls_supplemental.json'
    if supp_path.exists():
        try:
            supp = json.loads(supp_path.read_text(encoding='utf-8'))
            for r in supp.get('rows', []):
                if r.get('method') and r.get('dataset'):
                    rows.append(r)
                    # Also add a method-level entry to all_papers if not already present
                    seen = any(p.get('method') == r['method'] for p in all_papers)
                    if not seen:
                        all_papers.append({k: v for k, v in r.items() if k != 'dataset'})
        except Exception as e:
            log(f'  cls_supplemental merge failed: {e}')
    payload = {'datasets': dataset_sheets, 'all_papers': all_papers, 'rows': rows}
    CLS_JSON.write_text(json.dumps(payload, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    return payload

def regenerate_as_json():
    import pandas as pd
    if not AS_XLSX.exists(): return None
    xl = pd.ExcelFile(AS_XLSX)
    all_papers = []
    if 'AS all papers' in xl.sheet_names:
        ap = pd.read_excel(xl, sheet_name='AS all papers')
        for _, row in ap.iterrows():
            rec = {
                'category':      _clean(row.get('類別')),
                'method':        _clean(row.get('方法')),
                'authors':       _clean(row.get('作者')),
                'venue':         _clean(row.get('發表')),
                'date':          str(_clean(row.get('年月')) or ''),
                'status':        _clean(row.get('狀態')),
                'synth_type':    _clean(row.get('合成類型')),
                'task_setting':  _clean(row.get('任務設定')),
                'datasets_used': _clean(row.get('資料集')),
                'split':         _clean(row.get('Split/Protocol')),
                'backbone':      _clean(row.get('Backbone/Detector')),
                'notes':         _clean(row.get('備註(特色/可比性)')),
                'link':          _clean(row.get('連結')),
                'github':        _clean(row.get('GitHub')),
            }
            if rec['method']: all_papers.append(rec)
    dataset_sheets = [s for s in xl.sheet_names if s != 'AS all papers']
    rows = []
    for sn in dataset_sheets:
        df = pd.read_excel(xl, sheet_name=sn)
        if '方法' not in df.columns: continue
        for _, row in df.iterrows():
            rec = {
                'dataset':       sn,
                'category':      _clean(row.get('類別')),
                'method':        _clean(row.get('方法')),
                'authors':       _clean(row.get('作者')),
                'venue':         _clean(row.get('發表')),
                'date':          str(_clean(row.get('年月')) or ''),
                'status':        _clean(row.get('狀態')),
                'synth_type':    _clean(row.get('合成類型')),
                'task_setting':  _clean(row.get('任務設定')),
                'split':         _clean(row.get('Split/Protocol')),
                'backbone':      _clean(row.get('Backbone/Detector')),
                'notes':         _clean(row.get('備註(特色/可比性)')),
                'link':          _clean(row.get('連結')),
                'github':        _clean(row.get('GitHub')),
            }
            if rec['method']: rows.append(rec)
    payload = {'datasets': dataset_sheets, 'all_papers': all_papers, 'rows': rows}
    AS_JSON.write_text(json.dumps(payload, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    return payload

# ====================================================================
# Email notification
# ====================================================================
def load_env(path):
    """Read simple KEY=VALUE pairs from a .env file (no quoting / interpolation)."""
    cfg = {}
    if not path.exists():
        return cfg
    for line in path.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        cfg[k.strip()] = v.strip().strip('"').strip("'")
    return cfg

def render_email_body(new_ad, new_od, ran_at, pushed):
    """Plain-text body listing the new papers."""
    lines = []
    lines.append(f'NTUT 自動化檢測實驗室 — 文獻自動更新通知')
    lines.append(f'執行時間: {ran_at}')
    lines.append(f'本次新增: AD {len(new_ad)} 篇 / OD {len(new_od)} 篇')
    lines.append(f'GitHub 推送: {"成功" if pushed else "未推送"}')
    lines.append('')
    lines.append('=' * 64)
    lines.append('Anomaly Detection 新增論文')
    lines.append('=' * 64)
    if new_ad:
        for i, p in enumerate(new_ad, 1):
            lines.append(f'{i:>2}. [{p["dataset"]} | {p["category"]}]')
            lines.append(f'    {p["title"]}')
            if p.get('authors'):
                lines.append(f'    Authors : {p["authors"]}')
            lines.append(f'    arXiv   : {p["link"]}')
            lines.append(f'    Date    : {p["date"]}')
            lines.append('')
    else:
        lines.append('  (本次無新增)')
        lines.append('')

    lines.append('=' * 64)
    lines.append('Object Detection 新增論文')
    lines.append('=' * 64)
    if new_od:
        for i, p in enumerate(new_od, 1):
            lines.append(f'{i:>2}. [{p["dataset"]} | {p["category"]}]')
            lines.append(f'    {p["title"]}')
            if p.get('authors'):
                lines.append(f'    Authors : {p["authors"]}')
            lines.append(f'    arXiv   : {p["link"]}')
            lines.append(f'    Date    : {p["date"]}')
            lines.append('')
    else:
        lines.append('  (本次無新增)')
        lines.append('')

    lines.append('-' * 64)
    lines.append('提示: 自動分類為 best-effort 啟發式, 指標數字 (AUROC / mAP 等)')
    lines.append('需手動填入。請參閱 scripts/update_log.txt 確認分類正確。')
    lines.append('網站: https://112378074.github.io/NTUT-IVI-Lab--SOTA-Literature-Viewer/')
    return '\n'.join(lines)

def render_email_html(new_ad, new_od, ran_at, pushed):
    def section(title, papers, cat_color):
        if not papers:
            return f'<h3 style="color:#0f172a">{title}</h3><p style="color:#64748b">本次無新增</p>'
        rows = ''
        for i, p in enumerate(papers, 1):
            rows += (f'<tr style="border-top:1px solid #e2e8f0">'
                     f'<td style="padding:8px;color:#64748b;width:32px;text-align:right">{i}</td>'
                     f'<td style="padding:8px">'
                     f'<div style="font-weight:600;color:#0f172a">{p["title"]}</div>'
                     f'<div style="color:#475569;font-size:13px;margin-top:2px">'
                     f'<span style="background:{cat_color};color:#fff;padding:1px 7px;border-radius:10px;font-size:11px;margin-right:6px">{p["category"]}</span>'
                     f'<span style="background:#e2e8f0;color:#334155;padding:1px 7px;border-radius:10px;font-size:11px;margin-right:6px">{p["dataset"]}</span>'
                     f'<span style="color:#64748b">{p["date"]}</span></div>'
                     f'<div style="color:#64748b;font-size:12px;margin-top:4px">'
                     f'{(p.get("authors") or "")[:120]}</div>'
                     f'<div style="margin-top:4px"><a href="{p["link"]}" style="color:#2563eb;font-size:12px">arXiv ↗</a></div>'
                     f'</td></tr>')
        return (f'<h3 style="color:#0f172a;margin-bottom:8px">{title}</h3>'
                f'<table style="width:100%;border-collapse:collapse;background:#fff;'
                f'border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">{rows}</table>')

    push_pill = ('<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:10px;font-size:12px">已推送 GitHub</span>'
                 if pushed
                 else '<span style="background:#fef3c7;color:#854d0e;padding:2px 8px;border-radius:10px;font-size:12px">未推送</span>')
    html = f'''<!doctype html><html><body style="font-family:-apple-system,Segoe UI,Arial,sans-serif;background:#f6f8fc;padding:20px;color:#1a2233">
<div style="max-width:780px;margin:0 auto">
  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-bottom:16px">
    <div style="font-size:12px;color:#2563eb;font-weight:600;margin-bottom:8px">NTUT · IEM · Automated Visual Inspection System Lab</div>
    <h2 style="margin:0 0 6px;color:#0f172a">文獻自動更新通知</h2>
    <div style="color:#64748b;font-size:13px">執行時間 {ran_at} · 新增 AD {len(new_ad)} 篇 / OD {len(new_od)} 篇 · {push_pill}</div>
  </div>
  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-bottom:16px">
    {section("Anomaly Detection 新增論文", new_ad, "#2563eb")}
  </div>
  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-bottom:16px">
    {section("Object Detection 新增論文", new_od, "#7c3aed")}
  </div>
  <div style="color:#64748b;font-size:12px;text-align:center;line-height:1.6;padding:8px 0 16px">
    自動分類為 best-effort 啟發式; 指標數字需手動填入。<br>
    <a href="https://112378074.github.io/NTUT-IVI-Lab--SOTA-Literature-Viewer/" style="color:#2563eb">前往網站</a>
  </div>
</div></body></html>'''
    return html

def render_email_body_v2(new_ad, new_od, new_cls, new_as, ran_at, pushed):
    lines = []
    lines.append('NTUT 自動化檢測實驗室 — 文獻自動更新通知')
    lines.append(f'執行時間: {ran_at}')
    lines.append(f'本次新增: AD {len(new_ad)} / OD {len(new_od)} / CLS {len(new_cls)} / AS {len(new_as)}')
    lines.append(f'GitHub 推送: {"成功" if pushed else "未推送"}')
    def section(label, papers):
        lines.append(''); lines.append('=' * 60); lines.append(label); lines.append('=' * 60)
        if not papers:
            lines.append('  (本次無新增)'); return
        for i, p in enumerate(papers, 1):
            lines.append(f'{i:>2}. [{p["dataset"]} | {p["category"]}]')
            lines.append(f'    {p["title"]}')
            if p.get('authors'): lines.append(f'    Authors : {p["authors"]}')
            lines.append(f'    arXiv   : {p["link"]}')
            lines.append(f'    Date    : {p["date"]}')
            lines.append('')
    section('Anomaly Detection 新增論文',  new_ad)
    section('Object Detection 新增論文',   new_od)
    section('Classification 新增論文',     new_cls)
    section('Anomaly Synthesis 新增論文',  new_as)
    lines.append('-' * 60)
    lines.append('提示: 自動分類為 best-effort 啟發式; 新增列的指標皆留空待人工核對.')
    lines.append('網站: https://112378074.github.io/NTUT-IVI-Lab--SOTA-Literature-Viewer/')
    return '\n'.join(lines)

def render_email_html_v2(new_ad, new_od, new_cls, new_as, ran_at, pushed):
    def section(title, papers, color):
        if not papers:
            return f'<h3 style="color:#0f172a;margin:18px 0 8px">{title}</h3><p style="color:#64748b;font-size:13px">本次無新增</p>'
        rows = ''
        for i, p in enumerate(papers, 1):
            rows += (f'<tr style="border-top:1px solid #e2e8f0">'
                     f'<td style="padding:8px;color:#64748b;width:32px;text-align:right">{i}</td>'
                     f'<td style="padding:8px">'
                     f'<div style="font-weight:600;color:#0f172a">{p["title"]}</div>'
                     f'<div style="color:#475569;font-size:13px;margin-top:2px">'
                     f'<span style="background:{color};color:#fff;padding:1px 7px;border-radius:10px;font-size:11px;margin-right:6px">{p["category"][:36]}</span>'
                     f'<span style="background:#e2e8f0;color:#334155;padding:1px 7px;border-radius:10px;font-size:11px;margin-right:6px">{p["dataset"]}</span>'
                     f'<span style="color:#64748b">{p["date"]}</span></div>'
                     f'<div style="color:#64748b;font-size:12px;margin-top:4px">{(p.get("authors") or "")[:120]}</div>'
                     f'<div style="margin-top:4px"><a href="{p["link"]}" style="color:#2563eb;font-size:12px">arXiv ↗</a></div>'
                     f'</td></tr>')
        return (f'<h3 style="color:#0f172a;margin:18px 0 8px">{title}</h3>'
                f'<table style="width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">{rows}</table>')

    push_pill = ('<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:10px;font-size:12px">已推送 GitHub</span>'
                 if pushed
                 else '<span style="background:#fef3c7;color:#854d0e;padding:2px 8px;border-radius:10px;font-size:12px">未推送</span>')
    return ('<!doctype html><html><body style="font-family:-apple-system,Segoe UI,Arial,sans-serif;background:#f6f8fc;padding:20px;color:#1a2233">'
            '<div style="max-width:780px;margin:0 auto">'
            '<div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-bottom:16px">'
            '<div style="font-size:12px;color:#2563eb;font-weight:600;margin-bottom:8px">NTUT · IEM · Automated Visual Inspection System Lab</div>'
            '<h2 style="margin:0 0 6px;color:#0f172a">文獻自動更新通知</h2>'
            f'<div style="color:#64748b;font-size:13px">執行時間 {ran_at}</div>'
            f'<div style="color:#64748b;font-size:13px;margin-top:6px">本次新增 AD {len(new_ad)} / OD {len(new_od)} / CLS {len(new_cls)} / AS {len(new_as)} · {push_pill}</div>'
            '</div>'
            '<div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-bottom:16px">'
            f'{section("Anomaly Detection",  new_ad,  "#2563eb")}'
            f'{section("Object Detection",   new_od,  "#7c3aed")}'
            f'{section("Classification",     new_cls, "#15803d")}'
            f'{section("Anomaly Synthesis",  new_as,  "#ea580c")}'
            '</div>'
            '<div style="color:#64748b;font-size:12px;text-align:center;line-height:1.6;padding:8px 0 16px">'
            '新增列的指標皆留空待人工核對 (避免錯誤數字).<br>'
            '<a href="https://112378074.github.io/NTUT-IVI-Lab--SOTA-Literature-Viewer/" style="color:#2563eb">前往網站</a>'
            '</div></div></body></html>')

def send_notification(new_ad, new_od, ran_at, pushed, new_cls=None, new_as=None):
    new_cls = new_cls or []
    new_as  = new_as  or []
    cfg = load_env(ENV_FILE)
    host = cfg.get('SMTP_HOST', 'smtp.gmail.com')
    port = int(cfg.get('SMTP_PORT', '587'))
    user = cfg.get('SMTP_USER', '')
    pwd  = cfg.get('SMTP_PASSWORD', '')
    to   = cfg.get('NOTIFY_TO', 'azaz31855@gmail.com')

    if not user or not pwd:
        log(f'  email skipped: SMTP_USER / SMTP_PASSWORD not set in {ENV_FILE}')
        return False

    subject = (f'[AIL Auto-Update] +{len(new_ad)} AD / +{len(new_od)} OD / '
               f'+{len(new_cls)} CLS / +{len(new_as)} AS ({datetime.now().strftime("%Y-%m-%d")})')
    text = render_email_body_v2(new_ad, new_od, new_cls, new_as, ran_at, pushed)
    html = render_email_html_v2(new_ad, new_od, new_cls, new_as, ran_at, pushed)

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = user
    msg['To']      = to
    msg['Date']    = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid()
    msg.attach(MIMEText(text, 'plain', 'utf-8'))
    msg.attach(MIMEText(html, 'html',  'utf-8'))

    try:
        with smtplib.SMTP(host, port, timeout=30) as s:
            s.ehlo()
            s.starttls(context=SSL_CTX)
            s.ehlo()
            s.login(user, pwd)
            s.sendmail(user, [to], msg.as_string())
        log(f'  email sent to {to}')
        return True
    except Exception as e:
        log(f'  email send failed: {type(e).__name__}: {e}')
        return False

# ====================================================================
# Git push
# ====================================================================
def git_push(message):
    try:
        cwd = str(PROJECT_DIR)
        files = [INDEX_HTML.name, AD_XLSX.name, OD_XLSX.name, CLS_XLSX.name, AS_XLSX.name,
                 AD_JSON.name, OD_JSON.name, CLS_JSON.name, AS_JSON.name]
        # Filter to existing files (some xlsx may not be present)
        files = [f for f in files if (PROJECT_DIR / f).exists()]
        subprocess.run(['git', 'add'] + files, cwd=cwd, check=True)
        # Check if anything to commit
        diff = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=cwd)
        if diff.returncode == 0:
            log('  no changes to commit')
            return False
        subprocess.run(['git', 'commit', '-m', message], cwd=cwd, check=True)
        subprocess.run(['git', 'push'], cwd=cwd, check=True)
        log('  git push completed')
        return True
    except subprocess.CalledProcessError as e:
        log(f'  git operation failed: {e}')
        return False

# ====================================================================
# Main
# ====================================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry', action='store_true', help='Fetch and classify only; no writes')
    parser.add_argument('--no-push', action='store_true', help='Skip git push')
    parser.add_argument('--no-email', action='store_true', help='Skip email notification')
    args = parser.parse_args()
    ran_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    log('=' * 60)
    log(f'Run start (mode={"dry" if args.dry else "full"})')
    log(f'Project: {PROJECT_DIR}')

    # 1. Existing IDs
    ad_ids,  _ = collect_existing_arxiv_ids(AD_XLSX)
    od_ids,  _ = collect_existing_arxiv_ids(OD_XLSX)
    cls_ids, _ = collect_existing_arxiv_ids(CLS_XLSX)
    as_ids,  _ = collect_existing_arxiv_ids(AS_XLSX)
    log(f'Existing IDs — AD: {len(ad_ids)} | OD: {len(od_ids)} | CLS: {len(cls_ids)} | AS: {len(as_ids)}')

    # 2. Fetch + classify AD — require target dataset to be mentioned (avoid noise)
    new_ad = []
    for q in AD_QUERIES:
        root = fetch_arxiv(q)
        for e in parse_entries(root):
            if not e['arxiv_id']:                continue
            if e['arxiv_id'] in ad_ids:          continue
            if not is_within_window(e['published'], LOOKBACK_DAYS): continue
            text = e['title'] + ' ' + e['summary']
            if is_ad_excluded(text):
                log(f'  AD - skipped (excluded topic): {e["title"][:80]}')
                continue
            ds = detect_dataset(text, AD_DATASETS_BY_PATTERNS)
            if not ds:
                # No target dataset mentioned -> probably off-topic
                log(f'  AD - skipped (no target dataset): {e["title"][:80]}')
                continue
            cat = classify_category(text, AD_CATEGORIES_BY_KEYWORDS) or AD_DEFAULT_CATEGORY
            method = e['title'][:90]
            ad_ids.add(e['arxiv_id'])
            new_ad.append({**e, 'category': cat, 'dataset': ds, 'method': method,
                           'note': f'Auto-fetched ({e["date"]}); 待人工核對指標'})
        time.sleep(5)
    log(f'New AD candidates: {len(new_ad)}')
    for p in new_ad:
        log(f'  AD + [{p["dataset"]}|{p["category"]}] {p["title"][:80]}')

    # 3. Fetch + classify OD — require target dataset; exclude 3D/medical/etc.
    new_od = []
    for q in OD_QUERIES:
        root = fetch_arxiv(q)
        for e in parse_entries(root):
            if not e['arxiv_id']:                continue
            if e['arxiv_id'] in od_ids:          continue
            if not is_within_window(e['published'], LOOKBACK_DAYS): continue
            text = e['title'] + ' ' + e['summary']
            if is_od_excluded(text):
                log(f'  OD - skipped (excluded topic): {e["title"][:80]}')
                continue
            ds = detect_dataset(text, OD_DATASETS_BY_PATTERNS)
            if not ds:
                log(f'  OD - skipped (no target dataset): {e["title"][:80]}')
                continue
            cat = classify_category(text, OD_CATEGORIES_BY_KEYWORDS) or OD_DEFAULT_CATEGORY
            method = e['title'][:90]
            od_ids.add(e['arxiv_id'])
            new_od.append({**e, 'category': cat, 'dataset': ds, 'method': method,
                           'note': 'Auto-fetched from arXiv; metrics To verify'})
        time.sleep(5)
    log(f'New OD candidates: {len(new_od)}')
    for p in new_od:
        log(f'  OD + [{p["dataset"]}|{p["category"]}] {p["title"][:80]}')

    # 4. Fetch + classify CLS
    new_cls = []
    for q in CLS_QUERIES:
        root = fetch_arxiv(q)
        for e in parse_entries(root):
            if not e['arxiv_id']:                continue
            if e['arxiv_id'] in cls_ids:         continue
            if not is_within_window(e['published'], LOOKBACK_DAYS): continue
            text = e['title'] + ' ' + e['summary']
            if is_cls_excluded(text):
                log(f'  CLS - skipped (excluded topic): {e["title"][:80]}')
                continue
            ds = detect_dataset(text, CLS_DATASETS_BY_PATTERNS)
            if not ds:
                log(f'  CLS - skipped (no target dataset): {e["title"][:80]}')
                continue
            cat = classify_category(text, CLS_CATEGORIES_BY_KEYWORDS) or CLS_DEFAULT_CATEGORY
            method = e['title'][:90]
            cls_ids.add(e['arxiv_id'])
            new_cls.append({**e, 'category': cat, 'dataset': ds, 'method': method,
                            'note': 'Auto-fetched from arXiv; metrics To verify'})
        time.sleep(5)
    log(f'New CLS candidates: {len(new_cls)}')
    for p in new_cls:
        log(f'  CLS + [{p["dataset"]}|{p["category"]}] {p["title"][:80]}')

    # 5. Fetch + classify AS
    new_as = []
    for q in AS_QUERIES:
        root = fetch_arxiv(q)
        for e in parse_entries(root):
            if not e['arxiv_id']:                continue
            if e['arxiv_id'] in as_ids:          continue
            if not is_within_window(e['published'], LOOKBACK_DAYS): continue
            text = e['title'] + ' ' + e['summary']
            if is_as_excluded(text):
                log(f'  AS - skipped (excluded topic): {e["title"][:80]}')
                continue
            ds = detect_dataset(text, AS_DATASETS_BY_PATTERNS) or AS_DEFAULT_DATASET
            cat = classify_category(text, AS_CATEGORIES_BY_KEYWORDS) or AS_DEFAULT_CATEGORY
            method = e['title'][:90]
            as_ids.add(e['arxiv_id'])
            new_as.append({**e, 'category': cat, 'dataset': ds, 'method': method,
                           'note': 'Auto-fetched from arXiv; pending verification'})
        time.sleep(5)
    log(f'New AS candidates: {len(new_as)}')
    for p in new_as:
        log(f'  AS + [{p["dataset"]}|{p["category"]}] {p["title"][:80]}')

    # 3b. Scan CVF Open Access (CVPR/ICCV/ECCV/WACV proceedings)
    new_cvf = []
    try:
        new_cvf = process_cvf()
    except Exception as e:
        log(f'  CVF scan failed: {e}')

    if args.dry:
        log(f'Dry run complete; no files written (CVF would add {len(new_cvf)})')
        return 0

    if not new_ad and not new_od and not new_cls and not new_as and not new_cvf:
        log('No new papers found — nothing to update')
        if not args.no_email:
            send_notification([], [], ran_at, pushed=False, new_cls=[], new_as=[])
        return 0

    # 4. Append AD rows
    if new_ad:
        wb = load_workbook(AD_XLSX)
        for p in new_ad:
            sheet = p['dataset']
            if sheet == 'MVTec AD 2': sheet = 'MVTec AD2'
            ok = append_ad_row(wb, sheet, p)
            if not ok:
                log(f'  AD: skipped (sheet {sheet} not found): {p["title"][:60]}')
        wb.save(AD_XLSX)
        log(f'AD xlsx updated: +{len(new_ad)} rows')

    # 5. Append OD rows
    if new_od:
        wb = load_workbook(OD_XLSX)
        for p in new_od:
            append_od_dataset_row(wb, p['dataset'], p)
            append_od_all_papers_row(wb, p)
        wb.save(OD_XLSX)
        log(f'OD xlsx updated: +{len(new_od)} rows')

    # 6. Append CLS rows
    if new_cls and CLS_XLSX.exists():
        wb = load_workbook(CLS_XLSX)
        for p in new_cls:
            append_cls_dataset_row(wb, p['dataset'], p)
            append_cls_all_papers_row(wb, p)
        wb.save(CLS_XLSX)
        log(f'CLS xlsx updated: +{len(new_cls)} rows')

    # 7. Append AS rows
    if new_as and AS_XLSX.exists():
        wb = load_workbook(AS_XLSX)
        for p in new_as:
            append_as_dataset_row(wb, p['dataset'], p)
            append_as_all_papers_row(wb, p)
        wb.save(AS_XLSX)
        log(f'AS xlsx updated: +{len(new_as)} rows')

    # 7b. Append CVF Open Access papers to the All Papers sheets
    if new_cvf:
        try:
            append_cvf_rows(new_cvf)
        except Exception as e:
            log(f'  CVF append failed: {e}')

    # 8. Regenerate JSON & inject HTML
    ad_data  = regenerate_ad_json()
    od_data  = regenerate_od_json()
    cls_data = regenerate_cls_json()
    as_data  = regenerate_as_json()
    reinject_html(ad_data, od_data, cls_data, as_data)
    log(f'Regenerated: AD {len(ad_data)} | OD {len(od_data["all_papers"])} ({len(od_data["rows"])} rows) '
        f'| CLS {len(cls_data["all_papers"]) if cls_data else 0} | AS {len(as_data["all_papers"]) if as_data else 0}')

    # 9. Git push
    pushed = False
    if not args.no_push:
        cvf_n = len(new_cvf)
        msg = f'auto: +{len(new_ad)} AD / +{len(new_od)} OD / +{len(new_cls)} CLS / +{len(new_as)} AS / +{cvf_n} CVF papers ({datetime.now().strftime("%Y-%m-%d")})'
        pushed = git_push(msg)

    # 10. Email notification
    if not args.no_email:
        send_notification(new_ad, new_od, ran_at, pushed, new_cls=new_cls, new_as=new_as)

    log('Run complete')
    return 0

if __name__ == '__main__':
    try:
        sys.exit(main())
    except Exception as e:
        log(f'FATAL: {type(e).__name__}: {e}')
        import traceback; log(traceback.format_exc())
        sys.exit(1)
