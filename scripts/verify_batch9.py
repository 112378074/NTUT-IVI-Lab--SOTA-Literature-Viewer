#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_batch9.py — 4 fills + 10 removes."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

FILLS = [
    # AST @ MVTec AD (Tables 2, 4) — image 99.2, pixel 95.0
    ('MVTec AD', 93, 99.2, 95.0, None, None,
     '✅ 已驗證 (AST WACV23 Table 2+4)：Asymmetric Student-Teacher (RGB+3D NF teacher)；MVTec AD I=99.2 P=95.0 [Student-Teacher]'),
    # VQ-Flow @ MVTec AD (Table I avg, ConvNeXt backbone, multi-class)
    ('MVTec AD', 23, 99.5, 98.3, None, None,
     '✅ 已驗證 (VQ-Flow Table I multi-class avg, ConvNeXt)：Hierarchical Vector Quantization Flow；99.5/98.3 @ 33 FPS [基於 NF]'),
    # VQ-Flow @ VisA (Table IV mention multi-class)
    ('VisA', 70, 95.9, 98.9, None, None,
     '✅ 已驗證 (VQ-Flow Table IV multi-class)：VisA I=95.9 P=98.9 [基於 NF]'),
    # ReContrast @ MVTec LOCO (Table A7 mean)
    ('MVTec LOCO', 18, 82.1, None, None, None,
     '✅ 已驗證 (ReContrast NeurIPS23 Table A7)：Mean I-AUROC=82.1 (struct 90.7 + logic 73.4)；非專為 LOCO 設計但優於多 baseline [Student-Teacher 對比]'),
]

DELETES = [
    # TVD ×6 — paper is about Phys-AD (video anomaly), not MVTec/VisA/MPDD/BTAD/3D/LOCO
    ('MVTec AD',   104, 'TVD paper proposes Phys-AD dataset (video-level), not tested on standard industrial AD'),
    ('VisA',        77, 'TVD = Phys-AD video paper, not VisA'),
    ('MPDD',        26, 'TVD = Phys-AD video paper, not MPDD'),
    ('BTAD',        32, 'TVD = Phys-AD video paper, not BTAD'),
    ('MVTec 3D',    24, 'TVD = Phys-AD video paper, not MVTec 3D'),
    ('MVTec LOCO',  24, 'TVD = Phys-AD video paper, not MVTec LOCO'),
    # CFM / M3DM / M3DM-NR / EasyNet @ MVTec AD — all multi-modal RGB+3D methods for MVTec 3D-AD
    ('MVTec AD',    89, 'CFM (CVPR24) only evaluates MVTec 3D-AD (multi-modal RGB+3D)'),
    ('MVTec AD',    91, 'M3DM (CVPR23) is multi-modal RGB+3D for MVTec 3D-AD'),
    ('MVTec AD',    90, 'M3DM-NR is the noisy-resistant version of M3DM, also MVTec 3D-AD only'),
    ('MVTec AD',    94, 'EasyNet primarily evaluates MVTec 3D-AD + Eyecandies (RGB-D)'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> i={i}, p={p}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason[:60]}...)')

    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))

if __name__ == '__main__':
    main()
