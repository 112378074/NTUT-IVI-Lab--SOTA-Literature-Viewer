#!/usr/bin/env python3
"""Add ECCV 2026 + recent arXiv 2026 AD papers to the xlsx database.

Run from repo root:
    python scripts/_add_eccv2026_papers.py
"""
import sys, os
from pathlib import Path
import openpyxl

PROJECT_DIR = Path(__file__).resolve().parent.parent
AD_XLSX = PROJECT_DIR / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# ─── Paper definitions ────────────────────────────────────────────────────────
# Each entry: (dataset, category, method, authors, venue, year_month, status,
#              I_AUROC, P_AUROC, P_AP, P_PRO, FPS, note, link, github)
# Columns 1-15 matching the xlsx schema.

NEW_PAPERS = [
    # ── ArcAD (ECCV 2026) – MVTec AD ─────────────────────────────────────────
    (
        'MVTec AD, VisA, Real-IAD, MANTA',
        '基於表徵 (Representation)',
        'ArcAD (Anomaly-Rectified Cold-Start)',
        'Anonymous (LGC-AD team)',
        'ECCV 2026 / arXiv 2607.02252',
        '2026-07',
        '已發表',
        '99.7 / 98.9',   # MVTec/VisA
        '99.2 / 99.0',
        'N/A',
        'N/A',
        'N/A',
        ('⚠ 非標準協議 (Supervised cold-start) — 需提供少量正常樣本 + 少量異常樣本(含像素標注)。'
         '✅ Supervised/cold-start verified — GitHub README LGC-AD/ArcAD; '
         'MVTec AD: I-AUROC=99.7, P-AUROC=99.2; VisA: I-AUROC=98.9, P-AUROC=99.0; '
         'F1-max reported (68.9/54.9) instead of P-AP; AU-PRO not reported. '
         'NOT comparable to standard unsupervised AD. [基於表徵 + Supervised]'),
        'https://arxiv.org/abs/2607.02252',
        'https://github.com/LGC-AD/ArcAD',
    ),
    # ── CMDS-AD (ECCV 2026) – MVTec 3D-AD few-shot ───────────────────────────
    (
        'MVTec 3D-AD, EyeCandies (few-shot)',
        '基於 3D/RGB-D (3D/RGB-D)',
        'CMDS-AD (Cross-Modal Dual-Stream)',
        'Anonymous (arXiv 2606.20300)',
        'ECCV 2026 / arXiv 2606.20300',
        '2026-06',
        '已發表',
        'N/A',
        'N/A',
        'N/A',
        'N/A',
        'N/A',
        ('⏳ Pending absolute metrics — 1-shot: +5.7% I-AUROC, +2.0% AUPRO vs. prior SOTA on MVTec 3D-AD; '
         '+7.7%, +5.6% on EyeCandies; absolute numbers not accessible without arXiv HTML. '
         'Protocol: 1–4 shot anomaly detection with cross-modal (RGB+depth) dual-stream decoupling. '
         '[基於 3D/RGB-D + Few-shot]'),
        'https://arxiv.org/abs/2606.20300',
        '',
    ),
    # ── DPDiff-AD (arXiv 2605.24402) – standard unsupervised multi-class ──────
    (
        'MVTec AD, VisA',
        '基於擴散 (Diffusion)',
        'DPDiff-AD (Dual Prototype-Conditioned Diffusion)',
        'Yaoxuan Feng, Yuxin Li, Weijiang Lv, Zixuan Zhao, Yubiao Wang, Wenchao Chen, Bo Chen, Hongwei Liu',
        'arXiv 2605.24402 (2026-05)',
        '2026-05',
        '預印本',
        '99.5',
        '98.7',
        '65.6',
        '95.9',
        '101',
        ('✅ Standard verified (indirect) — metrics extracted from arXiv HTML 2605.24402 via web search; '
         'MVTec AD multi-class unsupervised: I-AUROC=99.5, P-AUROC=98.7, P-AP=65.6, AU-PRO=95.9; '
         'inference 101 FPS; VisA: I-AUROC>ReContrast by +2.2pts (absolute not retrieved); '
         'SOTA claim on large-category (160-cat) benchmark (+5.3 I-AUROC, +2.9 P-AUROC vs. Dinomaly+); '
         'code not yet released. [基於擴散]'),
        'https://arxiv.org/abs/2605.24402',
        '',
    ),
    # ── MATCH (ECCV 2026) – multi-view flow matching on Real-IAD + MANTA ──────
    (
        'Real-IAD, MANTA',
        '基於 NF (Normalizing Flow)',
        'MATCH (Flow Matching Multi-View)',
        'Mathis Kruse, Melissa Schween, Bodo Rosenhahn',
        'ECCV 2026 / arXiv 2606.24375',
        '2026-06',
        '已發表',
        'N/A',
        'N/A',
        'N/A',
        'N/A',
        'N/A',
        ('⏳ Pending absolute metrics — first multi-view AD method using Flow Matching; '
         'evaluates on Real-IAD and MANTA-Tiny; claims SOTA on both; '
         'absolute metric table not accessible without arXiv HTML. '
         '[基於 NF / Flow Matching + Multi-view]'),
        'https://arxiv.org/abs/2606.24375',
        '',
    ),
]

# Separate rows per-dataset for dataset-specific sheets
DATASET_ROWS = [
    # ArcAD MVTec AD
    {
        'sheet': 'MVTec AD',
        'row': (
            'MVTec AD',
            '基於表徵 (Representation)',
            'ArcAD (Anomaly-Rectified Cold-Start)',
            'Anonymous (LGC-AD team)',
            'ECCV 2026 / arXiv 2607.02252',
            '2026-07',
            '已發表',
            '99.7',
            '99.2',
            'N/A',
            'N/A',
            'N/A',
            ('⚠ 非標準 (Supervised cold-start) — NOT comparable to unsupervised AD. '
             '✅ Supervised/cold-start verified — GitHub README LGC-AD/ArcAD; '
             'MVTec-AD: I-AUROC=99.7, P-AUROC=99.2, F1-max=68.9; AU-PRO not reported. '
             '[基於表徵 + Supervised cold-start]'),
            'https://arxiv.org/abs/2607.02252',
            'https://github.com/LGC-AD/ArcAD',
        ),
    },
    # ArcAD VisA
    {
        'sheet': 'VisA',
        'row': (
            'VisA',
            '基於表徵 (Representation)',
            'ArcAD (Anomaly-Rectified Cold-Start)',
            'Anonymous (LGC-AD team)',
            'ECCV 2026 / arXiv 2607.02252',
            '2026-07',
            '已發表',
            '98.9',
            '99.0',
            'N/A',
            'N/A',
            'N/A',
            ('⚠ 非標準 (Supervised cold-start) — NOT comparable to unsupervised AD. '
             '✅ Supervised/cold-start verified — GitHub README LGC-AD/ArcAD; '
             'VisA: I-AUROC=98.9, P-AUROC=99.0, F1-max=54.9; AU-PRO not reported. '
             '[基於表徵 + Supervised cold-start]'),
            'https://arxiv.org/abs/2607.02252',
            'https://github.com/LGC-AD/ArcAD',
        ),
    },
    # CMDS-AD MVTec 3D
    {
        'sheet': 'MVTec 3D',
        'row': (
            'MVTec 3D-AD',
            '基於 3D/RGB-D (3D/RGB-D)',
            'CMDS-AD (Cross-Modal Dual-Stream)',
            'Anonymous (arXiv 2606.20300)',
            'ECCV 2026 / arXiv 2606.20300',
            '2026-06',
            '已發表',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            ('⏳ Pending absolute metrics — few-shot (1-4 shot) protocol; '
             '1-shot: +5.7% I-AUROC, +2.0% AUPRO vs. prior SOTA on MVTec 3D-AD; '
             'absolute table not accessible without arXiv HTML. '
             '[基於 3D/RGB-D + Few-shot]'),
            'https://arxiv.org/abs/2606.20300',
            '',
        ),
    },
    # DPDiff-AD MVTec AD
    {
        'sheet': 'MVTec AD',
        'row': (
            'MVTec AD',
            '基於擴散 (Diffusion)',
            'DPDiff-AD (Dual Prototype-Conditioned Diffusion)',
            'Yaoxuan Feng et al.',
            'arXiv 2605.24402 (2026-05)',
            '2026-05',
            '預印本',
            '99.5',
            '98.7',
            '65.6',
            '95.9',
            '101',
            ('✅ Standard verified (indirect via web search from arXiv HTML 2605.24402); '
             'MVTec AD multi-class unsupervised: I-AUROC=99.5, P-AUROC=98.7, P-AP=65.6, AU-PRO=95.9; '
             '101 FPS; code not yet released. [基於擴散]'),
            'https://arxiv.org/abs/2605.24402',
            '',
        ),
    },
]


def append_row(ws, row_data):
    """Append a tuple as one row, skipping the last empty trailing cells."""
    ws.append(list(row_data))


def paper_exists(ws, method_name):
    """Return True if a row with this method name already exists in the sheet."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] and method_name.lower() in str(row[2]).lower():
            return True
    return False


def main():
    wb = openpyxl.load_workbook(AD_XLSX)
    ws_all = wb['總覽 All Papers']

    added = []

    # ── Add to 總覽 ─────────────────────────────────────────────────────────
    for paper in NEW_PAPERS:
        method_name = paper[2]
        if paper_exists(ws_all, method_name.split('(')[0].strip()):
            print(f'SKIP (already in 總覽): {method_name}')
            continue
        append_row(ws_all, paper)
        added.append(('總覽', method_name))
        print(f'ADDED to 總覽: {method_name}')

    # ── Add to dataset-specific sheets ─────────────────────────────────────
    for entry in DATASET_ROWS:
        sheet_name = entry['sheet']
        row = entry['row']
        method_name = row[2]
        if sheet_name not in wb.sheetnames:
            print(f'SKIP (no sheet): {sheet_name} for {method_name}')
            continue
        ws = wb[sheet_name]
        if paper_exists(ws, method_name.split('(')[0].strip()):
            print(f'SKIP (already in {sheet_name}): {method_name}')
            continue
        append_row(ws, row)
        added.append((sheet_name, method_name))
        print(f'ADDED to {sheet_name}: {method_name}')

    if not added:
        print('Nothing new to add.')
        return

    wb.save(AD_XLSX)
    print(f'\nSaved {AD_XLSX}')
    print(f'Rows added: {len(added)}')
    for sheet, method in added:
        print(f'  [{sheet}] {method}')


if __name__ == '__main__':
    main()
