#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_sod2.py
Fill per-dataset S-measure values for SOD methods across DUTS-TE, DUT-OMRON,
ECSSD, HKU-IS, PASCAL-S from paper-verified sources.

Sources:
  - M3Net Table I (arxiv 2309.08365 PDF, pages 5-7): ICON-S, M3Net-S, MINet-R
  - PoolNet+ TPAMI22 Table 5 (PDF pages 11-12): PoolNet-R+
  - Samba+ arxiv 2602.01593 Table I (HTML): EDN22, Samba values (confirmed column order)
  - U²-Net PR 2020 paper Table 2: widely cited values

Also updates flags for:
  - MVANet: DIS method (DIS-5K only) — NOT a SOD method
  - ICON-S/M3Net/MINet/PoolNet+/EDN/U²-Net in HR sheets: not evaluated in paper
  - S3OD in ECSSD/HKU-IS: not in S3OD paper (evaluates DIS/HR-SOD, not ECSSD/HKU-IS)
  - Samba in DUTS-TE/ECSSD/HKU-IS: fill with CVPR25 values
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# =============================================================================
# VERIFIED S-measure values per (method_base, dataset)
# Sources noted in parentheses
# =============================================================================
SOD5_DATA = {
    # --- ICON-S (Swin-B backbone) — M3Net Table I ---
    'ICON-S': {
        'DUTS-TE':   (0.917, 'Sα 0.917 / MAE 0.025 / Em 0.960 / Fw 0.886', 'ICON-S Swin-B, M3Net arxiv 2309.08365 Table I'),
        'DUT-OMRON': (0.869, 'Sα 0.869 / MAE 0.043 / Em 0.906 / Fw 0.804', 'ICON-S Swin-B, M3Net arxiv 2309.08365 Table I'),
        'ECSSD':     (0.941, 'Sα 0.941 / MAE 0.023 / Em 0.972 / Fw 0.936', 'ICON-S Swin-B, M3Net arxiv 2309.08365 Table I'),
        'HKU-IS':    (0.935, 'Sα 0.935 / MAE 0.022 / Em 0.974 / Fw 0.925', 'ICON-S Swin-B, M3Net arxiv 2309.08365 Table I'),
        'PASCAL-S':  (0.885, 'Sα 0.885 / MAE 0.048 / Em 0.930 / Fw 0.854', 'ICON-S Swin-B, M3Net arxiv 2309.08365 Table I'),
    },
    # --- M3Net (M3Net-S Swin-B 384×384) — M3Net Table I & III ---
    'M3Net': {
        'DUTS-TE':   (0.927, 'Sα 0.927 / MAE 0.024 / Em 0.960 / Fw 0.902', 'M3Net-S SwinB 384², M3Net arxiv 2309.08365 Table I/III'),
        'DUT-OMRON': (0.872, 'Sα 0.872 / MAE 0.045 / Em 0.903 / Fw 0.811', 'M3Net-S SwinB 384², M3Net arxiv 2309.08365 Table I/III'),
        'ECSSD':     (0.948, 'Sα 0.948 / MAE 0.021 / Em 0.974 / Fw 0.947', 'M3Net-S SwinB 384², M3Net arxiv 2309.08365 Table I/III'),
        'HKU-IS':    (0.943, 'Sα 0.943 / MAE 0.019 / Em 0.977 / Fw 0.937', 'M3Net-S SwinB 384², M3Net arxiv 2309.08365 Table I/III'),
        'PASCAL-S':  (0.889, 'Sα 0.889 / MAE 0.047 / Em 0.932 / Fw 0.864', 'M3Net-S SwinB 384², M3Net arxiv 2309.08365 Table I/III'),
    },
    # --- MINet (MINet-R ResNet50) — M3Net Table I ---
    'MINet': {
        'DUTS-TE':   (0.884, 'Sα 0.884 / MAE 0.037 / Em 0.927 / Fw 0.825', 'MINet-R ResNet50, M3Net arxiv 2309.08365 Table I'),
        'DUT-OMRON': (0.833, 'Sα 0.833 / MAE 0.056 / Em 0.869 / Fw 0.738', 'MINet-R ResNet50, M3Net arxiv 2309.08365 Table I'),
        'ECSSD':     (0.925, 'Sα 0.925 / MAE 0.033 / Em 0.957 / Fw 0.911', 'MINet-R ResNet50, M3Net arxiv 2309.08365 Table I'),
        'HKU-IS':    (0.919, 'Sα 0.919 / MAE 0.029 / Em 0.960 / Fw 0.897', 'MINet-R ResNet50, M3Net arxiv 2309.08365 Table I'),
        'PASCAL-S':  (0.856, 'Sα 0.856 / MAE 0.064 / Em 0.903 / Fw 0.809', 'MINet-R ResNet50, M3Net arxiv 2309.08365 Table I'),
    },
    # --- PoolNet+ (PoolNet-R+ ResNet50) — PoolNet+ TPAMI22 Table 5 ---
    'PoolNet+': {
        'DUTS-TE':   (0.890, 'Sα 0.890 / MAE 0.039 / Fβ 0.894', 'PoolNet-R+ ResNet50, PoolNet+ TPAMI22 Table 5'),
        'DUT-OMRON': (0.842, 'Sα 0.842 / MAE 0.056 / Fβ 0.831', 'PoolNet-R+ ResNet50, PoolNet+ TPAMI22 Table 5'),
        'ECSSD':     (0.925, 'Sα 0.925 / MAE 0.040 / Fβ 0.949', 'PoolNet-R+ ResNet50, PoolNet+ TPAMI22 Table 5'),
        'HKU-IS':    (0.921, 'Sα 0.921 / MAE 0.034 / Fβ 0.941', 'PoolNet-R+ ResNet50, PoolNet+ TPAMI22 Table 5'),
        'PASCAL-S':  (0.864, 'Sα 0.864 / MAE 0.068 / Fβ 0.879', 'PoolNet-R+ ResNet50, PoolNet+ TPAMI22 Table 5'),
    },
    # --- EDN (TIP 2022) — Samba+ arxiv 2602.01593 Table I (column order: DUTS,DUT-O,HKU-IS,PASCAL-S,ECSSD) ---
    'EDN': {
        'DUTS-TE':   (0.892, 'Sα 0.892 / Fm 0.893 / Em 0.933', 'EDN TIP22, Samba+ arxiv 2602.01593 Table I'),
        'DUT-OMRON': (0.849, 'Sα 0.849 / Fm 0.821 / Em 0.884', 'EDN TIP22, Samba+ arxiv 2602.01593 Table I'),
        'ECSSD':     (0.927, 'Sα 0.927 / Fm 0.950 / Em 0.957', 'EDN TIP22, Samba+ arxiv 2602.01593 Table I'),
        'HKU-IS':    (0.924, 'Sα 0.924 / Fm 0.940 / Em 0.963', 'EDN TIP22, Samba+ arxiv 2602.01593 Table I'),
        'PASCAL-S':  (0.864, 'Sα 0.864 / Fm 0.879 / Em 0.907', 'EDN TIP22, Samba+ arxiv 2602.01593 Table I'),
    },
    # --- U²-Net (PR 2020) — paper Table 2, well-established in SOD community ---
    'U^2-Net': {
        'DUTS-TE':   (0.873, 'Sα 0.873 / MAE 0.047 / Fβ 0.842 (U²-Net 176MB model)', 'U²-Net PR 2020 Table 2'),
        'DUT-OMRON': (0.844, 'Sα 0.844 / MAE 0.054 / Fβ 0.779', 'U²-Net PR 2020 Table 2'),
        'ECSSD':     (0.923, 'Sα 0.923 / MAE 0.041 / Fβ 0.931', 'U²-Net PR 2020 Table 2'),
        'HKU-IS':    (0.916, 'Sα 0.916 / MAE 0.035 / Fβ 0.919', 'U²-Net PR 2020 Table 2'),
        'PASCAL-S':  (0.843, 'Sα 0.843 / MAE 0.074 / Fβ 0.835', 'U²-Net PR 2020 Table 2'),
    },
    # --- Samba (CVPR 2025) — Samba+ arxiv 2602.01593 Table I ---
    'Samba': {
        'DUTS-TE':   (0.932, 'Sα 0.932 / Fm 0.935 / Em 0.958', 'Samba CVPR25, Samba+ arxiv 2602.01593 Table I'),
        'DUT-OMRON': (0.889, 'Sα 0.889 / Fm 0.880 / Em 0.928', 'Samba CVPR25, Samba+ arxiv 2602.01593 Table I'),
        'ECSSD':     (0.953, 'Sα 0.953 / Fm 0.965 / Em 0.975', 'Samba CVPR25, Samba+ arxiv 2602.01593 Table I'),
        'HKU-IS':    (0.945, 'Sα 0.945 / Fm 0.958 / Em 0.975', 'Samba CVPR25, Samba+ arxiv 2602.01593 Table I'),
        'PASCAL-S':  (0.892, 'Sα 0.892 / Fm 0.904 / Em 0.937', 'Samba CVPR25, Samba+ arxiv 2602.01593 Table I'),
    },
}

STANDARD_SOD = ['DUTS-TE', 'DUT-OMRON', 'ECSSD', 'HKU-IS', 'PASCAL-S']
HR_SOD       = ['HRSOD', 'DAVIS-S', 'UHRSD']
SHADOW_SHEETS = ['SBU-Refine', 'ISTD']

# Methods that do NOT evaluate on HR-SOD benchmarks
NOT_ON_HR = {'ICON-S', 'M3Net', 'MINet', 'PoolNet+', 'EDN', 'U^2-Net', 'MINet'}

# =============================================================================
# Flag-only updates
# =============================================================================
def mvanet_note(sn):
    return (f'⚠️ MVANet (CVPR 2024, arxiv 2404.07445) is a Dichotomous Image Segmentation '
            f'(DIS) method — evaluated on DIS-5K dataset ONLY. '
            f'It does NOT report on {sn} SOD benchmark. '
            f'This entry is misplaced; should be in a DIS sheet, not SOD.')

def not_on_hr_note(method, sn):
    return (f'⚠️ {method} does not evaluate on {sn} in its original paper. '
            f'{sn} is a High-Resolution SOD benchmark; '
            f'{method} reports only on standard SOD benchmarks '
            f'(DUTS-TE/DUT-OMRON/ECSSD/HKU-IS/PASCAL-S).')

S3OD_NOTE = ('⚠️ S3OD (ICLR 2026, arxiv 2510.21605) evaluates on DIS subsets '
             '(DIS-1 through DIS-4) and HR-SOD benchmarks '
             '(DAVIS-S, HRSOD-TE, UHRSD-TE, DUTS-TE, DUT-OMRON). '
             'It does NOT report ECSSD or HKU-IS results in its paper. '
             'This entry is misplaced.')


def main():
    wb = load_workbook(XLSX)
    filled = 0
    updated = 0

    all_sheets = STANDARD_SOD + HR_SOD + SHADOW_SHEETS

    for sn in all_sheets:
        if sn not in wb.sheetnames:
            continue
        s = wb[sn]

        for i, row in enumerate(s.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            if not (len(row) > 1 and row[1]):
                continue

            method_cell = str(row[1]).strip()
            # Normalise base name (strip parenthetical suffixes)
            base = method_cell.split('(')[0].strip()

            cur_sm   = row[6] if len(row) > 6 else None
            cur_note = str(row[10] or '') if len(row) > 10 else ''

            has_sm   = cur_sm and str(cur_sm).strip() not in ('', 'N/A')
            verified = '✅' in cur_note

            # ── MVANet: update flag in ALL SOD sheets ─────────────────────
            if base == 'MVANet' and sn in STANDARD_SOD + HR_SOD:
                if '⚠️' in cur_note and 'DIS' not in cur_note:
                    s.cell(row=i, column=11).value = mvanet_note(sn)
                    if not has_sm:
                        s.cell(row=i, column=7).value = None  # clear wrong duplicate
                        s.cell(row=i, column=8).value = 'N/A (DIS method, not SOD — see note)'
                    updated += 1
                    print(f'UPDATE [MVANet] [{sn}] r{i}')
                continue

            # ── S3OD: update flag in ECSSD / HKU-IS ──────────────────────
            if base == 'S3OD (synthetic)' and sn in ('ECSSD', 'HKU-IS'):
                if not verified:
                    s.cell(row=i, column=11).value = S3OD_NOTE
                    updated += 1
                    print(f'UPDATE [S3OD] [{sn}] r{i}')
                continue

            # ── Standard SOD sheets: fill verified Sm values ──────────────
            if sn in STANDARD_SOD and base in SOD5_DATA:
                if sn in SOD5_DATA[base] and not verified:
                    sm, ap_text, src = SOD5_DATA[base][sn]
                    s.cell(row=i, column=7).value = sm
                    s.cell(row=i, column=8).value = ap_text
                    s.cell(row=i, column=11).value = (
                        f'✅ verified ({src}): {sn} Sα={sm}'
                    )
                    filled += 1
                    print(f'FILL [{sn}] r{i} {method_cell}: Sα={sm}')
                continue

            # ── HR sheets: update flag for methods not evaluated there ─────
            if sn in HR_SOD and base in NOT_ON_HR:
                if not verified and '不評估' not in cur_note and 'does not evaluate' not in cur_note:
                    s.cell(row=i, column=11).value = not_on_hr_note(base, sn)
                    updated += 1
                    print(f'UPDATE [not-on-HR] [{sn}] r{i} {method_cell}')

    wb.save(XLSX)
    print(f'\nTotal: filled={filled}, flags_updated={updated}')


if __name__ == '__main__':
    main()
