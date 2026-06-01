# -*- coding: utf-8 -*-
import openpyxl, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
f = 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'
wb = openpyxl.load_workbook(f)
ws = wb['MVTec AD2']

SRC = '來源: VAND 3.0 報告 arXiv 2509.17615 Table 2 / RoBiS arXiv 2505.21152 Table 1'
def note(seg, extra=''):
    return ('✅ verified — AD2 官方指標=SegF1(像素F1，排名依據)；SegF1_priv=%s（%s）。'
            'ClassF1/AU-PRO@0.05 官方僅 RoBiS 公布。%s' % (seg, SRC, extra))
PEND = '⏳ Pending — 非 VAND 官方 leaderboard 條目；AD2 SegF1 待官方表核對（不放佔位值）。'

# cols: 資料集,類別,方法,作者,發表,年月,狀態,ClassF1,(P-AUROC NA),SegF1,AU-PRO@0.05,FPS,備註,連結,GitHub
R = [
 ['MVTec AD 2','基於表徵 (Representation)','ISVL (VAND 2025 第一名)','Xingao Wang et al.','CVPR 2025 VAND 3.0 (Track1 第一名)','2025-06','✅ 已驗證 (第一名)','N/A','N/A',53.81,'N/A','N/A',note('53.81','[基於表徵]'),'https://github.com/ISVL119/isvl/tree/main','https://github.com/ISVL119/isvl/tree/main'],
 ['MVTec AD 2','基於表徵 (Representation)','RoBiS (VAND 2025 第二名)','Xurui Li, Zhongsheng Jiang, Tingxuan Ai, Yu Zhou','CVPR 2025 VAND 3.0 (Track1 第二名)','2025-05','✅ 已驗證 (第二名)',79.62,'N/A',51.00,67.25,'N/A','✅ verified — ClassF1 79.62 / SegF1 51.00 / AU-PRO@0.05 67.25 (private)。來源: RoBiS arXiv 2505.21152 Table 1&2。[基於表徵]','https://arxiv.org/abs/2505.21152','https://github.com/xrli-U/RoBiS'],
 ['MVTec AD 2','基於表徵 (Representation)','ASEG (VAND 2025 第三名)','VAND 3.0 team','CVPR 2025 VAND 3.0 (Track1 第三名)','2025-06','✅ 已驗證 (第三名)','N/A','N/A',50.43,'N/A','N/A',note('50.43'),'https://arxiv.org/abs/2509.17615','N/A'],
 ['MVTec AD 2','基於表徵 (Representation)','SuperAD (Training-free DINOv2)','Hugo H. D. Roggeveen et al.','CVPR 2025 VAND 3.0 (Track1 第四名)','2025-05','✅ 已驗證 (第四名)','N/A','N/A',47.75,'N/A','N/A',note('47.75','原 47.18 已修正'),'https://arxiv.org/abs/2505.19750','N/A'],
 ['MVTec AD 2','基於表徵 (Representation)','Filter-Former (VAND 2025 第五名)','VAND 3.0 team','CVPR 2025 VAND 3.0 (Track1 第五名)','2025-06','✅ 已驗證 (第五名)','N/A','N/A',44.43,'N/A','N/A',note('44.43'),'https://arxiv.org/abs/2509.17615','N/A'],
 ['MVTec AD 2','基於 NF (Normalizing Flow)','MSFlow (AD2 baseline)','Zhou et al.','IEEE TNNLS 2024 (AD2 baseline)','2024-01','✅ 已驗證 (AD2 baseline)','N/A','N/A',21.8,'N/A','N/A',note('21.8','AD2 官方 baseline 中最佳。[基於 NF]'),'https://arxiv.org/abs/2308.15300','https://github.com/cool-xuan/msflow'],
 ['MVTec AD 2','基於知識蒸餾 (Knowledge Distillation)','RD++ (AD2 baseline)','Tran Dinh Tien et al.','CVPR 2023 (AD2 baseline)','2023-06','✅ 已驗證 (AD2 baseline)','N/A','N/A',19.2,'N/A','N/A',note('19.2','[基於知識蒸餾]'),'https://arxiv.org/abs/2206.15476','https://github.com/tientrandinh/Revisiting-Reverse-Distillation'],
 ['MVTec AD 2','基於重構 (Reconstruction)','RD4AD (AD2 baseline)','Hanqiu Deng, Xingyu Li','CVPR 2022 (AD2 baseline)','2022-06','✅ 已驗證 (AD2 baseline)','N/A','N/A',18.1,'N/A','N/A',note('18.1','原 14.2 已修正。[基於重構]'),'https://arxiv.org/abs/2201.10703','https://github.com/hq-deng/RD4AD'],
 ['MVTec AD 2','基於表徵 (Representation)','SimpleNet (AD2 baseline)','Zhikang Liu et al.','CVPR 2023 (AD2 baseline)','2023-03','✅ 已驗證 (AD2 baseline)','N/A','N/A',17.7,'N/A','N/A',note('17.7','原 11.8 已修正。[基於表徵]'),'https://arxiv.org/abs/2303.15140','https://github.com/DonaldRR/SimpleNet'],
 ['MVTec AD 2','基於表徵 (Representation)','EfficientAD (AD2 baseline)','Kilian Batzner, Lars Heckler, Rebecca König','WACV 2024 (AD2 baseline)','2023-03','✅ 已驗證 (AD2 baseline)','N/A','N/A',15.4,'N/A','N/A',note('15.4','原 16.9 已修正。[基於表徵]'),'https://arxiv.org/abs/2303.14535','https://github.com/nelson1425/EfficientAD'],
 ['MVTec AD 2','基於資料擴增 (Data Augmentation)','DSR (AD2 baseline)','Vitjan Zavrtanik et al.','ECCV 2022 (AD2 baseline)','2022-10','✅ 已驗證 (AD2 baseline)','N/A','N/A',10.9,'N/A','N/A',note('10.9','[基於資料擴增]'),'https://arxiv.org/abs/2208.01521','https://github.com/VitjanZ/DSR_anomaly_detection'],
 ['MVTec AD 2','基於表徵 (Representation)','PatchCore (AD2 baseline)','Karsten Roth et al.','CVPR 2022 (AD2 baseline)','2022-06','✅ 已驗證 (AD2 baseline)','N/A','N/A',3.7,'N/A','N/A',note('3.7','原 13.4 已修正（高解析度下大幅退化）。[基於表徵]'),'https://arxiv.org/abs/2106.08265','https://github.com/amazon-science/patchcore-inspection'],
 ['MVTec AD 2','基於擴散 (Diffusion)','DDAD (AD2 baseline)','Arian Mousakhan, Thomas Brox, Jawad Tayyub','BMVC 2024','2023-05','⏳ Pending','N/A','N/A','N/A','N/A','N/A',PEND+' [基於擴散]','https://arxiv.org/abs/2307.08816','https://github.com/arimousa/DDAD'],
 ['MVTec AD 2','基於資料擴增 (Data Augmentation)','DRAEM (AD2 baseline)','Vitjan Zavrtanik, Matej Kristan, Danijel Skocaj','ICCV 2021','2021-08','⏳ Pending','N/A','N/A','N/A','N/A','N/A',PEND+' [基於資料擴增]','https://arxiv.org/abs/2108.07610','https://github.com/VitjanZ/DRAEM'],
 ['MVTec AD 2','基於表徵 (Representation)','INP-Former (AD2 baseline)','Wei Luo, Yunkang Cao et al.','CVPR 2025 / arXiv 2503.02424','2025-03','⏳ Pending','N/A','N/A','N/A','N/A','N/A',PEND+' 原 SegF1=21.8 疑似誤植 MSFlow 數值，已清除。[基於表徵]','https://arxiv.org/abs/2503.02424','https://github.com/luow23/INP-Former'],
 ['MVTec AD 2','基於表徵 (Representation)','SuperADD (VAND 4.0)','Lukas Roming et al.','arXiv preprint (CVPR 2026 VAND 4.0 Workshop)','2026-05','⏳ Pending','N/A','N/A','N/A','N/A','N/A',PEND+' arXiv preprint，未列入 VAND 3.0 官方表。[基於表徵]','https://arxiv.org/abs/2605.14808','N/A'],
]

ws.cell(row=1, column=1).value = ('資料集：MVTec AD 2  (排序：依 SegF1 由高至低。官方指標 = ClassF1(影像F1) / '
    'SegF1(像素F1，排名依據) / AU-PRO@0.05，非 AUROC。欄位對應：I-AUROC欄→ClassF1、P-AP欄→SegF1、P-PRO欄→AU-PRO@0.05)')

maxr = ws.max_row
for rr in range(4, maxr + 1):
    for cc in range(1, 16):
        ws.cell(row=rr, column=cc).value = None
for i, row in enumerate(R):
    for cc, val in enumerate(row, 1):
        ws.cell(row=4 + i, column=cc).value = val
wb.save(f)
print('AD2 rebuilt: %d rows. max_row now %d' % (len(R), ws.max_row))
