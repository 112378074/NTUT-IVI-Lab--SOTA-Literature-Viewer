#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_voc_lvis.py
Fill verified metrics and update specific flags for:
  - PASCAL VOC 2007: flag mainstream methods that do NOT report standard
    closed-set 20-class mAP@0.5 on VOC 2007 test set
  - LVIS v1.0 val: fill Co-DETR (ViT-L) verified AP; flag misplaced/
    non-standard entries

Sources:
  - Co-DETR: arxiv 2211.12860 abstract → "67.9% AP on LVIS val" (ViT-L)
  - OW-OVD: CVF CVPR2025 abstract → evaluates on M-OWODB/S-OWODB (not standard VOC mAP)
  - SEEN-DA: CVF CVPR2025 → domain adaptation protocol (VOC→target)
  - All other modern detectors (RF-DETR, Grounding DINO, RTMDet-X, etc.)
    confirmed COCO/LVIS focused; no standard VOC 2007 evaluation
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# =============================================================================
# VERIFIED VALUES
# =============================================================================
LVIS_VERIFIED = {
    # Co-DETR (ViT-L) — arxiv 2211.12860 abstract: "67.9% AP on LVIS val"
    'Co-DETR (ViT-L)': (
        67.9,
        'Co-DETR ViT-L 67.9 AP on LVIS val (COCO+LVIS jointly trained)',
        '✅ verified (Co-DETR arxiv 2211.12860 abstract): LVIS val AP=67.9'
    ),
}

# =============================================================================
# FLAG NOTES — PASCAL VOC 2007
# =============================================================================

# Methods evaluated on COCO only; not PASCAL VOC 2007 standard closed-set
COCO_ONLY_NOTE = lambda m, venue, detail='': (
    f'⚠️ {m} ({venue}) evaluates on MS COCO (and KITTI/LVIS/ODinW where applicable); '
    f'does NOT report PASCAL VOC 2007 standard 20-class mAP@0.5 in its paper. '
    f'{detail}This entry is misplaced or pending migration to a COCO sheet.'
)

VOC_FLAGS = {
    # r13/r14 — RF-DETR
    'RF-DETR (Large)': COCO_ONLY_NOTE('RF-DETR', 'ICLR 2026',
        'RF-DETR focuses on COCO and Roboflow100-VL benchmarks for real-time detection. '),
    'RF-DETR (Nano)': COCO_ONLY_NOTE('RF-DETR Nano', 'ICLR 2026',
        'RF-DETR focuses on COCO and Roboflow100-VL benchmarks for real-time detection. '),

    # r16 — Grounding DINO
    'Grounding DINO': (
        '⚠️ Grounding DINO (Liu et al. ECCV 2024, arxiv 2303.05499) is an open-vocabulary '
        'detector evaluated on COCO val2017, LVIS v1.0, and ODinW-13 benchmarks. '
        'It does NOT report standard PASCAL VOC 2007 closed-set 20-class mAP@0.5 '
        '(open-vocabulary protocol is not comparable to closed-set VOC evaluation). '
        'The "zero-shot VOC" row r12 captures its transfer performance separately.'
    ),

    # r17 — Relation-DETR
    'Relation-DETR (FocalNet-L)': COCO_ONLY_NOTE('Relation-DETR', 'ECCV 2024',
        'Evaluated on COCO val2017 and test-dev; FocalNet-L backbone. '),

    # r18 — Grounding DINO 1.5 Pro
    'Grounding DINO 1.5 Pro': (
        '⚠️ Grounding DINO 1.5 Pro (arxiv 2405.10300) is an open-vocabulary detector '
        'evaluated on COCO, LVIS, ODinW; does NOT report standard PASCAL VOC 2007 '
        'closed-set mAP@0.5. Open-vocabulary results are not comparable to closed-set VOC ranking.'
    ),

    # r20 — RTMDet-X
    'RTMDet-X': COCO_ONLY_NOTE('RTMDet-X', 'arXiv 2022',
        'RTMDet-X evaluated on COCO val2017 (52.8 AP) and CrowdHuman; '
        'no PASCAL VOC 2007 evaluation in paper. '),

    # r21 — Meta-DETR (TPAMI 2022)
    'Meta-DETR': (
        '⚠️ Meta-DETR (Zhang et al. TPAMI 2022, arxiv 2208.00219) is a few-shot OD method. '
        'It evaluates on MS-COCO novel AP at K-shot (1/5/10/30) — see MS-COCO few-shot sheets. '
        'It does NOT report standard PASCAL VOC 2007 20-class closed-set mAP@0.5. '
        'This entry is misplaced; should be in COCO few-shot sheets only.'
    ),

    # r25 — PoolNet+ (TPAMI 2021/2022)
    'PoolNet+': (
        '⚠️ PoolNet+ (Liu et al. TPAMI 2022, arxiv 1904.09569) is a Salient Object Detection '
        '(SOD) method — NOT a standard object detection method. It evaluates on SOD benchmarks '
        '(DUTS-TE, DUT-OMRON, ECSSD, HKU-IS, PASCAL-S) using S-measure/MAE/F-measure. '
        'It does NOT evaluate on PASCAL VOC 2007 20-class mAP. '
        'This entry is misplaced; should only appear in SOD sheets.'
    ),

    # r28 — EfficientDet-D7
    'EfficientDet-D7': COCO_ONLY_NOTE('EfficientDet-D7', 'CVPR 2020',
        'Tan et al. CVPR 2020: evaluated on COCO test-dev only (52.2 AP). '
        'No PASCAL VOC 2007 evaluation in the EfficientDet paper. '),

    # r31 — CenterNet
    'CenterNet': COCO_ONLY_NOTE('CenterNet (Objects as Points)', 'arXiv 2019',
        'Zhou et al. arXiv 1904.07850: evaluated on MS COCO val2017 and KITTI (3D OD) and '
        'COCO keypoints. No standard PASCAL VOC 2007 AP evaluation in paper. '),

    # r15 — YOLOv13-X
    'YOLOv13-X': (
        '⚠️ YOLOv13-X (arxiv 2506.17733) performs cross-domain evaluation: models trained '
        'on COCO are tested on PASCAL VOC 2007 as a generalization benchmark, not as a '
        'standard closed-set training+test protocol. '
        'Cross-domain transfer AP is not directly comparable to methods trained on VOC 2007 trainval. '
        'The specific cross-domain transfer AP number needs verification from the full PDF.'
    ),
}

# CVPR/WACV 2025 specialized papers — specific reason each does NOT report standard VOC 2007 mAP
VOC_SPECIALIZED = {
    # r33
    'OW-OVD': (
        '⚠️ OW-OVD (Xi et al. CVPR 2025) is an Open-World + Open-Vocabulary OD paper. '
        'Evaluation is on M-OWODB and S-OWODB benchmarks (open-world metrics: '
        'U-Recall, U-mAP for unknown classes). '
        'Does NOT report standard PASCAL VOC 2007 closed-set 20-class mAP@0.5. '
        'VOC categories appear only as the "known class" base set in the OWOD benchmark split.'
    ),

    # r34
    'SEEN-DA': (
        '⚠️ SEEN-DA (Li et al. CVPR 2025) is a Domain Adaptive OD paper. '
        'Evaluation uses cross-domain transfer protocols (e.g. Cityscapes→KITTI, '
        'VOC→Clipart, VOC→Watercolor) — reporting domain-adaptation mAP, '
        'NOT standard closed-set PASCAL VOC 2007 test-set mAP@0.5. '
        'This entry is misplaced in the standard VOC 2007 ranking.'
    ),

    # r35
    'Test-Time Backdoor Detection for Object Detection Models': (
        '⚠️ Test-Time Backdoor Detection for OD Models (Zhang et al. CVPR 2025) is a security '
        'paper about backdoor attack detection. VOC 2007 is used as an attack-evaluation target '
        '(not as a standard benchmark for detection mAP). '
        'Does NOT report standard 20-class mAP@0.5; reports attack-success rate / detection accuracy.'
    ),

    # r36
    'SET': (
        '⚠️ SET — Spectral Enhancement for Tiny Object Detection (Sun et al. CVPR 2025) '
        'is a tiny OD paper. Evaluated on tiny-object benchmarks (e.g. VisDrone, AI-TOD, DOTA). '
        'Does NOT report PASCAL VOC 2007 standard 20-class mAP@0.5.'
    ),

    # r37
    'Style Evolving along Chain-of-Thought for Unknown-Domain Obj': (
        '⚠️ Style Evolving along Chain-of-Thought for Unknown-Domain OD (Zhang et al. CVPR 2025) '
        'is a domain generalization / style adaptation paper. '
        'Evaluates on unknown-domain transfer benchmarks (not standard closed-set VOC 2007 mAP@0.5). '
        'VOC may appear as a source domain but results are domain-shift metrics.'
    ),

    # r38
    'Towards RAW Object Detection in Diverse Conditions': (
        '⚠️ Towards RAW Object Detection in Diverse Conditions (Li et al. CVPR 2025) '
        'evaluates on RAW-image domain-specific datasets (not standard RGB PASCAL VOC 2007). '
        'Does NOT report standard 20-class VOC mAP@0.5; focuses on RAW-domain detection metrics.'
    ),

    # r39
    'Revisiting Generative Replay for Class Incremental Object De': (
        '⚠️ Revisiting Generative Replay for Class Incremental OD (Zhang et al. CVPR 2025) '
        'uses VOC 2007 in a class-incremental learning protocol (multi-task splits, '
        'e.g. first 10 classes then next 10). '
        'Reports "current task AP" / "all-task AP" under incremental constraints — '
        'NOT directly comparable to standard closed-set 20-class mAP@0.5.'
    ),

    # r40
    'Open-World Objectness Modeling Unifies Novel Object Detectio': (
        '⚠️ Open-World Objectness Modeling (Zhang et al. CVPR 2025) uses OWOD benchmark '
        '(M-OWODB / S-OWODB, derived from VOC+COCO categories). '
        'Reports open-world metrics (U-Recall, WI, AOSE, mAP on known classes). '
        'NOT comparable to standard closed-set PASCAL VOC 2007 20-class mAP@0.5.'
    ),

    # r41
    'Percept, Memory, and Imagine': (
        '⚠️ Percept, Memory, and Imagine — World Feature Simulating for Open-Domain Unknown '
        '(Wu et al. CVPR 2025) is an open-domain unknown OD paper. '
        'Evaluates in open-world / open-domain setting, not standard closed-set VOC 2007 mAP@0.5.'
    ),

    # r42 — SimLTD is a long-tail paper evaluating on LVIS, not VOC
    'SimLTD': (
        '⚠️ SimLTD (Tran et al. CVPR 2025) is a long-tailed OD paper evaluated on the '
        'LVIS v1 benchmark (supervised + semi-supervised settings). '
        'Does NOT evaluate on PASCAL VOC 2007. This entry is misplaced; '
        'should be in the LVIS v1.0 val sheet (r23 there).'
    ),

    # r43
    'Learning Endogenous Attention for Incremental Object Detecti': (
        '⚠️ Learning Endogenous Attention for Incremental OD (Song et al. CVPR 2025) '
        'uses VOC 2007 in a class-incremental / continual learning protocol (task splits). '
        'Reports forgetting / plasticity metrics — NOT standard closed-set 20-class mAP@0.5. '
        'Not directly comparable to standard VOC ranking.'
    ),

    # r44
    'Bit-Flip Induced Latency Attacks in Object Detection': (
        '⚠️ Bit-Flip Induced Latency Attacks in OD (Sistla et al. WACV 2025) '
        'is a security / adversarial paper. '
        'VOC and COCO appear as target datasets for bit-flip attack evaluation '
        '(attack success rate / latency degradation), NOT standard detection mAP. '
        'Does NOT report standard closed-set PASCAL VOC 2007 mAP@0.5.'
    ),

    # r45
    'ERUP-YOLO': (
        '⚠️ ERUP-YOLO (Ogino et al. WACV 2025) enhances YOLO robustness for adverse weather. '
        'Evaluated on COCO and adverse-weather augmented test sets (fog / rain / etc.). '
        'Results are robustness metrics under augmented conditions, NOT standard '
        'closed-set PASCAL VOC 2007 mAP@0.5 training-on-VOC protocol. '
        'Needs manual paper PDF check to confirm if standard VOC test AP is reported.'
    ),

    # r46
    'Decoupled PROB': (
        '⚠️ Decoupled PROB (Inoue et al. WACV 2025) is an open-world OD paper '
        '(variant of PROB — Probabilistic OWL OD). '
        'Evaluated on OW-DETR / M-OWODB benchmark (open-world metrics: U-Recall, WI, AOSE). '
        'VOC categories appear as "known" classes in the OWOD split — '
        'NOT standard closed-set 20-class mAP@0.5.'
    ),

    # r47
    'Cross-Domain Multi-Modal Few-Shot Object Detection via Rich': (
        '⚠️ Cross-Domain Multi-Modal Few-Shot OD via Rich Text (Shangguan et al. WACV 2025) '
        'is a cross-domain few-shot OD paper. '
        'Uses VOC and COCO in few-shot cross-domain protocol (K-shot novel class AP). '
        'NOT standard closed-set PASCAL VOC 2007 20-class mAP@0.5. '
        'This entry is misplaced; should be in few-shot OD sheets.'
    ),
}

# =============================================================================
# FLAG NOTES — LVIS v1.0 val
# =============================================================================
LVIS_FLAGS = {
    # r13 — Exploring Hierarchical Consistency
    'Exploring Hierarchical Consistency and Unbiased Objectness for Open-Vocabulary Object Dete': (
        '⚠️ Exploring Hierarchical Consistency and Unbiased Objectness for Open-Vocabulary OD '
        '(arxiv 2604.23344) is an open-vocabulary OD paper that likely reports LVIS AP. '
        'Specific LVIS val AP not extractable without PDF access (very recent paper, May 2026). '
        'Needs manual PDF verification to obtain exact AP number.'
    ),

    # r19 — TFA w/cos is misplaced in LVIS sheet
    'TFA w/cos': (
        '⚠️ TFA w/cos (Wang et al. ICML 2020, arxiv 2003.06957) is a few-shot OD method. '
        'It evaluates on MS-COCO novel K-shot AP (see MS-COCO few-shot sheets) and '
        'possibly VOC few-shot K-shot AP. '
        'It does NOT evaluate on LVIS v1.0 val standard benchmark. '
        'This entry is misplaced in the LVIS sheet.'
    ),

    # r20 — Search and Detect (long-tail LVIS)
    'Search and Detect': (
        '⚠️ Search and Detect (Sidhu et al. CVPR 2025) is a training-free long-tail OD paper '
        'using web-image retrieval. It likely evaluates on LVIS v1.0 val. '
        'Specific LVIS val AP not extractable without PDF access. '
        'Needs manual paper PDF check to fill exact AP number.'
    ),

    # r21 — Fractal Calibration (long-tail LVIS)
    'Fractal Calibration for Long-tailed Object Detection': (
        '⚠️ FRACAL — Fractal Calibration for Long-tailed OD (Alexandridis et al. CVPR 2025, '
        'arxiv 2410.11774) evaluates on LVIS v1.0, COCO, V3Det, OpenImages. '
        'Reports new SOTA on LVIS (surpasses all previous, boosting rare AP by 8.6%). '
        'Exact overall LVIS val AP not in abstract — needs full PDF for specific number. '
        'Needs manual paper PDF check.'
    ),

    # r22 — Open-World Objectness (OWOD, not standard LVIS)
    'Open-World Objectness Modeling Unifies Novel Object Detectio': (
        '⚠️ Open-World Objectness Modeling (Zhang et al. CVPR 2025) uses OWOD benchmark '
        '(M-OWODB / S-OWODB). Reports open-world metrics (U-Recall, WI, AOSE), '
        'NOT standard LVIS v1.0 val detection AP. '
        'This entry is misplaced in the standard LVIS ranking.'
    ),

    # r23 — SimLTD (long-tail LVIS)
    'SimLTD': (
        '⚠️ SimLTD (Tran et al. CVPR 2025, arxiv 2412.20047) evaluates on LVIS v1 benchmark '
        '(supervised + semi-supervised long-tail detection). Reports new SOTA on LVIS v1 '
        '(+10.6 AP_r over previous SOTA). '
        'Specific overall LVIS v1 val AP not in abstract — needs full PDF for exact number.'
    ),

    # r24 — Enhancing Novel OD via Cooperative Foundational Models
    'Enhancing Novel Object Detection via Cooperative Foundationa': (
        '⚠️ Enhancing Novel OD via Cooperative Foundational Models (Bharadwaj et al. WACV 2025) '
        'focuses on novel object detection using foundational model cooperation. '
        'Evaluation protocol may use LVIS or COCO novel class splits rather than '
        'standard LVIS v1.0 val full-set AP. '
        'Needs manual paper PDF check to confirm dataset split and AP value.'
    ),

    # r25 — Enhancing Embodied OD
    'Enhancing Embodied Object Detection with Spatial Feature Mem': (
        '⚠️ Enhancing Embodied OD with Spatial Feature Memory (Chapman et al. WACV 2025) '
        'is an embodied/egocentric OD paper. '
        'Evaluates on embodied/robot-facing datasets (e.g. AI2-THOR, RoboTHOR, Ego4D). '
        'Does NOT evaluate on standard LVIS v1.0 val benchmark. '
        'This entry is misplaced in the LVIS v1.0 val sheet.'
    ),
}


def main():
    wb = load_workbook(XLSX)
    filled = 0
    updated = 0

    # ── LVIS v1.0 val ─────────────────────────────────────────────────────────
    if 'LVIS v1.0 val' in wb.sheetnames:
        s = wb['LVIS v1.0 val']
        for i, row in enumerate(s.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            if not (len(row) > 1 and row[1]):
                continue
            method_full = str(row[1]).strip()
            cur_note = str(row[10] or '') if len(row) > 10 else ''
            cur_m = row[6] if len(row) > 6 else None
            if '✅' in cur_note:
                continue

            # Fill verified
            if method_full in LVIS_VERIFIED and not cur_m:
                ap, ap_text, note = LVIS_VERIFIED[method_full]
                s.cell(row=i, column=7).value = ap
                s.cell(row=i, column=8).value = ap_text
                s.cell(row=i, column=11).value = note
                filled += 1
                print(f'FILL [LVIS v1.0 val] r{i} {method_full}: AP={ap}')
                continue

            # Update flags (only if not already verified)
            # Use prefix matching for long truncated names
            matched_flag = None
            for key, flag_note in LVIS_FLAGS.items():
                if method_full.startswith(key[:50]):
                    matched_flag = flag_note
                    break

            if matched_flag and '⚠' in cur_note and not cur_m:
                # Update with more specific note if current note is generic
                if 'LVIS v1.0 val AP 數字需 paper PDF' in cur_note:
                    s.cell(row=i, column=11).value = matched_flag
                    updated += 1
                    print(f'UPDATE [LVIS v1.0 val] r{i} {method_full[:50]}')

    # ── PASCAL VOC 2007 ───────────────────────────────────────────────────────
    if 'PASCAL VOC 2007' in wb.sheetnames:
        s = wb['PASCAL VOC 2007']
        for i, row in enumerate(s.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            if not (len(row) > 1 and row[1]):
                continue
            method_full = str(row[1]).strip()
            base = method_full.split('(')[0].strip()
            cur_note = str(row[10] or '') if len(row) > 10 else ''
            cur_m = row[6] if len(row) > 6 else None
            if '✅' in cur_note:
                continue
            if cur_m is not None and str(cur_m).strip() not in ('', 'N/A', 'To verify'):
                continue  # already has real number

            new_note = None

            # Check mainstream methods first (exact match)
            if method_full in VOC_FLAGS:
                new_note = VOC_FLAGS[method_full]
            elif base in VOC_FLAGS:
                new_note = VOC_FLAGS[base]

            # Check specialized CVPR/WACV 2025 papers (prefix match for truncated names)
            if new_note is None:
                for key, flag_note in VOC_SPECIALIZED.items():
                    if method_full.startswith(key[:50]) or key.startswith(method_full[:50]):
                        new_note = flag_note
                        break

            if new_note and '⚠' in cur_note and 'PDF 個別驗證' in cur_note:
                s.cell(row=i, column=11).value = new_note
                updated += 1
                print(f'UPDATE [PASCAL VOC 2007] r{i} {method_full[:60]}')

    wb.save(XLSX)
    print(f'\nTotal: filled={filled}, updated={updated}')


if __name__ == '__main__':
    main()
