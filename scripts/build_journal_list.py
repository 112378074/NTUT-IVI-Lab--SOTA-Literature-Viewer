# -*- coding: utf-8 -*-
"""Build a journal list workbook: Image/Computer-Vision + Smart-Manufacturing."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Columns: Domain, Tier, Journal full name, Abbrev, Publisher, IF(JCR 2024/25, verified only),
#          Quartile (confident only), ISSN, Scope, IVAD relevance
# IF: only values verified via web search this session are filled; others blank.
# Relevance legend: ◎◎ 最對口 / ◎ 高 / ○ 中 / △ 邊緣

CV = "影像 / 電腦視覺"
SM = "智慧製造"

rows = [
    # ---- Image / Computer Vision ----
    (CV, "Tier1", "IEEE Transactions on Pattern Analysis and Machine Intelligence", "IEEE TPAMI", "IEEE", "", "Q1", "0162-8828", "電腦視覺、模式辨識、機器學習方法論（領域最高影響力）", "◎"),
    (CV, "Tier1", "International Journal of Computer Vision", "IJCV", "Springer", "", "Q1", "0920-5691", "純電腦視覺、資料集/benchmark（MVTec AD 2 發表於此）", "◎"),
    (CV, "Tier1", "IEEE Transactions on Image Processing", "IEEE TIP", "IEEE", "", "Q1", "1057-7149", "影像處理、重建、分割、復原、影像層級視覺", "◎"),
    (CV, "Tier1", "Pattern Recognition", "PR", "Elsevier", "7.6", "Q1", "0031-3203", "模式辨識、電腦視覺、機器學習，應用廣", "◎"),
    (CV, "Tier2", "IEEE Transactions on Circuits and Systems for Video Technology", "IEEE TCSVT", "IEEE", "", "Q1", "1051-8215", "視訊/影像分析、偵測、異常合成", "◎"),
    (CV, "Tier2", "IEEE Transactions on Multimedia", "IEEE TMM", "IEEE", "", "Q1", "1520-9210", "多媒體、影像/視訊理解", "○"),
    (CV, "Tier2", "Computer Vision and Image Understanding", "CVIU", "Elsevier", "3.5", "Q1", "1077-3142", "電腦視覺 + 影像處理 + 模式辨識", "○"),
    (CV, "Tier2", "Image and Vision Computing", "IVC", "Elsevier", "3.1", "Q1", "0262-8856", "影像與視覺計算", "○"),
    (CV, "Tier2", "IEEE Transactions on Neural Networks and Learning Systems", "IEEE TNNLS", "IEEE", "", "Q1", "2162-237X", "深度學習（視覺應用大宗）、表徵學習、distillation", "○"),
    (CV, "Tier2", "Pattern Recognition Letters", "PRL", "Elsevier", "", "", "0167-8655", "模式辨識短篇，發表快", "○"),
    (CV, "Tier2", "Neurocomputing", "", "Elsevier", "", "", "0925-2312", "神經網路、電腦視覺/影像通用", "○"),
    (CV, "Tier2/3", "IEEE Transactions on Computational Imaging", "IEEE TCI", "IEEE", "", "", "", "計算成像、影像重建", "○"),
    (CV, "Tier2/3", "IEEE Signal Processing Letters", "IEEE SPL", "IEEE", "", "", "1070-9908", "訊號/影像處理短篇", "△"),
    (CV, "Tier2/3", "Signal Processing: Image Communication", "SP:IC", "Elsevier", "", "", "0923-5965", "影像/視訊編碼與通訊", "△"),
    (CV, "Tier2/3", "Journal of Visual Communication and Image Representation", "JVCIR", "Elsevier", "", "", "1047-3203", "視覺通訊、影像表徵", "△"),
    (CV, "Tier2/3", "Signal, Image and Video Processing", "SIViP", "Springer", "", "", "1863-1703", "影像/視訊處理", "△"),
    (CV, "Tier2/3", "IET Image Processing", "IET-IPR", "IET / Wiley", "", "", "1751-9659", "影像處理（OA 友善）", "△"),
    (CV, "Tier2/3", "Journal of Mathematical Imaging and Vision", "JMIV", "Springer", "", "", "0924-9907", "影像數學理論", "△"),
    (CV, "Tier3", "Machine Vision and Applications", "MVA", "Springer", "", "", "0932-8092", "機器視覺應用", "○"),
    (CV, "Tier3", "The Visual Computer", "TVC", "Springer", "", "", "0178-2789", "視覺計算、電腦圖學", "△"),
    (CV, "Tier3", "Pattern Analysis and Applications", "PAA", "Springer", "", "", "1433-7541", "模式分析應用", "△"),
    (CV, "Tier3", "Multimedia Tools and Applications", "MTA", "Springer", "", "", "1380-7501", "多媒體/影像（量大）", "△"),
    (CV, "Tier3", "IEEE Access", "", "IEEE", "", "", "2169-3536", "開放取用、跨領域電腦視覺", "△"),
    (CV, "Tier3", "Sensors", "", "MDPI", "", "", "1424-8220", "感測、視覺（OA，量大需篩選）", "△"),
    (CV, "子領域選用", "IEEE Transactions on Medical Imaging", "IEEE TMI", "IEEE", "", "Q1", "0278-0062", "醫療影像（若涉及醫療 CV）", "選用"),
    (CV, "子領域選用", "Medical Image Analysis", "MedIA", "Elsevier", "", "Q1", "1361-8415", "醫療影像分析（若涉及醫療 CV）", "選用"),
    (CV, "子領域選用", "IEEE Transactions on Geoscience and Remote Sensing", "IEEE TGRS", "IEEE", "", "Q1", "0196-2892", "遙測視覺（若涉及遙測）", "選用"),
    (CV, "子領域選用", "ISPRS Journal of Photogrammetry and Remote Sensing", "ISPRS J", "Elsevier", "", "Q1", "0924-2716", "攝影測量、遙測（若涉及遙測）", "選用"),

    # ---- Smart Manufacturing ----
    (SM, "Tier1", "Robotics and Computer-Integrated Manufacturing", "RCIM", "Elsevier", "15.4", "Q1", "0736-5845", "機器人、整合製造、智慧產線（含視覺導引/檢測）", "◎"),
    (SM, "Tier1", "Journal of Manufacturing Systems", "JMS", "Elsevier", "14.2", "Q1", "0278-6125", "製造系統、智慧製造、品質檢測", "◎"),
    (SM, "Tier1", "IEEE Transactions on Industrial Informatics", "IEEE TII", "IEEE", "9.9", "Q1", "1551-3203", "工業資訊、AI 檢測、Industry 4.0（工業視覺異常核心）", "◎◎"),
    (SM, "Tier1", "Computers in Industry", "CinI", "Elsevier", "9.1", "Q1", "0166-3615", "工業電腦應用、視覺檢測", "◎◎"),
    (SM, "Tier1/2", "Journal of Intelligent Manufacturing", "JIM", "Springer", "7.4", "Q1", "0956-5515", "智慧製造、機器學習、瑕疵分類", "◎"),
    (SM, "Tier1/2", "Journal of Manufacturing Processes", "JMP", "Elsevier", "6.8", "Q1", "1526-6125", "製程、加工、線上檢測", "○"),
    (SM, "Tier1/2", "IEEE Transactions on Automation Science and Engineering", "IEEE T-ASE", "IEEE", "", "Q1", "1545-5955", "自動化、產線檢測", "○"),
    (SM, "Tier1/2", "IEEE Transactions on Industrial Electronics", "IEEE TIE", "IEEE", "", "Q1", "0278-0046", "工業電子、視覺/感測檢測", "○"),
    (SM, "Tier1/2", "IEEE Transactions on Instrumentation and Measurement", "IEEE TIM", "IEEE", "", "Q1", "0018-9456", "量測式表面瑕疵檢測", "◎"),
    (SM, "Tier1/2", "Journal of Manufacturing Technology Management", "JMTM", "Emerald", "", "Q1", "1741-038X", "製造管理、Industry 4.0 策略", "△"),
    (SM, "Tier2", "International Journal of Computer Integrated Manufacturing", "IJCIM", "Taylor & Francis", "", "", "0951-192X", "電腦整合製造、Industry 4.0→5.0", "○"),
    (SM, "Tier2", "Computers & Industrial Engineering", "C&IE", "Elsevier", "", "Q1", "0360-8352", "工業工程、最佳化、品質", "△"),
    (SM, "Tier2", "CIRP Journal of Manufacturing Science and Technology", "CIRP-JMST", "Elsevier", "", "", "1755-5817", "製造科學（CIRP 系統）", "○"),
    (SM, "Tier2", "International Journal of Production Research", "IJPR", "Taylor & Francis", "", "Q1", "0020-7543", "生產系統、製造研究", "△"),
    (SM, "Tier2", "International Journal of Advanced Manufacturing Technology", "IJAMT", "Springer", "", "", "0268-3768", "先進製造技術（量大）", "○"),
    (SM, "Tier2", "Advances in Manufacturing", "AiM", "Springer", "", "", "2095-3127", "製造前沿", "△"),
    (SM, "Tier2", "Manufacturing Letters", "ManLet", "Elsevier", "", "", "2213-8463", "製造短篇快報", "△"),
    (SM, "Tier2/3", "Engineering Applications of Artificial Intelligence", "EAAI", "Elsevier", "", "Q1", "0952-1976", "AI 工程應用（含製造檢測）", "○"),
    (SM, "Tier2/3", "Advanced Engineering Informatics", "AEI", "Elsevier", "", "Q1", "1474-0346", "工程資訊、數位孿生", "○"),
    (SM, "Tier2/3", "Expert Systems with Applications", "ESWA", "Elsevier", "", "Q1", "0957-4174", "AI 系統應用", "○"),
    (SM, "Tier2/3", "Measurement", "", "Elsevier", "", "", "0263-2241", "量測、瑕疵檢測", "○"),
    (SM, "Tier2/3", "Applied Sciences", "", "MDPI", "", "", "2076-3417", "開放取用、IIoT、製造視覺", "△"),
    (SM, "Tier2/3", "Machines", "", "MDPI", "", "", "2075-1702", "機械、智慧製造（OA）", "△"),
]

HEADERS = ["領域", "Tier", "期刊全名", "縮寫", "出版社",
           "IF (JCR 2024/25)", "分區", "ISSN", "範疇 / Scope", "對 IVAD 關聯"]

# ---- styling helpers ----
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
CV_FILL = PatternFill("solid", fgColor="E8F1FB")
SM_FILL = PatternFill("solid", fgColor="FCEFE3")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = [16, 11, 52, 14, 18, 15, 8, 13, 46, 14]


def style_sheet(ws, data):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(data, 2):
        fill = CV_FILL if row[0] == CV else SM_FILL
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 9)))
            if c in (6, 7, 8, 10):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(data)+1}"
    ws.row_dimensions[1].height = 30


wb = openpyxl.Workbook()

# Sheet 1: combined
ws_all = wb.active
ws_all.title = "全部期刊"
style_sheet(ws_all, rows)

# Sheet 2: CV only
ws_cv = wb.create_sheet("影像_電腦視覺")
style_sheet(ws_cv, [r for r in rows if r[0] == CV])

# Sheet 3: Smart Manufacturing only
ws_sm = wb.create_sheet("智慧製造")
style_sheet(ws_sm, [r for r in rows if r[0] == SM])

# Sheet 4: notes / sources
ws_note = wb.create_sheet("說明")
notes = [
    ["說明 / Notes", ""],
    ["", ""],
    ["資料整理日期", "2026-06-01"],
    ["領域", "兩大類：影像/電腦視覺、智慧製造（含工業視覺檢測）"],
    ["IF 欄位", "僅填入本次經 web 查證的 JCR 2024/2025 影響因子；未查證者留白，請至 JCR/Scimago 官方確認。"],
    ["分區 欄位", "僅標示確定為 Q1 的旗艦期刊與高把握者；留白＝未於本次確認（分區依 JCR 學科分類年度而異）。"],
    ["IF 已查證之期刊", "Pattern Recognition 7.6 / CVIU 3.5 / Image and Vision Computing 3.1 / "
                        "RCIM ~15.4 / J. Manufacturing Systems 14.2 / IEEE TII 9.9 / "
                        "Computers in Industry 9.1 / J. Intelligent Manufacturing 7.4 / J. Manufacturing Processes 6.8"],
    ["對 IVAD 關聯", "◎◎ 最對口投稿/檢索目標 / ◎ 高 / ○ 中 / △ 邊緣或量大需篩選 / 選用＝特定子領域"],
    ["ISSN", "為各刊常用 (print) ISSN，僅供建檔參考；投稿前請至官方頁面再次確認。"],
    ["最對口期刊（IVAD/瑕疵偵測）", "IEEE TII、Computers in Industry、J. Manufacturing Systems、"
                                 "J. Intelligent Manufacturing、IEEE TIM、RCIM、TPAMI、TIP、IJCV、Pattern Recognition、CVIU、TCSVT"],
    ["", ""],
    ["資料來源 (查證 IF)", ""],
    ["Pattern Recognition (IF 7.6, Q1)", "https://www.journalmetrics.org/journal/pattern-recognition"],
    ["Computer Vision and Image Understanding (IF 3.5, Q1)", "https://research.com/journal/computer-vision-and-image-understanding"],
    ["Image and Vision Computing", "https://research.com/journal/image-and-vision-computing"],
    ["International Journal of Computer Vision (SJR 3.136, Q1)", "https://www.resurchify.com/impact/details/72242"],
    ["Robotics and Computer-Integrated Manufacturing (IF ~15.4)", "https://www.bioxbio.com/journal/ROBOT-CIM-INT-MANUF"],
    ["Journal of Manufacturing Systems (IF 14.2, Q1)", "https://www.journalmetrics.org/journal/journal-of-manufacturing-systems"],
    ["IEEE Transactions on Industrial Informatics (IF 9.9, Q1)", "https://www.journalmetrics.org/journal/ieee-transactions-on-industrial-informatics"],
    ["Computers in Industry (IF 9.1, Q1)", "https://www.bioxbio.com/journal/COMPUT-IND"],
    ["Journal of Intelligent Manufacturing (IF 7.4, Q1)", "https://www.resurchify.com/impact/details/24363"],
    ["Journal of Manufacturing Processes (IF 6.8, Q1)", "https://www.journalmetrics.org/journal/journal-of-manufacturing-processes"],
]
for r, (a, b) in enumerate(notes, 1):
    ca = ws_note.cell(row=r, column=1, value=a)
    cb = ws_note.cell(row=r, column=2, value=b)
    ca.alignment = Alignment(vertical="top", wrap_text=True)
    cb.alignment = Alignment(vertical="top", wrap_text=True)
    if r == 1:
        ca.font = Font(bold=True, size=14, color="1F4E79")
    elif a and not b:
        ca.font = Font(bold=True, color="1F4E79")
    else:
        ca.font = Font(bold=True)
ws_note.column_dimensions["A"].width = 42
ws_note.column_dimensions["B"].width = 78

out = "期刊清單_影像CV_智慧製造.xlsx"
wb.save(out)
print("saved:", out)
print("CV rows:", sum(1 for r in rows if r[0] == CV), "| SM rows:", sum(1 for r in rows if r[0] == SM), "| total:", len(rows))
