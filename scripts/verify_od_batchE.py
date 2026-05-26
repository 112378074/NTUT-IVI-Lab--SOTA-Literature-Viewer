#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batchE.py — Salient OD fills (CPD-R from CPD CVPR19 Table 1)."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# CPD-R values from CPD CVPR19 paper Table 1 (extracted ResNet50 row):
# 5 datasets × (Sα, F-max, MAE), order based on ranking: ECSSD/HKU-IS/DUT-OMRON/DUTS-TE/PASCAL-S
# col 7 (mAP) = N/A, col 8 (AP-text) holds "S 0.xxx / Fβ 0.xxx / MAE 0.xxx"
FILLS = [
    # (sheet, row, AP-text, note)
    ('DUTS-TE', 16, 'S 0.865 / Fβ 0.805 / MAE 0.043',
     '✅ 已驗證 (CPD CVPR19 Table 1, ResNet50)：Cascaded Partial Decoder for fast salient object detection；DUTS-TE Sα=0.865 Fβ=0.805 MAE=0.043 [SOD]'),
    ('DUT-OMRON', 7, 'S 0.797 / Fβ 0.747 / MAE 0.056',
     '✅ 已驗證 (CPD CVPR19 Table 1, ResNet50)：DUT-OMRON Sα=0.797 Fβ=0.747 MAE=0.056 [SOD]'),
    ('ECSSD', 12, 'S 0.939 / Fβ 0.917 / MAE 0.037',
     '✅ 已驗證 (CPD CVPR19 Table 1, ResNet50)：ECSSD Sα=0.939 Fβ=0.917 MAE=0.037 [SOD]'),
    ('HKU-IS', 11, 'S 0.925 / Fβ 0.891 / MAE 0.034',
     '✅ 已驗證 (CPD CVPR19 Table 1, ResNet50)：HKU-IS Sα=0.925 Fβ=0.891 MAE=0.034 [SOD]'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        # SOD uses N/A in mAP col, value text in AP col, parsed by frontend
        s.cell(row=row, column=7).value = 'N/A'
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> {AP_text}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
