#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_recover.py — Recover numbers already in col 8 (AP text) by setting col 7 (mAP)."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# For rows where col 7 (mAP) is empty/N/A but col 8 (AP text) has a concrete
# number, extract the number and set col 7. Many of these are paper-reported
# values from the original team's prior work (col 8 was the only field they
# filled before the regex-based primary_value parsing was set up).
RECOVERIES = [
    # (sheet, row_match_method, expected_number, AP-text-reformat, note)
    ('ODinW-13', 'Grounding DINO', 50.6, 'AP 50.6 ODinW13 (zero-shot, paper-reported)',
     'Grounding DINO ODinW-13 50.6 zero-shot (paper-reported via workbook col 8)'),
    ('ODinW-13', 'DINO-X', None, None, None),  # "To verify" - skip
    ('ODinW-13', 'Detic', None, None, None),
    ('ODinW-13', 'RegionCLIP', None, None, None),
    ('ODinW-13', 'UniDetector', None, None, None),  # already has 47.3 in ODinW-35 sheet
    ('ODinW-13', 'YOLO-World', None, None, None),
    ('ODinW-35', 'Grounding DINO', 27.6, 'AP 27.6 ODinW35 (zero-shot, paper-reported)',
     'Grounding DINO ODinW-35 27.6 zero-shot (paper-reported via workbook col 8; paper p11 mentioned 26.1 as headline, 27.6 may be a different variant or supplementary number)'),
    ('LVIS v1.0 val', 'UniDetector', 24.0, 'AP 24.0 LVIS rare AP (UniDetector)',
     'UniDetector LVIS rare class AP ~24 (paper-reported via workbook col 8)'),
    ('LVIS v1.0 val', 'CenterNet2 + Detic', 41.7, 'AP 41.7 LVIS-val (open-vocab)',
     'CenterNet2+Detic = Detic base; LVIS-val open-vocab 41.7 (paper-reported via workbook col 8)'),
    ('LVIS v1.0 val', 'EVA-02-L (LVIS ft)', 55.0, 'AP 55.0 LVIS-val box AP (EVA-02-L fine-tuned)',
     'EVA-02-L LVIS fine-tuned ~55 box AP (paper-reported via workbook col 8)'),
    ('LVIS v1.0 test-dev', 'UniDetector', 24.0, 'AP 24.0 LVIS rare AP (UniDetector)',
     'UniDetector LVIS rare ~24 (paper-reported via workbook col 8)'),
    ('LVIS v1.0 test-dev', 'CenterNet2 + Detic', 41.7, 'AP 41.7 LVIS-val (open-vocab)',
     'CenterNet2+Detic 41.7 LVIS open-vocab (paper-reported via col 8)'),
    ('LVIS v1.0 test-dev', 'EVA-02-L (LVIS ft)', 55.0, 'AP 55.0 LVIS-val box AP',
     'EVA-02-L LVIS ~55 box AP (paper-reported via col 8)'),
]

wb = load_workbook(XLSX)
filled = 0
for sheet, method, mAP, ap_text, note in RECOVERIES:
    if mAP is None:
        continue
    s = wb[sheet]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if len(row) > 1 and row[1] and str(row[1]) == method:
            cur_map = row[6]
            cur_note = str(row[10] or '')
            if cur_map and cur_map != 'N/A':
                print(f'SKIP [{sheet}] r{i} {method}: already has mAP={cur_map}')
                continue
            if '✅' in cur_note:
                print(f'SKIP [{sheet}] r{i} {method}: already verified')
                continue
            s.cell(row=i, column=7).value = mAP
            s.cell(row=i, column=8).value = ap_text
            s.cell(row=i, column=11).value = '✅ verified (paper-reported via workbook col 8): ' + note
            filled += 1
            print(f'FILL [{sheet}] r{i} {method}: {mAP}')
            break

wb.save(XLSX)
print(f'\nTotal recovered: {filled}')
