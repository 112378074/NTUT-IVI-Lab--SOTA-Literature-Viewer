#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_final.py — Final OD verification batch: fill + bulk flag."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

wb = load_workbook(XLSX)
filled = 0
flagged = 0

# === FILLS ===

# ODinW-35: Grounding DINO L 26.1 (paper p11 text)
s = wb['ODinW-35']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if m == 'Grounding DINO' and 'verified' not in cur.lower() and '✅' not in cur:
            s.cell(row=i, column=7).value = 26.1
            s.cell(row=i, column=8).value = 'AP 26.1 ODinW (Swin-L, zero-shot, 35 datasets)'
            s.cell(row=i, column=11).value = '✅ verified (Grounding DINO ECCV24 paper p11)：Grounding DINO L 26.1 AP ODinW zero-shot across 35 datasets [Open-set OD]'
            filled += 1
            print(f'FILL ODinW-35 r{i}: Grounding DINO -> 26.1')

# ODinW-13: UniDetector 47.3 (Table 3)
s = wb['ODinW-13']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if m == 'UniDetector' and '✅' not in cur:
            s.cell(row=i, column=7).value = 47.3
            s.cell(row=i, column=8).value = 'AP 47.3 ODinW13 (zero-shot)'
            s.cell(row=i, column=11).value = '✅ verified (UniDetector CVPR23 Table 3)：UniDetector 47.3 avg AP across 13 ODinW datasets [Universal OD]'
            filled += 1
            print(f'FILL ODinW-13 r{i}: UniDetector -> 47.3')

# === FLAGS ===

# ODinW-35 remaining
s = wb['ODinW-35']
ODINW35_FLAGS = {
    'DINO-X': 'DINO-X paper Figure 2/text only reports COCO + LVIS-minival + LVIS-val zero-shot; ODinW-35 not in main results',
    'YOLO-World': 'YOLO-World CVPR24 paper Table 5 reports LVIS only; ODinW not in main results',
    'Detic': 'Detic CVPR22 paper Table 4 reports open-vocab COCO + LVIS; ODinW not in main results',
    'OWL-ViT': 'OWL-ViT paper does not directly report ODinW-35 (cited in OWLv2 Table 1 as ODinW-13=48.4 only)',
    'OWLv2': 'OWLv2 paper Table 1 only reports ODinW-13 (50.1 best variant); does not report ODinW-35',
    'RegionCLIP': 'RegionCLIP paper does not report ODinW (OWLv2 Table 1 cite shows only LVIS=32.3)',
    'MM-Grounding-DINO': 'arxiv:2404.00000 link is placeholder (not a real arxiv ID). Need correct paper reference',
}
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur: continue
        if m in ODINW35_FLAGS:
            s.cell(row=i, column=11).value = '⚠️ ' + ODINW35_FLAGS[m]
            flagged += 1

# ODinW-13 remaining
s = wb['ODinW-13']
ODINW13_FLAGS = {
    'Grounding DINO': 'Grounding DINO paper Tables 4/6/12-14 report ODinW-35 only (35 datasets); no direct ODinW-13 average in main paper. Cross-cite Grounding DINO L 51.4 ODinW13 from GD 1.5 Pro Table 1',
    'DINO-X': 'DINO-X paper Figure 2 only reports COCO + LVIS; ODinW not in main results',
    'Detic': 'Detic paper does not directly report ODinW-13 average',
    'RegionCLIP': 'RegionCLIP paper does not report ODinW',
    'YOLO-World': 'YOLO-World CVPR24 paper Table 5 reports LVIS only; ODinW-13 not in main results',
}
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur: continue
        if m in ODINW13_FLAGS:
            s.cell(row=i, column=11).value = '⚠️ ' + ODINW13_FLAGS[m]
            flagged += 1

# COCO-O remaining 6
s = wb['COCO-O']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur or '⚠' in cur: continue
        s.cell(row=i, column=11).value = '⚠️ ' + f'{m} not directly tested in COCO-O paper Table 7 (53 models pre-2024). Needs updated COCO-O benchmark or third-party reproduction'
        flagged += 1

# CrowdHuman EOD + YOLOv8x
s = wb['CrowdHuman']
CH_FLAGS = {
    'EOD': 'EOD link arxiv:2309.00000 is placeholder (not real arxiv ID). Need correct paper reference',
    'YOLOv8x': 'YOLOv8x Ultralytics docs do not directly report CrowdHuman AP; needs third-party CrowdHuman benchmark',
}
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur or '⚠' in cur: continue
        if m in CH_FLAGS:
            s.cell(row=i, column=11).value = '⚠️ ' + CH_FLAGS[m]
            flagged += 1

# PASCAL VOC YOLOv8x / YOLO-NAS / Grounding DINO ZS
s = wb['PASCAL VOC 2007']
PV = {
    'YOLOv8x': 'YOLOv8x Ultralytics docs do not directly report PASCAL VOC 2007 closed-set mAP',
    'YOLO-NAS': 'YOLO-NAS (Deci-AI) does not directly report PASCAL VOC 2007 mAP in official materials',
    'Grounding DINO (zero-shot)': 'Grounding DINO paper main benchmarks are COCO/LVIS/ODinW; PASCAL VOC 2007 closed-set not a tested benchmark',
}
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur or '⚠' in cur: continue
        if m in PV:
            s.cell(row=i, column=11).value = '⚠️ ' + PV[m]
            flagged += 1

# CPPE-5 remaining 3
s = wb['CPPE-5']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and not row[6]:
        m = str(row[1])
        cur = str(row[10] or '')
        if '✅' in cur or '⚠' in cur: continue
        if m in ('YOLOv5x','DETR','Cascade R-CNN'):
            s.cell(row=i, column=11).value = '⚠️ ' + f'{m} not in CPPE-5 paper Table 4/5 (which tests SSD/YOLOv3/Faster R-CNN/Deformable DETR/FCOS). Needs third-party reproduction'
            flagged += 1

# LVIS CenterNet2+Detic / EVA-02-L
for sn in ('LVIS v1.0 val', 'LVIS v1.0 test-dev'):
    s = wb[sn]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if len(row) > 1 and row[1] and not row[6]:
            m = str(row[1])
            cur = str(row[10] or '')
            if '✅' in cur or '⚠' in cur: continue
            if m == 'CenterNet2 + Detic':
                s.cell(row=i, column=11).value = '⚠️ CenterNet2+Detic = Detic w/ CenterNet2 backbone; Detic paper Table 5 LVIS open-vocab APr=17.8 (rare-class novel). Overall mAP needs per-paper read'
                flagged += 1
            elif m == 'EVA-02-L (LVIS ft)':
                s.cell(row=i, column=11).value = '⚠️ EVA-02-Det paper Table 7 reports EVA-02-L classification AP=90.0; for LVIS detection, paper Table 8 has mask AP. Specific box AP needs per-paper check'
                flagged += 1

# CAMO-FS bulk flag
s = wb['CAMO-FS']
for i, row in enumerate(s.iter_rows(values_only=True), start=1):
    if len(row) > 1 and row[1] and i > 1 and not row[6]:
        cur = str(row[10] or '')
        if '✅' in cur or '⚠' in cur: continue
        s.cell(row=i, column=11).value = '⚠️ CAMO-FS sheet naming likely wrong (CD-FSOD methods test ArTaxOr/Clipart1k/DIOR/DeepFish/NEU-DET/UODD; no CAMO-FS dataset)'
        flagged += 1

# MS-COCO few-shot 17 remaining
SHOTS = ['MS-COCO 1-shot', 'MS-COCO 5-shot', 'MS-COCO 10-shot', 'MS-COCO 30-shot', 'COCO 2017 FSOD']
SHOT_NOTES = {
    'FSRW': 'FSRW (Kang 2019) Table 2 only tests 10-shot=5.6 and 30-shot=9.1 COCO novel AP; 1/5-shot entries not in original paper',
    'FM-FSOD': 'arxiv:2402.10433 link wrong (Str2Str protein paper, not FM-FSOD). Need correct FM-FSOD paper',
    'FII-DETR': 'FII-DETR (Information Fusion journal) no public arXiv (sciencedirect paywalled); cannot verify',
    'hANMCL': 'arxiv:2404.00700 link wrong (Kleinian groups math paper, not hANMCL). Need correct hANMCL paper',
}
for sn in SHOTS:
    s = wb[sn]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if len(row) > 1 and row[1] and i > 1 and not row[6]:
            m = str(row[1])
            cur = str(row[10] or '')
            if '✅' in cur or '⚠' in cur: continue
            base = m.split('(')[0].strip()
            if base in SHOT_NOTES:
                s.cell(row=i, column=11).value = '⚠️ ' + SHOT_NOTES[base]
                flagged += 1

# UGRAN @ 10 SOD sheets
SOD = ['DUTS-TE','DUT-OMRON','ECSSD','HKU-IS','PASCAL-S','HRSOD','DAVIS-S','UHRSD','SBU-Refine','ISTD']
for sn in SOD:
    s = wb[sn]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if len(row) > 1 and row[1] and i > 1 and not row[6]:
            m = str(row[1])
            cur = str(row[10] or '')
            if '✅' in cur or '⚠' in cur: continue
            if m == 'UGRAN':
                s.cell(row=i, column=11).value = '⚠️ UGRAN (arxiv 2407.13680) PDF text extraction failed — tables rendered as images / special encoding (pypdf returns 0 decimals across all pages). Needs manual paper read or OCR to fill S-measure/MAE'
                flagged += 1

wb.save(XLSX)
print(f'\nTotal filled: {filled}, flagged: {flagged}')
