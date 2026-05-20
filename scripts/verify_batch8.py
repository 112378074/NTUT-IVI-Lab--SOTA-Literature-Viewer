#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_batch8.py — 17 fills + 12 removes for AD."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# (sheet, row_1based, i_auroc, p_auroc, p_ap, p_pro, note)
FILLS = [
    # PRN @ MVTec AD (already has 99.4 — confirm + mark verified)
    ('MVTec AD', 24, 99.4, 99.0, 78.6, 96.1,
     '✅ 已驗證 (PRN CVPR2023 Table 1+2 average)：Prototypical Residual Network；多尺度殘差表徵+異常生成 [監督式異常定位]'),
    # PRN @ BTAD (Table 8)
    ('BTAD', 22, 94.7, 97.1, 54.0, 78.0,
     '✅ 已驗證 (PRN CVPR2023 Table 8)：Prototypical Residual Network；BTAD I=94.7 P=97.1 [監督式]'),
    # Anomaly-OV (TZS) zero-shot Table 2
    ('MVTec AD', 104, 94.0, None, None, None,
     '✅ 已驗證 (Anomaly-OV CVPR25 Table 2, zero-shot)：MLLM-based ZSAD；Look-Twice Feature Matching [基於 MLLM]'),
    ('VisA',     77, 91.1, None, None, None,
     '✅ 已驗證 (Anomaly-OV CVPR25 Table 2, zero-shot)：MLLM-based ZSAD [基於 MLLM]'),
    ('MPDD',     24, 81.7, None, None, None,
     '✅ 已驗證 (Anomaly-OV CVPR25 Table 2, zero-shot) [基於 MLLM]'),
    ('BTAD',     30, 89.0, None, None, None,
     '✅ 已驗證 (Anomaly-OV CVPR25 Table 2, zero-shot) [基於 MLLM]'),
    # BSMB (Beyond Single-Modal Boundary) Table 1 zero-shot AUROC
    ('MVTec AD', 105, 92.3, None, 96.3, None,
     '✅ 已驗證 (BSMB CVPR25 Table 1, zero-shot)：跨模態 Transferable Visual Prototype；MVTec/VisA/BTAD/MPDD/3D 通用 [跨模態 ZSAD]'),
    ('VisA',     78, 86.8, None, 89.3, None,
     '✅ 已驗證 (BSMB CVPR25 Table 1, zero-shot)：跨模態 TVP [跨模態 ZSAD]'),
    ('MPDD',     25, 72.8, None, 79.7, None,
     '✅ 已驗證 (BSMB CVPR25 Table 1, zero-shot) [跨模態 ZSAD]'),
    ('BTAD',     31, 97.5, None, 98.9, None,
     '✅ 已驗證 (BSMB CVPR25 Table 1, zero-shot) [跨模態 ZSAD]'),
    ('MVTec 3D', 25, 79.6, None, 93.1, None,
     '✅ 已驗證 (BSMB CVPR25 Table 1, zero-shot) [跨模態 ZSAD]'),
    # Bayes-PFL (BPFL) Table 1 zero-shot
    ('MVTec AD', 109, 92.3, 91.8, 48.3, None,
     '✅ 已驗證 (Bayes-PFL CVPR25 Table 1, zero-shot)：Bayesian Prompt Flow Learning；多模態 prompt distribution [基於 VLM]'),
    ('VisA',     82, 87.0, 95.6, 29.8, None,
     '✅ 已驗證 (Bayes-PFL CVPR25 Table 1, zero-shot)：[基於 VLM]'),
    ('BTAD',     33, 93.2, 93.9, 47.1, None,
     '✅ 已驗證 (Bayes-PFL CVPR25 Table 1, zero-shot)：[基於 VLM]'),
    # INP-Former / "Exploring Intrinsic Normal Prototypes" Table 2 multi-class
    ('MVTec AD', 106, 99.7, 98.5, 71.0, 94.9,
     '✅ 已驗證 (INP-Former CVPR25 Table 2 multi-class)：Intrinsic Normal Prototypes 單張影像內提取；ViT-B/14 DINOv2 [基於重構]'),
    ('VisA',     79, 98.9, 98.9, 51.2, 94.4,
     '✅ 已驗證 (INP-Former CVPR25 Table 2 multi-class)：[基於重構]'),
    # LogSAD / TFA on MVTec LOCO full-data
    ('MVTec LOCO', 26, 90.2, None, None, None,
     '✅ 已驗證 (LogSAD CVPR25 Table 1, full-data)：Multi-granularity (patch + interest + composition) training-free VLM AD；MVTec LOCO full-data I=90.2 [基於 VLM]'),
]

# (sheet, row_1based, reason)
DELETES = [
    # ISVL — VAND 2025 Challenge winner; only tested on MVTec AD 2 private/mixed test set
    ('MVTec AD',   102, 'ISVL is VAND 2025 Challenge Category 1 winner — only evaluated on MVTec AD 2 (TESTpriv/TESTpriv_mix). Not tested on standard MVTec AD.'),
    ('VisA',        76, 'ISVL only evaluated on MVTec AD 2'),
    ('MVTec 3D',    23, 'ISVL only evaluated on MVTec AD 2'),
    ('MVTec LOCO',  24, 'ISVL only evaluated on MVTec AD 2 (VAND 2025 Category 1)'),
    # SuperAD — VAND 2025 challenge submission; only MVTec AD 2
    ('MVTec AD',   101, 'SuperAD is a VAND 2025 Challenge submission — only evaluated on MVTec AD 2 TESTpriv/mixed'),
    ('VisA',        75, 'SuperAD only evaluated on MVTec AD 2'),
    # CSAD — paper only evaluates MVTec LOCO
    ('MVTec AD',    90, 'CSAD (BMVC2024) only evaluates MVTec LOCO AD; MVTec/VisA listed in intro but not tested'),
    ('VisA',        72, 'CSAD only evaluates MVTec LOCO AD'),
    # TZS @ MVTec 3D — Anomaly-OV Table 2 covers MVTec/VisA/AITEX/ELPV/BTAD/MPDD + medical, NOT MVTec 3D-AD
    ('MVTec 3D',    24, 'Anomaly-OV (TZS) Table 2 does not include MVTec 3D-AD in zero-shot benchmark'),
    # LogSAD @ MVTec AD/VisA — only 4-shot tested, not full-data unsupervised
    ('MVTec AD',   108, 'LogSAD only tests MVTec AD in 4-shot setting (Table 4), not full-data unsupervised. Not comparable to unsupervised AD ranking.'),
    ('VisA',        81, 'LogSAD only tests VisA in 4-shot setting (Table 4)'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []

    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} → i={i}, p={p}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason[:60]}...)')

    wb.save(XLSX)
    print('Saved', XLSX)
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
