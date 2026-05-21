#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_batch2.py — more COCO AP fills from DEIM/RT-DETRv3/DEIMv2/YOLOX PDFs."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

FILLS = [
    # DEIM-D-FINE-L (DEIM-L) on COCO val2017 (DEIM Table 1)
    ('COCO 2017 val', 18, 54.7, 'AP 54.7 val',
     '✅ 已驗證 (DEIM CVPR25 Table 1, COCO val2017)：DEIM-D-FINE-L 54.7% AP, 8.07ms [End-to-end RT-OD]'),
    ('COCO 2017', 17, 54.7, 'AP 54.7 val2017',
     '✅ 已驗證 (DEIM CVPR25 Table 1, COCO val2017)：DEIM-D-FINE-L 54.7% AP [End-to-end RT-OD]'),
    # DEIMv2 (DINOv3) — DEIMv2-X (largest) on COCO val2017 (DEIMv2 Table 3)
    ('COCO 2017 val', 16, 57.8, 'AP 57.8 val (DEIMv2-X)',
     '✅ 已驗證 (DEIMv2 arXiv25 Table 3, COCO val2017)：DEIMv2-X 57.8% AP, DINOv3 backbone [End-to-end RT-OD]'),
    ('COCO 2017', 15, 57.8, 'AP 57.8 val2017 (DEIMv2-X)',
     '✅ 已驗證 (DEIMv2 arXiv25 Table 3, COCO val2017)：DEIMv2-X 57.8% AP, DINOv3 backbone [End-to-end RT-OD]'),
    # RT-DETRv3-R101 on COCO val2017 (RT-DETRv3 Table 1, 6x schedule)
    ('COCO 2017 val', 20, 54.6, 'AP 54.6 val (6x)',
     '✅ 已驗證 (RT-DETRv3 Table 1, COCO val2017 6x)：RT-DETRv3-R101 54.6% AP, 13.5ms [End-to-end RT-OD]'),
    ('COCO 2017', 20, 54.6, 'AP 54.6 val2017 (6x)',
     '✅ 已驗證 (RT-DETRv3 Table 1, COCO val2017 6x)：RT-DETRv3-R101 54.6% AP [End-to-end RT-OD]'),
    # RT-DETRv3 (general, no specific backbone) → use R101 as headline number
    ('COCO 2017 val', 42, 54.6, 'AP 54.6 val (R101 6x)',
     '✅ 已驗證 (RT-DETRv3 Table 1)：headline number from R101 6x setting [End-to-end RT-OD]'),
    ('COCO 2017', 94, 54.6, 'AP 54.6 val2017 (R101 6x)',
     '✅ 已驗證 (RT-DETRv3 Table 1)：headline number from R101 6x setting [End-to-end RT-OD]'),
    # YOLOX-X on COCO val (YOLOX Table 3)
    ('COCO 2017', 44, 51.2, 'AP 51.2 val',
     '✅ 已驗證 (YOLOX arXiv21 Table 3, COCO val)：YOLOX-X 51.2% AP, 99.1M params [Real-Time OD]'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, mAP_v, AP_text, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=2).value
        s.cell(row=row, column=7).value = mAP_v
        s.cell(row=row, column=8).value = AP_text
        s.cell(row=row, column=11).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> mAP={mAP_v}')
    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))

if __name__ == '__main__':
    main()
