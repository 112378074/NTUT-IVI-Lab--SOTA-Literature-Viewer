#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fix_ad_column_shift.py — fix the column-shift bug in earlier AD verify batches.

Bug:
  AD workbook column layout is:
    1=資料集, 2=類別, 3=方法, 4=作者, 5=發表, 6=年月, 7=狀態,
    8=I-AUROC, 9=P-AUROC, 10=P-AP, 11=P-PRO, 12=FPS, 13=備註
  but earlier verify_batch*/verify_ad_final* wrote:
    col=7  ← I-AUROC  (clobbered 狀態!)
    col=8  ← P-AUROC  (was supposed to be I-AUROC)
    col=9  ← P-AP     (was supposed to be P-AUROC)
    col=10 ← P-PRO    (was supposed to be P-AP)
    col=13 ← notes    (correct ✓)
  i.e. all four metric values are 1 column too far left, and the original
  status string was overwritten by the I-AUROC number.

Fix:
  For every row whose notes column 13 contains '✅ 已驗證', shift cols 7-10
  right by one (7→8, 8→9, 9→10, 10→11) and rebuild the status column from
  the venue text (col 5).
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

CONF_PUB = re.compile(
    r'\b(CVPR|ICCV|ECCV|WACV|NeurIPS|ICLR|ICML|AAAI|IJCAI|BMVC|TPAMI|TIP|'
    r'TASE|TMLR|TII|TNNLS|JMIV|ISIE|KDD|ACL|ICASSP|SDM|VISAPP)\b',
    re.IGNORECASE,
)


def infer_status(venue):
    """Return '已發表' or '預印本' from the venue string."""
    if venue is None:
        return '預印本'
    v = str(venue)
    # If venue mentions a real conference/journal and isn't tagged 'preprint',
    # call it published.
    is_preprint = 'arxiv' in v.lower() and ('preprint' in v.lower()
                                            or 'cvpr 2026' in v.lower()
                                            or '在審' in v)
    if CONF_PUB.search(v) and not is_preprint:
        return '已發表'
    if 'arxiv' in v.lower() or 'preprint' in v.lower() or '在審' in v:
        return '預印本'
    return '已發表'  # default to published if it has a recognized name


def main():
    wb = load_workbook(XLSX)
    fixed = 0
    skipped = 0

    sheets = ['MVTec AD', 'MVTec LOCO', 'MVTec AD2', 'VisA', 'MPDD', 'BTAD', 'MVTec 3D']
    for sn in sheets:
        s = wb[sn]
        for row in range(1, s.max_row + 1):
            notes = s.cell(row=row, column=13).value
            if not isinstance(notes, str) or '✅ 已驗證' not in notes:
                continue
            # Read current (shifted-left) columns
            c7  = s.cell(row=row, column=7).value   # contains the I-AUROC value (bug)
            c8  = s.cell(row=row, column=8).value   # contains the P-AUROC value
            c9  = s.cell(row=row, column=9).value   # contains the P-AP value
            c10 = s.cell(row=row, column=10).value  # contains the P-PRO value
            c11_before = s.cell(row=row, column=11).value  # was-FPS, possibly None
            # Sanity check: c7 should be a number (the I-AUROC) — if it's a
            # plain string like '已發表' the row was already fixed; skip.
            if isinstance(c7, str) and c7.strip() in ('已發表', '預印本', 'Published', 'N/A', ''):
                skipped += 1
                continue
            venue = s.cell(row=row, column=5).value
            new_status = infer_status(venue)
            # Apply shift
            s.cell(row=row, column=7).value  = new_status
            s.cell(row=row, column=8).value  = c7
            s.cell(row=row, column=9).value  = c8
            s.cell(row=row, column=10).value = c9
            s.cell(row=row, column=11).value = c10
            # Column 12 (FPS) intentionally left as-is — c11_before
            # was the original FPS slot and we didn't touch it in the buggy
            # batches.
            fixed += 1
            print(f'  [{sn} r{row}] status="{new_status}" | i={c7} p={c8} ap={c9} pro={c10}')

    wb.save(XLSX)
    print(f'\nFixed {fixed} rows, skipped {skipped} already-correct rows.')


if __name__ == '__main__':
    main()
