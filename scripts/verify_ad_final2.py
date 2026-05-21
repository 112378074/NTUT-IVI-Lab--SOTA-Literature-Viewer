#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_ad_final2.py — finishing AD: 6 fills + 9 removes."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

FILLS = [
    # DiAD @ MVTec 3D (Table 5)
    ('MVTec 3D', 21, 84.6, 96.4, 25.3, 87.8,
     '✅ 已驗證 (DiAD AAAI24 Table 5)：Diffusion-based multi-class AD；MVTec 3D I=84.6 P=96.4 AP=25.3 PRO=87.8 [基於擴散]'),
    # AMI-Net @ BTAD (Table II)
    ('BTAD', 22, 95.1, 97.1, None, None,
     '✅ 已驗證 (AMI-Net IEEE TASE24 Table II)：Adaptive Masked Inpainting；BTAD Average I=95.1 P=97.1 [基於重構]'),
    # InvAD-Diff @ MPDD (Table 3 single-class)
    ('MPDD', 18, 96.5, 98.3, 40.1, 91.2,
     '✅ 已驗證 (InvAD-Diff CVPR26投 Table 3 single-class)：Inversion-based Diffusion；MPDD I=96.5 P=98.3 AP=40.1 PRO=91.2 [基於擴散+重構]'),
    # UniMMAD @ MVTec 3D (Table 14 image + Table 15 pixel)
    ('MVTec 3D', 17, 92.5, 99.0, 44.1, 95.9,
     '✅ 已驗證 (UniMMAD arXiv25 Tables 14+15)：MoE Multi-Modal AD；MVTec 3D I=92.5 P=99.0 AP=44.1 PRO=95.9 [多模態 MoE]'),
    # ADL @ MVTec AD (Table 1, 10% data contamination — fair comp setting)
    ('MVTec AD', 102, 92.4, None, None, None,
     '✅ 已驗證 (ADL WACV25 Table 1, 10% contamination)：Adaptive Deviation Learning under contaminated data；MVTec AD avg AUROC 0.924 [基於監督式 + Noisy normal]'),
    # SSASDM (DTU-Net) @ MVTec AD (Table 2)
    ('MVTec AD', 100, 99.17, None, None, None,
     '✅ 已驗證 (SSASDM WACV25 Table 2)：DTU-Net 自監督擴散+動態 Transformer UNet；MVTec AD 15-class avg AUROC 99.17 [基於擴散自監督]'),
]

DELETES = [
    # SINBAD @ MVTec AD — paper only tests MVTec LOCO AD (Table 6 in appendix)
    ('MVTec AD', 91, 'SINBAD ICML24 Table 6 only tests MVTec LOCO AD (logical anomaly), not MVTec AD'),
    # BTF @ MVTec AD — paper only tests MVTec 3D-AD (multimodal RGB+3D)
    ('MVTec AD', 92, 'BTF "Back to the Feature" CVPRW23 only evaluates MVTec 3D-AD multimodal, not MVTec AD'),
    # RoBiS @ MVTec AD — VAND 2025 Cat-1 paper, only MVTec AD 2
    ('MVTec AD', 93, 'RoBiS is VAND 2025 Cat-1 submission (Table 1 is MVTec AD 2 TESTpriv/mix), not standard MVTec AD'),
    # Triad @ MVTec LOCO — paper only tests MVTec AD + WFDD (Table 1)
    ('MVTec LOCO', 19, 'Triad ICCV25 Table 1 only tests MVTec-AD + WFDD, not MVTec LOCO'),
    # RGBias @ MVTec AD — paper only tests CIFAR-10/100 (one-class semantic AD)
    ('MVTec AD', 101, 'RGBias WACV25 Table 1 only tests CIFAR-10/100 one-class semantic AD, not MVTec AD industrial'),
    # U-Flow @ BTAD — paper only tests MVTec AD (no BTAD mention found)
    ('BTAD', 23, 'U-Flow paper Table 1 only evaluates MVTec AD pixel AUROC 98.74%, no BTAD result'),
    # VT-ADL @ MVTec AD — paper Table III reports PRO score only, not I-AUROC ranking metric
    ('MVTec AD', 89, 'VT-ADL ISIE21 Table III reports PRO score only (MVTec mean 80.7), not I-AUROC; primary BTAD baseline'),
    # TailedCore @ MVTec AD + VisA — long-tail noisy specialized settings (Pareto/step K=1/K=4), not standard
    ('MVTec AD', 99, 'TailedCore CVPR25 evaluates only under long-tail noisy settings (Pareto, step K=1/K=4); not directly comparable to standard ranking'),
    ('VisA', 77, 'TailedCore CVPR25 specialized long-tail noisy setting (Tables 1-2 across multiple tail types); not standard'),
    # SLDFC @ MVTec AD — paper Table 1 only 5 texture classes (not full 15)
    ('MVTec AD', 103, 'SLDFC WACV25 Table 1 only evaluates 5 texture classes of MVTec AD (mean 99.6 textures); not full 15-class ranking'),
]


def main():
    wb = load_workbook(XLSX)
    summary = []
    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} -> i={i}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason[:60]}...)')

    wb.save(XLSX)
    print('Saved')
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
