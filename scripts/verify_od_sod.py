#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_od_sod.py — Fix SOD sheets: replace duplicated wrong values with per-dataset values from papers.

ROOT CAUSE: All 10 SOD sheets (DUTS-TE, DUT-OMRON, ECSSD, HKU-IS, PASCAL-S, HRSOD,
DAVIS-S, UHRSD, SBU-Refine, ISTD) currently show IDENTICAL S-measure values for each
method (e.g., BiRefNet shows S=0.939 across all 10 datasets). This is from a copy-paste
error in the original sheet — each method actually reports DIFFERENT S-measure per
dataset.

This script:
  1. Replaces with correct per-dataset values for BiRefNet (its own paper Table 5)
  2. Replaces with correct per-dataset values for InSPyReNet (its own paper Table 2/3)
  3. Replaces VST, BASNet values from InSPyReNet's citation Table 2 (uniform protocol)
  4. Flags SBU-Refine and ISTD as not-in-SOD-scope (these are shadow detection benchmarks)
  5. Flags remaining methods (MVANet, M3Net, ICON-S, EDN, PoolNet+, MINet, U^2-Net) as
     needing per-paper extraction.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'

# =============================================================================
# VERIFIED VALUES per (sheet, method) -> (S_alpha, full_text, note_suffix)
# =============================================================================
# BiRefNet (Zheng et al. CAAI AI 2024, Table 5)
#   Train set 1 (DUTS-TR only): DAVIS-S/HRSOD-TE/UHRSD-TE/DUTS-TE/DUT-OMRON
#   Train set 1,2,3 (best multi-set): same datasets
#   Paper does NOT evaluate on ECSSD/HKU-IS/PASCAL-S/SBU-Refine/ISTD
# We use Train set 1,2,3 (best reported config)
BIREFNET_DATA = {
    'DAVIS-S':   (0.975, 'Sα 0.975 / Fmax 0.979 / Em 0.989 / MAE 0.006 (BiRefNet TR=D,H,U, paper Table 5)'),
    'HRSOD':     (0.962, 'Sα 0.962 / Fmax 0.961 / Em 0.973 / MAE 0.013 (BiRefNet TR=D,H,U, paper Table 5 HRSOD-TE)'),
    'UHRSD':     (0.957, 'Sα 0.957 / Fmax 0.963 / Em 0.969 / MAE 0.016 (BiRefNet TR=D,H,U, paper Table 5 UHRSD-TE)'),
    'DUTS-TE':   (0.944, 'Sα 0.944 / Fmax 0.943 / Em 0.962 / MAE 0.018 (BiRefNet TR=D,H,U, paper Table 5)'),
    'DUT-OMRON': (0.882, 'Sα 0.882 / Fmax 0.839 / Em 0.896 / MAE 0.038 (BiRefNet TR=D,H,U, paper Table 5)'),
}

# InSPyReNet (Kim et al. ACCV 2022, SwinB backbone, Table 2 + Table 3)
INSPYRENET_DATA = {
    'DUTS-TE':   (0.931, 'Sα 0.931 / Fmax 0.927 / MAE 0.024 (InSPyReNet SwinB, paper Table 2)'),
    'DUT-OMRON': (0.875, 'Sα 0.875 / Fmax 0.832 / MAE 0.045 (InSPyReNet SwinB, paper Table 2)'),
    'ECSSD':     (0.949, 'Sα 0.949 / Fmax 0.960 / MAE 0.023 (InSPyReNet SwinB, paper Table 2)'),
    'HKU-IS':    (0.944, 'Sα 0.944 / Fmax 0.955 / MAE 0.021 (InSPyReNet SwinB, paper Table 2)'),
    'PASCAL-S':  (0.893, 'Sα 0.893 / Fmax 0.893 / MAE 0.048 (InSPyReNet SwinB, paper Table 2)'),
    # HR benchmarks - Table 3, DUTS-TR training
    'DAVIS-S':   (0.946, 'Sα 0.946 / Fmax 0.940 / MAE 0.014 (InSPyReNet SwinB, paper Table 3, TR=D)'),
    'HRSOD':     (0.930, 'Sα 0.930 / Fmax 0.926 / MAE 0.025 (InSPyReNet SwinB, paper Table 3 HRSOD-TE, TR=D)'),
    'UHRSD':     (0.925, 'Sα 0.925 / Fmax 0.931 / MAE 0.029 (InSPyReNet SwinB, paper Table 3 UHRSD-TE, TR=D)'),
}

# VST (Liu et al. ICCV 2021, T2T-ViT-14) - values from InSPyReNet Table 2 (citation, uniform protocol)
VST_DATA = {
    'DUTS-TE':   (0.896, 'Sα 0.896 / Fmax 0.878 / MAE 0.037 (VST T2T-ViT-14, cited in InSPyReNet Table 2)'),
    'DUT-OMRON': (0.850, 'Sα 0.850 / Fmax 0.800 / MAE 0.058 (VST T2T-ViT-14, cited in InSPyReNet Table 2)'),
    'ECSSD':     (0.932, 'Sα 0.932 / Fmax 0.944 / MAE 0.033 (VST T2T-ViT-14, cited in InSPyReNet Table 2)'),
    'HKU-IS':    (0.928, 'Sα 0.928 / Fmax 0.937 / MAE 0.029 (VST T2T-ViT-14, cited in InSPyReNet Table 2)'),
    'PASCAL-S':  (0.872, 'Sα 0.872 / Fmax 0.864 / MAE 0.061 (VST T2T-ViT-14, cited in InSPyReNet Table 2)'),
}

# BASNet (Qin et al. CVPR 2019, ResNet-34) - values from InSPyReNet Table 2
BASNET_DATA = {
    'DUTS-TE':   (0.866, 'Sα 0.866 / Fmax 0.838 / MAE 0.048 (BASNet ResNet-34, cited in InSPyReNet Table 2)'),
    'DUT-OMRON': (0.836, 'Sα 0.836 / Fmax 0.779 / MAE 0.056 (BASNet ResNet-34, cited in InSPyReNet Table 2)'),
    'ECSSD':     (0.916, 'Sα 0.916 / Fmax 0.931 / MAE 0.037 (BASNet ResNet-34, cited in InSPyReNet Table 2)'),
    'HKU-IS':    (0.909, 'Sα 0.909 / Fmax 0.919 / MAE 0.032 (BASNet ResNet-34, cited in InSPyReNet Table 2)'),
    'PASCAL-S':  (0.838, 'Sα 0.838 / Fmax 0.835 / MAE 0.076 (BASNet ResNet-34, cited in InSPyReNet Table 2)'),
}

# =============================================================================
# Method -> verified-data mapping
# =============================================================================
VERIFIED = {
    'BiRefNet':   (BIREFNET_DATA, 'BiRefNet ECCV24 paper Table 5'),
    'InSPyReNet': (INSPYRENET_DATA, 'InSPyReNet ACCV22 paper Tables 2-3 (SwinB backbone)'),
    'VST':        (VST_DATA, 'VST ICCV21, values cited in InSPyReNet Table 2'),
    'BASNet':     (BASNET_DATA, 'BASNet CVPR19, values cited in InSPyReNet Table 2'),
}

# Methods to flag (need per-paper extraction or out-of-scope)
FLAG_METHODS = {
    'MVANet':   'MVANet (CVPR24) paper PDF text-extraction failed (tables likely rendered as images). Original sheet showed identical S=0.927 across all 10 SOD datasets, which is duplicated/wrong. Needs per-dataset value from MVANet paper Table',
    'M3Net':    'M3Net (Mixed-Multi-level Mixed-attention Net) paper has tables on p6-7 but original sheet showed identical S=0.918 across all 10 SOD datasets (duplicated/wrong). Needs per-dataset extraction',
    'ICON-S':   'ICON-S (Zhuge et al., TPAMI 2023) paper text-extraction failed. Original sheet showed identical S=0.917 across all 10 SOD datasets (duplicated/wrong)',
    'EDN':      'EDN (Wu et al., TIP 2022) paper has tables but original sheet showed identical S=0.892 across all 10 SOD datasets (duplicated/wrong). Needs per-dataset extraction',
    'PoolNet+': 'PoolNet+ (Liu et al., TPAMI 2022 — note: + version) original sheet showed identical S=0.886 across all 10 SOD datasets (duplicated/wrong). Differs from original PoolNet (CVPR19); needs per-dataset extraction from PoolNet+ paper',
    'MINet':    'MINet (Pang et al., CVPR 2020) original sheet showed identical S=0.884 across all 10 SOD datasets (duplicated/wrong). Per InSPyReNet Table 2: DUTS-TE Sα=0.896, OMRON Sα=0.843, ECSSD Sα=0.931, HKU-IS Sα=0.923, PASCAL-S Sα=0.865 — needs careful per-row assignment',
    'U^2-Net':  'U2-Net (Qin et al., PR 2020) original sheet showed identical S=0.874 across all 10 SOD datasets (duplicated/wrong). Needs per-dataset extraction from U2-Net paper',
}

# Datasets that are NOT SOD benchmarks (shadow detection / mismatched)
NON_SOD_DATASETS = {
    'SBU-Refine': 'SBU-Refine is a SHADOW DETECTION dataset (Hu et al. 2021), not SOD. The SOD methods listed (BiRefNet/InSPyReNet/MVANet/etc.) do not evaluate on SBU-Refine — sheet needs reclassification to shadow detection or removal',
    'ISTD':       'ISTD (Image Shadow Triplets Dataset, Wang et al. CVPR18) is a SHADOW DETECTION dataset, not SOD. The SOD methods listed do not evaluate on ISTD — sheet needs reclassification or removal',
}

SOD_SHEETS = ['DUTS-TE','DUT-OMRON','ECSSD','HKU-IS','PASCAL-S','HRSOD','DAVIS-S','UHRSD','SBU-Refine','ISTD']

wb = load_workbook(XLSX)
filled = 0
flagged = 0
ood = 0  # out-of-domain (SBU/ISTD)

for sn in SOD_SHEETS:
    s = wb[sn]
    is_non_sod = sn in NON_SOD_DATASETS
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not (len(row) > 1 and row[1]): continue
        method = str(row[1]).strip()
        cur_note = str(row[10] or '') if len(row) > 10 else ''
        if '✅' in cur_note:
            continue  # already verified

        # Case 1: Non-SOD sheet — flag every row with cross-domain note
        if is_non_sod:
            s.cell(row=i, column=11).value = '⚠️ ' + NON_SOD_DATASETS[sn] + f'. (Method "{method}" is SOD, not applicable here)'
            # Also clear the wrong duplicate value in col 8
            ap_cell = s.cell(row=i, column=8).value or ''
            if 'S ' in str(ap_cell) and 'F' in str(ap_cell):
                s.cell(row=i, column=8).value = 'N/A (out-of-domain: shadow detection benchmark, not SOD)'
            ood += 1
            continue

        # Case 2: Verified method with per-dataset value
        if method in VERIFIED:
            data, src = VERIFIED[method]
            if sn in data:
                sa, full_text = data[sn]
                s.cell(row=i, column=7).value = sa
                s.cell(row=i, column=8).value = full_text
                s.cell(row=i, column=11).value = f'✅ verified ({src}): {sn} Sα={sa}'
                filled += 1
                print(f'FILL [{sn}] r{i} {method}: Sα={sa}')
            else:
                # method exists in verified but not for this dataset
                s.cell(row=i, column=8).value = f'N/A (not evaluated in {src})'
                s.cell(row=i, column=11).value = f'⚠️ {method} does not report on {sn} in its paper. {src} covers other datasets only'
                ap_cell = s.cell(row=i, column=8).value or ''
                flagged += 1
            continue

        # Case 3: Method needs flagging
        if method in FLAG_METHODS:
            s.cell(row=i, column=11).value = '⚠️ ' + FLAG_METHODS[method]
            flagged += 1
            continue

wb.save(XLSX)
print(f'\nTotal: filled={filled}, flagged={flagged}, out-of-domain={ood}')
