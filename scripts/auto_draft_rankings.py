#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
auto_draft_rankings.py  —  Option 1: automated draft ranking extraction.

For every paper in the All-Papers sheets (AD / OD / CLS) that is not already
placed into the per-dataset ranking sheets, this script:
  1. downloads the paper PDF (arXiv or CVF Open Access),
  2. detects which TARGET datasets the paper evaluates on (reliable keyword
     match on the full text),
  3. best-effort extracts the proposed method's primary metric per dataset
     from the results tables (PyMuPDF find_tables),
  4. appends a DRAFT row to each detected dataset sheet, flagged
     '⚠️ 自動抽取・未驗證' in the notes column.

Rows already flagged '✅ 已驗證' (hand-verified) are never touched.
This is a long-running batch job — run it in the background.

Usage:
    python scripts/auto_draft_rankings.py --domain ad            # one domain
    python scripts/auto_draft_rankings.py --domain ad --limit 5  # test
    python scripts/auto_draft_rankings.py --all                  # ad+od+cls
    python scripts/auto_draft_rankings.py --all --regen          # + regen html
"""
import argparse, json, re, ssl, sys, time, urllib.request
from copy import copy
from pathlib import Path
from openpyxl import load_workbook

try:
    import certifi
    SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except Exception:
    SSL_CTX = ssl.create_default_context()

PROJECT = Path(__file__).resolve().parent.parent
AD_XLSX  = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'
OD_XLSX  = PROJECT / 'Object_Detection_Papers_Ranking_2021_2026.xlsx'
CLS_XLSX = PROJECT / 'Image_Classification_Papers_Ranking_2021_2026.xlsx'
LOG      = PROJECT / 'scripts' / 'auto_draft_log.txt'
STATE    = PROJECT / 'scripts' / '.auto_draft_state.json'
UA = 'NTUT-AIL-PaperBot/1.0 (literature review; azaz31855@gmail.com)'

DRAFT_TAG = '⚠️ 自動抽取・指標待驗證 (已偵測使用此資料集)'
VERIFIED_TAG = '✅ 已驗證'

# ---- target datasets per domain: (sheet_name, [regex patterns]) ----
AD_DATASETS = [
    ('MVTec AD 2', [r'mvtec\s*ad\s*2', r'mvtec\s*ad2']),
    ('MVTec LOCO', [r'mvtec\s*loco', r'\bloco\b']),
    ('MVTec 3D',   [r'mvtec\s*3d']),
    ('MVTec AD',   [r'mvtec[\s-]*ad\b', r'\bmvtec\b']),
    ('VisA',       [r'\bvisa\b']),
    ('MPDD',       [r'\bmpdd\b']),
    ('BTAD',       [r'\bbtad\b', r'beantech']),
]
OD_DATASETS = [
    ('COCO test-dev', [r'coco\s*test-?dev']),
    ('COCO 2017 val', [r'coco\s*2017\s*val', r'val2017']),
    ('COCO 2017',     [r'coco\s*2017', r'ms-?coco', r'\bcoco\b']),
    ('PASCAL VOC 2007',[r'pascal\s*voc', r'\bvoc\s*2007\b']),
    ('LVIS v1.0 val', [r'\blvis\b']),
    ('CrowdHuman',    [r'crowdhuman']),
    ('DUTS-TE',       [r'duts-te', r'\bduts\b']),
    ('DUT-OMRON',     [r'dut-omron']),
    ('ECSSD',         [r'\becssd\b']),
    ('HKU-IS',        [r'hku-is']),
    ('ODinW-13',      [r'odinw-?13']),
    ('ODinW-35',      [r'odinw-?35']),
]
CLS_DATASETS = [
    ('ImageNet-1K',    [r'imagenet-?1k', r'imagenet\b']),
    ('ImageNet-V2',    [r'imagenet-?v2']),
    ('CIFAR-100',      [r'cifar-?100']),
    ('CIFAR-10',       [r'cifar-?10\b']),
    ('CUB-200-2011',   [r'cub-?200', r'\bcub\b']),
    ('Stanford Cars',  [r'stanford\s*cars']),
    ('FGVC-Aircraft',  [r'fgvc-?aircraft', r'\baircraft\b']),
    ('Food-101',       [r'food-?101']),
    ('STL-10',         [r'stl-?10']),
    ('miniImageNet 5w-1s', [r'mini-?imagenet']),
]


def log(msg):
    line = f'[{time.strftime("%Y-%m-%d %H:%M:%S")}] {msg}'
    print(line, flush=True)
    with open(LOG, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def fetch_pdf(link, retries=2):
    """Resolve a paper link to a PDF URL and download the bytes."""
    if not link or link in ('N/A', '—'):
        return None
    pdf_url = None
    if link.endswith('.pdf'):
        pdf_url = link
    elif 'openaccess.thecvf.com' in link and '/html/' in link:
        pdf_url = link.replace('/html/', '/papers/').replace('_paper.html', '_paper.pdf')
    else:
        m = re.search(r'arxiv\.org/abs/(\d+\.\d+)', link)
        if m:
            pdf_url = f'https://arxiv.org/pdf/{m.group(1)}'
    if not pdf_url:
        return None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(pdf_url, headers={'User-Agent': UA})
            with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as r:
                return r.read()
        except Exception as e:
            if attempt == 0:
                time.sleep(6)
                continue
            log(f'    pdf fetch failed: {e}')
    return None


def detect_datasets(text, ds_list):
    found = []
    tl = text.lower()
    for sheet, pats in ds_list:
        if any(re.search(p, tl) for p in pats):
            found.append(sheet)
    # Keep only the most specific MVTec variant if several match
    if 'MVTec AD' in found and ('MVTec AD 2' in found):
        pass
    return found


def extract_primary_metric(doc, method, domain):
    """Best-effort: scan tables for a row matching the method / 'Ours' and
    return the largest plausible primary-metric value found in that row.
    Returns None when nothing confident is found."""
    mkey = re.sub(r'[^a-z0-9]', '', (method or '').lower())[:12]
    best = None
    for page in doc:
        try:
            tabs = page.find_tables()
        except Exception:
            continue
        for tab in tabs.tables:
            try:
                grid = tab.extract()
            except Exception:
                continue
            for row in grid:
                label = ' '.join(str(c) for c in row[:2] if c).lower()
                if not label:
                    continue
                if 'ours' in label or (mkey and mkey in re.sub(r'[^a-z0-9]', '', label)):
                    nums = []
                    for c in row:
                        if c is None:
                            continue
                        for m in re.finditer(r'\b(\d{2,3}\.\d)\b', str(c)):
                            v = float(m.group(1))
                            if domain in ('ad', 'od') and 30.0 <= v <= 100.0:
                                nums.append(v)
                            elif domain == 'cls' and 30.0 <= v <= 100.0:
                                nums.append(v)
                    if nums:
                        cand = max(nums)
                        if best is None or cand > best:
                            best = cand
    return best


def style_from(sh, src, dest, n):
    for c in range(1, n + 1):
        s = sh.cell(row=src, column=c)
        d = sh.cell(row=dest, column=c)
        if s.has_style:
            d.font = copy(s.font); d.fill = copy(s.fill); d.border = copy(s.border)
            d.alignment = copy(s.alignment); d.number_format = s.number_format


def last_row(sh, kc):
    last = sh.max_row
    while last > 1 and sh.cell(row=last, column=kc).value in (None, ''):
        last -= 1
    return last


# ---- per-domain config: xlsx, all-papers sheet, method col, dataset list, ncols ----
DOMAIN_CFG = {
    'ad':  dict(xlsx=AD_XLSX,  allsheet='總覽 All Papers', mcol=3, datasets=AD_DATASETS,  ncols=15),
    'od':  dict(xlsx=OD_XLSX,  allsheet='OD all papers',   mcol=2, datasets=OD_DATASETS,  ncols=13),
    'cls': dict(xlsx=CLS_XLSX, allsheet='Classification all papers', mcol=2, datasets=CLS_DATASETS, ncols=15),
}


def existing_in_sheet(sh, mcol):
    s = set()
    for row in sh.iter_rows(values_only=True):
        if len(row) >= mcol and isinstance(row[mcol-1], str):
            s.add(re.sub(r'[^a-z0-9]', '', row[mcol-1].lower()))
    return s


def write_ad_row(sh, paper, ds, metric):
    nr = last_row(sh, 3) + 1
    vals = [ds, paper['category'], paper['method'], paper['authors'], paper['venue'],
            paper['date'], '預印本', metric, None, None, None, None,
            f'{DRAFT_TAG}：{paper["title"][:80]}', paper['link'], paper['github']]
    for c, v in enumerate(vals, 1):
        sh.cell(row=nr, column=c).value = v
    style_from(sh, max(4, nr-1), nr, 15)


def write_od_row(sh, paper, metric):
    nr = last_row(sh, 2) + 1
    # OD per-dataset sheet: 類別/方法/作者/發表/年月/狀態/mAP/AP/FPS/params/備註/連結/GitHub
    vals = [paper['category'], paper['method'], paper['authors'], paper['venue'],
            paper['date'], 'Preprint', metric, None, None, None,
            f'{DRAFT_TAG}：{paper["title"][:80]}', paper['link'], paper['github']]
    for c, v in enumerate(vals, 1):
        sh.cell(row=nr, column=c).value = v
    style_from(sh, max(2, nr-1), nr, 13)


def write_cls_row(sh, paper, metric):
    nr = last_row(sh, 2) + 1
    vals = [paper['category'], paper['method'], paper['authors'], paper['venue'],
            paper['date'], '預印本', metric, None, None, None, None, None,
            f'{DRAFT_TAG}：{paper["title"][:80]}', paper['link'], paper['github']]
    for c, v in enumerate(vals, 1):
        sh.cell(row=nr, column=c).value = v
    style_from(sh, max(2, nr-1), nr, 15)


def process_domain(domain, limit=None):
    import fitz
    cfg = DOMAIN_CFG[domain]
    wb = load_workbook(cfg['xlsx'])
    allsh = wb[cfg['allsheet']]
    mcol = cfg['mcol']

    # State (resume across runs)
    state = {}
    if STATE.exists():
        try:
            state = json.loads(STATE.read_text(encoding='utf-8'))
        except Exception:
            state = {}
    done = set(state.get(domain, []))

    # Read all-papers rows
    papers = []
    header_row = 1
    for r in range(header_row + 1, allsh.max_row + 1):
        method = allsh.cell(row=r, column=mcol).value
        if not method or not str(method).strip():
            continue
        # column positions differ per domain; read generically
        if domain == 'ad':
            cat = allsh.cell(row=r, column=2).value
            authors = allsh.cell(row=r, column=4).value
            venue = allsh.cell(row=r, column=5).value
            date = allsh.cell(row=r, column=6).value
            notes = allsh.cell(row=r, column=13).value
            link = allsh.cell(row=r, column=14).value
            github = allsh.cell(row=r, column=15).value
        elif domain == 'od':
            cat = allsh.cell(row=r, column=1).value
            authors = allsh.cell(row=r, column=3).value
            venue = allsh.cell(row=r, column=4).value
            date = allsh.cell(row=r, column=5).value
            notes = allsh.cell(row=r, column=6).value
            link = allsh.cell(row=r, column=7).value
            github = allsh.cell(row=r, column=8).value
        else:  # cls
            cat = allsh.cell(row=r, column=1).value
            authors = allsh.cell(row=r, column=3).value
            venue = allsh.cell(row=r, column=4).value
            date = allsh.cell(row=r, column=5).value
            notes = allsh.cell(row=r, column=13).value
            link = allsh.cell(row=r, column=14).value
            github = allsh.cell(row=r, column=15).value
        papers.append(dict(method=str(method).strip(), category=cat or '',
                            authors=authors or '—', venue=venue or '',
                            date=str(date or ''), notes=str(notes or ''),
                            link=link or '', github=github or 'N/A',
                            title=str(notes or method)))

    log(f'[{domain}] {len(papers)} papers in All Papers')
    processed = 0
    for p in papers:
        key = re.sub(r'[^a-z0-9]', '', p['method'].lower())
        if key in done:
            continue
        if VERIFIED_TAG in p['notes']:
            done.add(key); continue
        if limit and processed >= limit:
            break
        processed += 1
        log(f'[{domain}] ({processed}) {p["method"][:55]}')
        data = fetch_pdf(p['link'])
        if not data:
            log('    no PDF — skipped')
            done.add(key); continue
        try:
            doc = fitz.open(stream=data, filetype='pdf')
        except Exception as e:
            log(f'    pdf open failed: {e}')
            done.add(key); continue
        full = ''.join(pg.get_text() for pg in doc)
        ds_found = detect_datasets(full, cfg['datasets'])
        if not ds_found:
            log('    no target dataset detected')
            done.add(key); continue
        # Metrics are intentionally left BLANK for draft rows: a single
        # global number cannot be correctly attributed per-dataset, and a
        # wrong number is worse than none. Draft rows give COVERAGE (which
        # datasets a paper used); verified batches fill the real numbers.
        metric = None
        log(f'    datasets={ds_found} (metric blank — draft)')
        for ds in ds_found:
            if ds not in wb.sheetnames:
                continue
            sh = wb[ds]
            # skip if this method already in that dataset sheet
            ex = existing_in_sheet(sh, 3 if domain == 'ad' else 2)
            if key in ex:
                continue
            if domain == 'ad':
                write_ad_row(sh, p, ds, metric)
            elif domain == 'od':
                write_od_row(sh, p, metric)
            else:
                write_cls_row(sh, p, metric)
        done.add(key)
        # checkpoint every 10 papers
        if processed % 10 == 0:
            wb.save(cfg['xlsx'])
            state[domain] = sorted(done)
            STATE.write_text(json.dumps(state, ensure_ascii=False), encoding='utf-8')
            log(f'    [checkpoint saved @ {processed}]')
        time.sleep(2)

    wb.save(cfg['xlsx'])
    state[domain] = sorted(done)
    STATE.write_text(json.dumps(state, ensure_ascii=False), encoding='utf-8')
    log(f'[{domain}] done — processed {processed} this run')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--domain', choices=['ad', 'od', 'cls'])
    ap.add_argument('--all', action='store_true')
    ap.add_argument('--limit', type=int)
    ap.add_argument('--regen', action='store_true')
    args = ap.parse_args()

    try:
        import fitz  # noqa
    except ImportError:
        print('PyMuPDF required: pip install pymupdf', file=sys.stderr)
        return 1

    log('=' * 60)
    domains = ['ad', 'od', 'cls'] if args.all else ([args.domain] if args.domain else [])
    if not domains:
        print('specify --domain or --all', file=sys.stderr)
        return 1
    for d in domains:
        process_domain(d, limit=args.limit)

    if args.regen:
        sys.path.insert(0, str(PROJECT / 'scripts'))
        import update_papers as up
        ad = up.regenerate_ad_json(); od = up.regenerate_od_json()
        cls = up.regenerate_cls_json(); asd = up.regenerate_as_json()
        up.reinject_html(ad, od, cls, asd)
        log('regenerated JSON + index.html')
    log('all done')
    return 0


if __name__ == '__main__':
    sys.exit(main())
