#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verify_batch7.py — fill PDF-verified metrics + remove misplaced auto-draft rows.

Fills (5):
  - AA-CLIP @ MPDD, BTAD              (Tables 1 & 2, full-shot column)
  - PatchEAD @ MPDD, BTAD             (Table 2, PatchEAD+ DINOv2 zero-shot)
  - InvAD @ MVTec 3D                  (Appendix Table A2, MUAD setting)

Removes (6 — paper does not test the listed dataset):
  - InvAD @ BTAD, MVTec LOCO          (paper only tests MVTec/VisA/3D/Uni-Medical/Real-IAD/COCO-AD)
  - Multi-Flow @ MVTec AD, VisA, LOCO, 3D (paper only evaluates Real-IAD)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

PROJECT = Path(__file__).resolve().parent.parent
XLSX = PROJECT / 'AnomalyDetection_Papers_Summary_v10_20260425.xlsx'

# (sheet, row_1based, i_auroc, p_auroc, p_ap, p_pro, note)
FILLS = [
    ('MPDD', 20, 68.6, 96.6, 71.8, 89.6,
     '✅ 已驗證 (PatchEAD Table 2, PatchEAD+ DINOv2 zero-shot)：Patch-exclusive 訓練免訓練 AD；DINOv2 ViT-B/14 backbone [零樣本 Foundation Model]'),
    ('MPDD', 21, 75.1, 96.7, None, None,
     '✅ 已驗證 (AA-CLIP Tables 1 & 2, full-shot)：Anomaly-Aware CLIP 兩階段 prompt+residual adapter；零樣本/few-shot 通用 [基於 VLM]'),
    ('BTAD', 24, None, None, None, None,
     '⚠️ 移除：InvAD paper 僅在 MVTec AD/VisA/MVTec 3D/Uni-Medical/Real-IAD/COCO-AD 評估，BTAD 僅在 Related Work 提及，並未實際測試'),
    ('BTAD', 28, 91.2, 97.0, 94.3, 80.0,
     '✅ 已驗證 (PatchEAD Table 2, PatchEAD+ DINOv2 zero-shot)：Patch-exclusive 訓練免訓練 AD；BTAD 91.2/97.0 [零樣本 Foundation Model]'),
    ('BTAD', 29, 94.8, 97.0, None, None,
     '✅ 已驗證 (AA-CLIP Tables 1 & 2, full-shot)：Anomaly-Aware CLIP；BTAD 94.8/97.0 (full-shot) [基於 VLM]'),
    ('MVTec 3D', 20, 86.1, 98.8, 37.8, 94.7,
     '✅ 已驗證 (InvAD Appendix Table A2, MUAD)：Feature Inversion 多類別異常偵測；MVTec 3D AD I=86.1, P=98.8 [基於重構]'),
]

# Rows to DELETE (misplaced — paper does NOT actually evaluate on this dataset)
DELETES = [
    ('BTAD', 24, 'InvAD paper does not test BTAD'),                    # InvAD
    ('MVTec 3D', 23, 'Multi-Flow paper only evaluates Real-IAD'),      # Multi-Flow
    ('MVTec LOCO', 21, 'InvAD paper does not test MVTec LOCO'),
    ('MVTec LOCO', 25, 'Multi-Flow paper only evaluates Real-IAD'),
    ('MVTec AD', 97, 'Multi-Flow paper only evaluates Real-IAD'),
    ('VisA', 73, 'Multi-Flow paper only evaluates Real-IAD'),
]

# Drop the BTAD r24 fill (it's a deletion, not a fill)
FILLS = [f for f in FILLS if not (f[0] == 'BTAD' and f[1] == 24)]


def main():
    wb = load_workbook(XLSX)
    summary = []

    for sheet, row, i, p, ap, pro, note in FILLS:
        s = wb[sheet]
        cur = s.cell(row=row, column=3).value
        s.cell(row=row, column=7).value = 'N/A' if i is None else i
        s.cell(row=row, column=8).value = 'N/A' if p is None else p
        s.cell(row=row, column=9).value = 'N/A' if ap is None else ap
        s.cell(row=row, column=10).value = 'N/A' if pro is None else pro
        s.cell(row=row, column=13).value = note
        summary.append(f'  FILL [{sheet}] r{row}: {cur} → i={i}, p={p}, ap={ap}, pro={pro}')

    by_sheet = defaultdict(list)
    for sheet, row, reason in DELETES:
        by_sheet[sheet].append((row, reason))
    for sheet, items in by_sheet.items():
        s = wb[sheet]
        for row, reason in sorted(items, key=lambda x: -x[0]):
            cur = s.cell(row=row, column=3).value
            s.delete_rows(row, 1)
            summary.append(f'  DEL  [{sheet}] r{row}: {cur} ({reason})')

    wb.save(XLSX)
    print('Saved', XLSX)
    print('\n'.join(summary))


if __name__ == '__main__':
    main()
